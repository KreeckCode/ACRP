import json
import secrets
import logging
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from urllib import request
import zipfile
from .card_delivery import *
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib import messages
from django.contrib.contenttypes.models import ContentType
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.views.decorators.cache import cache_page, never_cache
from django.http import (
    HttpResponse, JsonResponse, Http404, HttpResponseForbidden, 
    HttpResponseBadRequest, FileResponse
)
from django.urls import reverse, reverse_lazy
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.mail import send_mail, EmailMessage
from django.core.cache import cache
from django.core.exceptions import ValidationError, PermissionDenied
from django.db import transaction, IntegrityError
from django.db.models import Q, Count, Avg, Sum, F
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.conf import settings
from django.template.loader import render_to_string

# Image processing
from PIL import Image, ImageDraw, ImageFont
import qrcode

# PDF generation
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet

from .models import (
    AffiliationCard, CardTemplate, CardVerification, CardDelivery,
    CardStatusChange, CardSystemSettings
)
from .forms import (
    CardAssignmentForm, CardPhotoUploadForm, CardStatusUpdateForm,
    BulkCardOperationForm, CardLookupForm, QRVerificationForm,
    CardDeliveryForm, CardDownloadForm, CardTemplateForm,
    SystemSettingsForm, CardReportForm, ContactForm
)
from mailjet_rest import Client
logger = logging.getLogger(__name__)
from .card_delivery import *

# ============================================================================
# UTILITY FUNCTIONS AND DECORATORS
# ============================================================================
def bulk_send_emails(cards, user, email_type='general', custom_subject=None, custom_message=None, request=None):
    """
    Bulk send emails to cardholders.
    
    Args:
        cards: QuerySet of AffiliationCard objects
        user: User performing the operation
        email_type: Type of email to send ('renewal_reminder', 'update', 'reactivation', 'welcome', 'custom')
        custom_subject: Custom email subject (for custom email type)
        custom_message: Custom email message (for custom email type)
    """
    success_count = 0
    failed_count = 0
    failed_emails = []
    
    for card in cards:
        try:
            # Skip cards without email addresses
            if not card.affiliate_email:
                failed_count += 1
                failed_emails.append(f"{card.card_number} - No email address")
                continue
            
            # Determine email content based on type
            if email_type == 'renewal_reminder':
                subject = f"ACRP Card Renewal Reminder - {card.card_number}"
                template = 'affiliationcard/emails/renewal_reminder.html'
                
            elif email_type == 'update':
                subject = f"ACRP Card System Update - {card.card_number}"
                template = 'affiliationcard/emails/system_update.html'
                
            elif email_type == 'reactivation':
                subject = f"Reactivate Your ACRP Card - {card.card_number}"
                template = 'affiliationcard/emails/reactivation.html'
                
            elif email_type == 'welcome':
                subject = f"Welcome to ACRP Digital Cards - {card.card_number}"
                template = 'affiliationcard/emails/welcome.html'
                
            elif email_type == 'custom':
                subject = custom_subject or f"ACRP Card Communication - {card.card_number}"
                template = 'affiliationcard/emails/custom.html'
                
            else:
                subject = f"ACRP Card Information - {card.card_number}"
                template = 'affiliationcard/emails/general.html'
            
            # Prepare email context
            context = {
                'card': card,
                'user': user,
                'email_type': email_type,
                'custom_message': custom_message,
                'current_year': timezone.now().year,
                'verification_url': request.build_absolute_uri(
                    reverse('affiliationcard:verify_token', args=[card.verification_token])
                ) if request else '#',
                'card_detail_url': request.build_absolute_uri(
                    reverse('affiliationcard:download_card', args=[card.verification_token])
                ) if request else '#',
            }
            
            # Render email content
            message = render_to_string(template, context)
            
            # Send email
            email = EmailMessage(
                subject=subject,
                body=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[card.affiliate_email]
            )
            email.content_subtype = "html"
            email.send()
            
            success_count += 1
            
            # Log the email sending
            logger.info(f"Bulk email sent to {card.affiliate_email} for card {card.card_number} by {user.username}")
            
        except Exception as e:
            failed_count += 1
            failed_emails.append(f"{card.card_number} - {str(e)}")
            logger.error(f"Failed to send bulk email for card {card.card_number}: {e}")
    
    # Prepare result message
    message = f'{success_count} emails sent successfully'
    if failed_count > 0:
        message += f', {failed_count} failed'
    
    return {
        'message': message,
        'success_count': success_count,
        'failed_count': failed_count,
        'failed_emails': failed_emails
    }


def bulk_regenerate_cards(cards, user, reason='bulk_regeneration', regenerate_tokens=True, regenerate_numbers=False, update_template=None):
    """
    Bulk regenerate card data (tokens, numbers, templates).
    
    Args:
        cards: QuerySet of AffiliationCard objects
        user: User performing the operation
        reason: Reason for regeneration
        regenerate_tokens: Whether to regenerate verification tokens
        regenerate_numbers: Whether to regenerate card numbers
        update_template: New template to apply (CardTemplate instance)
    """
    success_count = 0
    failed_count = 0
    failed_cards = []
    
    with transaction.atomic():
        for card in cards:
            try:
                old_data = {
                    'card_number': card.card_number,
                    'verification_token': card.verification_token,
                    'template': card.card_template.name if card.card_template else None
                }
                
                # Regenerate verification token
                if regenerate_tokens:
                    card.verification_token = secrets.token_urlsafe(32)
                
                # Regenerate card number (be careful with this!)
                if regenerate_numbers:
                    # Generate new card number using the same pattern as original
                    new_card_number = card.generate_card_number()
                    
                    # Ensure uniqueness
                    while AffiliationCard.objects.filter(card_number=new_card_number).exists():
                        new_card_number = card.generate_card_number()
                    
                    card.card_number = new_card_number
                
                # Update template if provided
                if update_template:
                    card.card_template = update_template
                
                # Update timestamps
                card.updated_at = timezone.now()
                card.save()
                
                # Log the regeneration
                CardStatusChange.objects.create(
                    card=card,
                    old_status=card.status,
                    new_status=card.status,  # Status doesn't change
                    reason=f"Bulk regeneration: {reason}",
                    changed_by=user,
                    ip_address='127.0.0.1',  # System operation
                    user_agent='System/BulkRegeneration',
                    additional_data=json.dumps({
                        'operation': 'bulk_regeneration',
                        'regenerated_tokens': regenerate_tokens,
                        'regenerated_numbers': regenerate_numbers,
                        'old_card_number': old_data['card_number'],
                        'old_verification_token': old_data['verification_token'][:10] + '...',  # Partial for security
                        'new_template': update_template.name if update_template else None
                    })
                )
                
                success_count += 1
                
                logger.info(f"Card {card.card_number} regenerated by {user.username}: {reason}")
                
            except Exception as e:
                failed_count += 1
                failed_cards.append(f"{card.card_number} - {str(e)}")
                logger.error(f"Failed to regenerate card {card.card_number}: {e}")
    
    # Prepare result message
    message = f'{success_count} cards regenerated successfully'
    if failed_count > 0:
        message += f', {failed_count} failed'
    
    return {
        'message': message,
        'success_count': success_count,
        'failed_count': failed_count,
        'failed_cards': failed_cards
    }


def bulk_send_emails_with_attachments(cards, user, email_type='card_delivery', attachment_format='pdf'):
    """
    Bulk send emails with card attachments.
    
    This is a specialized version for sending actual card files.
    """
    success_count = 0
    failed_count = 0
    failed_emails = []
    
    for card in cards:
        try:
            if not card.affiliate_email:
                failed_count += 1
                failed_emails.append(f"{card.card_number} - No email address")
                continue
            
            # Generate card file
            file_content, filename, content_type = generate_card_file(card, attachment_format)
            
            # Prepare email
            subject = f"Your ACRP Digital Card - {card.card_number}"
            context = {
                'card': card,
                'user': user
            }
            
            message = render_to_string('affiliationcard/emails/card_delivery.html', context)
            
            # Create email with attachment
            email = EmailMessage(
                subject=subject,
                body=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[card.affiliate_email]
            )
            email.content_subtype = "html"
            email.attach(filename, file_content, content_type)
            
            # Send email
            email.send()
            
            success_count += 1
            
            # Create delivery record
            CardDelivery.objects.create(
                card=card,
                delivery_type='email_pdf',
                recipient_email=card.affiliate_email,
                recipient_name=card.get_display_name(),
                initiated_by=user,
                file_format=attachment_format,
                status='completed',
                completed_at=timezone.now()
            )
            
            logger.info(f"Card attachment sent to {card.affiliate_email} for card {card.card_number}")
            
        except Exception as e:
            failed_count += 1
            failed_emails.append(f"{card.card_number} - {str(e)}")
            logger.error(f"Failed to send card attachment for {card.card_number}: {e}")
    
    message = f'{success_count} card emails sent successfully'
    if failed_count > 0:
        message += f', {failed_count} failed'
    
    return {
        'message': message,
        'success_count': success_count,
        'failed_count': failed_count,
        'failed_emails': failed_emails
    }


# Helper function to add to your existing bulk_operations view
def handle_bulk_send_emails(request, cards, form_data):
    """
    Handle bulk email sending with different options.
    """
    email_type = form_data.get('email_type', 'general')
    custom_subject = form_data.get('custom_subject')
    custom_message = form_data.get('custom_message')
    include_attachment = form_data.get('include_attachment', False)
    attachment_format = form_data.get('attachment_format', 'pdf')
    
    if include_attachment:
        return bulk_send_emails_with_attachments(cards, request.user, email_type, attachment_format)
    else:
        return bulk_send_emails(cards, request.user, email_type, custom_subject, custom_message)


def handle_bulk_regenerate_cards(request, cards, form_data):
    """
    Handle bulk card regeneration with different options.
    """
    reason = form_data.get('reason', 'Bulk regeneration')
    regenerate_tokens = form_data.get('regenerate_tokens', True)
    regenerate_numbers = form_data.get('regenerate_numbers', False)
    template_id = form_data.get('new_template_id')
    
    update_template = None
    if template_id:
        try:
            update_template = CardTemplate.objects.get(id=template_id)
        except CardTemplate.DoesNotExist:
            pass
    
    return bulk_regenerate_cards(
        cards, request.user, reason, 
        regenerate_tokens, regenerate_numbers, update_template
    )

def is_card_admin(user):
    """Check if user has card administration privileges."""
    return user.is_staff or user.acrp_role in {
        'GLOBAL_SDP', 'PROVIDER_ADMIN'
    }


def get_client_ip(request):
    """Extract client IP address safely."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def rate_limit_verification(request, max_attempts=10, window_minutes=60):
    """Simple rate limiting for verification attempts."""
    ip = get_client_ip(request)
    cache_key = f"card_verification_attempts:{ip}"
    
    attempts = cache.get(cache_key, 0)
    if attempts >= max_attempts:
        return False
    
    cache.set(cache_key, attempts + 1, window_minutes * 60)
    return True


# ============================================================================
# ADMIN CARD MANAGEMENT VIEWS
# ============================================================================

@login_required
@user_passes_test(is_card_admin)
def send_card(request, pk):
    """
    Enhanced card delivery configuration and sending.
    
    This view handles the complete card delivery workflow including:
    - Form validation and processing
    - Multiple delivery methods (email PDF, email link, direct download)
    - Error handling and user feedback
    - Delivery tracking and logging
    """
    card = get_object_or_404(AffiliationCard, pk=pk)
    
    # Check if card is in a valid state for delivery
    if card.status not in ['assigned', 'active']:
        messages.warning(request, f"Card status '{card.get_status_display()}' may not be suitable for delivery.")
    
    if request.method == 'POST':
        form = CardDeliveryForm(request.POST, card=card)
        
        if form.is_valid():
            try:
                logger.info(f"DEBUG: About to create delivery for card {card.card_number}")
                logger.info(f"DEBUG: Delivery method: {form.cleaned_data['delivery_method']}")
                logger.info(f"DEBUG: File format: {form.cleaned_data['file_format']}")
                

                # Extract form data
                delivery_data = {
                    'delivery_method': form.cleaned_data['delivery_method'],
                    'recipient_email': form.cleaned_data['recipient_email'],
                    'recipient_name': form.cleaned_data['recipient_name'],
                    'email_subject': form.cleaned_data.get('email_subject', ''),
                    'email_message': form.cleaned_data.get('email_message', ''),
                    'file_format': form.cleaned_data['file_format'],
                    'initiated_by': request.user,
                    'request': request,  # Pass request for URL building
                    'max_downloads': form.cleaned_data.get('max_downloads', 5)
                }
                
                # Right after extracting delivery_data, add:
                logger.info(f"DEBUG: About to call create_card_delivery function")
                logger.info(f"DEBUG: Function exists: {callable(create_card_delivery)}")

                try:
                    # Create and process delivery
                    delivery = create_card_delivery(card=card, **delivery_data)
                    logger.info(f"DEBUG: create_card_delivery returned successfully")
                except Exception as e:
                    logger.error(f"DEBUG: create_card_delivery failed with exception: {e}", exc_info=True)
                    raise

                logger.info(f"DEBUG: create_card_delivery returned: {delivery}")
                logger.info(f"DEBUG: Delivery status: {delivery.status}")
                logger.info(f"DEBUG: Delivery failure reason: {delivery.failure_reason}")
                
                
                # Provide user feedback based on delivery method
                if delivery.status == 'completed':
                    if delivery_data['delivery_method'] == 'email_pdf':
                        messages.success(
                            request, 
                            f"Card sent successfully as PDF attachment to {delivery.recipient_email}! "
                            f"Delivery ID: {delivery.id}"
                        )
                    elif delivery_data['delivery_method'] == 'email_link':
                        messages.success(
                            request,
                            f"Download link sent successfully to {delivery.recipient_email}! "
                            f"Link expires on {delivery.download_expires_at.strftime('%B %d, %Y')}. "
                            f"Delivery ID: {delivery.id}"
                        )
                    else:  # direct_download
                        messages.success(
                            request,
                            f"Card generated successfully for direct download! "
                            f"Delivery ID: {delivery.id}"
                        )
                elif delivery.status == 'ready_for_download':
                    messages.info(
                        request,
                        f"Card prepared for download. Delivery ID: {delivery.id}"
                    )
                else:
                    messages.warning(
                        request,
                        f"Card delivery initiated but status is '{delivery.get_status_display()}'. "
                        f"Please check the delivery details. Delivery ID: {delivery.id}"
                    )
                
                # Log successful delivery initiation
                logger.info(
                    f"Card delivery initiated by {request.user.username} for card {card.card_number} "
                    f"(delivery method: {delivery_data['delivery_method']}, "
                    f"recipient: {delivery.recipient_email}, "
                    f"delivery_id: {delivery.id})"
                )
                
                return redirect('affiliationcard:card_detail', pk=pk)
                
            except Exception as e:
                # Log the detailed error
                logger.error(
                    f"Card delivery failed for card {card.card_number} by {request.user.username}: {str(e)}", 
                    exc_info=True
                )
                
                # Provide user-friendly error message
                if 'Mailjet' in str(e):
                    messages.error(
                        request, 
                        "Failed to send email. Please check the email configuration and try again."
                    )
                elif 'generate' in str(e).lower():
                    messages.error(
                        request,
                        "Failed to generate card file. Please try a different format or contact support."
                    )
                else:
                    messages.error(
                        request, 
                        f"Failed to initiate card delivery: {str(e)}"
                    )
        else:
            # Form validation errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        # GET request - initialize form with card defaults
        initial_data = {
            'recipient_email': card.affiliate_email,
            'recipient_name': card.get_display_name(),
            'file_format': 'pdf',
            'delivery_method': 'email_pdf'
        }
        form = CardDeliveryForm(card=card, initial=initial_data)
    
    # Get recent deliveries for display
    recent_deliveries = card.deliveries.select_related('initiated_by').order_by('-initiated_at')[:5]
    
    context = {
        'form': form,
        'card': card,
        'recent_deliveries': recent_deliveries,
        'page_title': f'Send Card {card.card_number}',
        'delivery_stats': {
            'total_deliveries': card.deliveries.count(),
            'successful_deliveries': card.deliveries.filter(status='completed').count(),
            'failed_deliveries': card.deliveries.filter(status='failed').count(),
        }
    }
    
    return render(request, 'affiliationcard/admin/send_card.html', context)


@login_required
@user_passes_test(is_card_admin)
def card_dashboard(request):
    """
    Enhanced main dashboard for card administration.
    
    Shows overview statistics, recent activity, quick actions, and alerts
    for applications needing card assignment.
    """
    from enrollments.models import AssociatedApplication, DesignatedApplication, StudentApplication
    
    # Get summary statistics
    stats = {
        'total_cards': AffiliationCard.objects.count(),
        'active_cards': AffiliationCard.objects.filter(status='active').count(),
        'pending_cards': AffiliationCard.objects.filter(status='pending_assignment').count(),
        'expired_cards': AffiliationCard.objects.filter(status='expired').count(),
        'assigned_cards': AffiliationCard.objects.filter(status='assigned').count(),
        'suspended_cards': AffiliationCard.objects.filter(status='suspended').count(),
        'total_verifications': CardVerification.objects.count(),
        'recent_verifications': CardVerification.objects.filter(
            verified_at__gte=timezone.now() - timedelta(days=7)
        ).count(),
        'monthly_verifications': CardVerification.objects.filter(
            verified_at__gte=timezone.now() - timedelta(days=30)
        ).count(),
    }
    
    # Enhanced statistics with percentages and trends
    if stats['total_cards'] > 0:
        stats['active_percentage'] = round((stats['active_cards'] / stats['total_cards']) * 100, 1)
        stats['pending_percentage'] = round((stats['pending_cards'] / stats['total_cards']) * 100, 1)
    else:
        stats['active_percentage'] = 0
        stats['pending_percentage'] = 0
    
    # Get applications that are approved but don't have cards assigned
    approved_applications_without_cards = []
    
    # Check each application type more efficiently
    for model in [AssociatedApplication, DesignatedApplication, StudentApplication]:
        try:
            content_type = ContentType.objects.get_for_model(model)
            
            # Get approved applications that don't have cards
            approved_apps = model.objects.filter(
                status='approved'
            ).exclude(
                id__in=AffiliationCard.objects.filter(
                    content_type=content_type
                ).values_list('object_id', flat=True)
            ).order_by('-approved_at')[:10]  # Limit to most recent 10
            
            for app in approved_apps:
                app_type = model.__name__.replace('Application', '').lower()
                approved_applications_without_cards.append({
                    'application': app,
                    'app_type': app_type,
                    'content_type_id': content_type.id,
                    'object_id': app.id,
                    'days_since_approval': (timezone.now().date() - app.approved_at.date()).days if app.approved_at else 0
                })
        except Exception as e:
            logger.error(f"Error processing {model.__name__}: {e}")
            continue
    
    # Sort by approval date (most urgent first)
    approved_applications_without_cards.sort(key=lambda x: x['days_since_approval'], reverse=True)
    
    # Council breakdown with enhanced metrics
    council_stats = AffiliationCard.objects.values('council_code', 'council_name').annotate(
        total=Count('id'),
        active=Count('id', filter=Q(status='active')),
        pending=Count('id', filter=Q(status='pending_assignment')),
        expired=Count('id', filter=Q(status='expired')),
        recent_verifications=Count(
            'verifications',
            filter=Q(verifications__verified_at__gte=timezone.now() - timedelta(days=30))
        )
    ).order_by('-total')  # Order by total cards descending
    
    # Recent cards with enhanced data
    recent_cards = AffiliationCard.objects.select_related(
        'card_template', 'assigned_by'
    ).prefetch_related(
        'verifications'
    ).order_by('-created_at')[:10]
    
    # Recent verifications with enhanced data
    recent_verifications = CardVerification.objects.select_related(
        'card'
    ).order_by('-verified_at')[:15]
    
    # Cards expiring soon (next 30 days) with urgency classification
    expiring_soon = AffiliationCard.objects.filter(
        status='active',
        date_expires__lte=timezone.now().date() + timedelta(days=30),
        date_expires__gte=timezone.now().date()
    ).order_by('date_expires')[:15]
    
    # Add urgency classification to expiring cards
    for card in expiring_soon:
        days_until_expiry = (card.date_expires - timezone.now().date()).days
        if days_until_expiry <= 7:
            card.urgency = 'critical'
        elif days_until_expiry <= 14:
            card.urgency = 'high'
        else:
            card.urgency = 'medium'
    
    # Performance metrics for dashboard
    performance_metrics = {
        'avg_verifications_per_card': round(
            stats['total_verifications'] / max(stats['total_cards'], 1), 2
        ),
        'cards_needing_attention': stats['pending_cards'] + stats['expired_cards'],
        'system_health_score': calculate_system_health_score(stats),
        'approval_backlog_count': len(approved_applications_without_cards),
    }
    
    context = {
        'stats': stats,
        'performance_metrics': performance_metrics,
        'council_stats': council_stats,
        'recent_cards': recent_cards,
        'recent_verifications': recent_verifications,
        'expiring_soon': expiring_soon,
        'approved_applications_without_cards': approved_applications_without_cards[:5],  # Show top 5
        'total_approval_backlog': len(approved_applications_without_cards),
        'page_title': 'Card Administration Dashboard'
    }
    
    return render(request, 'affiliationcard/admin/dashboard.html', context)


def calculate_system_health_score(stats):
    """
    Calculate a system health score based on various metrics.
    
    Args:
        stats: Dictionary containing card statistics
    
    Returns:
        int: Health score from 0-100
    """
    try:
        score = 100
        total_cards = stats.get('total_cards', 0)
        
        if total_cards == 0:
            return 50  # Neutral score for empty system
        
        # Deduct points for problematic cards
        expired_ratio = stats.get('expired_cards', 0) / total_cards
        pending_ratio = stats.get('pending_cards', 0) / total_cards
        suspended_ratio = stats.get('suspended_cards', 0) / total_cards
        
        score -= int(expired_ratio * 30)  # Max 30 points deduction for expired cards
        score -= int(pending_ratio * 20)   # Max 20 points deduction for pending cards
        score -= int(suspended_ratio * 25) # Max 25 points deduction for suspended cards
        
        # Bonus points for high verification activity
        recent_verifications = stats.get('recent_verifications', 0)
        if recent_verifications > total_cards * 0.5:  # More than 50% verification rate
            score += 5
        
        return max(0, min(100, score))  # Ensure score is between 0-100
    except Exception as e:
        logger.error(f"Error calculating system health score: {e}")
        return 75  # Default safe score
    




@login_required
@user_passes_test(is_card_admin)
@require_http_methods(["GET", "POST"])
def assign_card(request, content_type_id, object_id):
    """
    Assign digital card to approved application.
    
    This is the main view called when admin clicks "Assign Digital Card"
    during application review in the enrollments app.
    """
    # Get the application
    try:
        content_type = ContentType.objects.get(pk=content_type_id)
        application = content_type.get_object_for_this_type(pk=object_id)
    except (ContentType.DoesNotExist, application.DoesNotExist):
        messages.error(request, "Application not found.")
        return redirect('enrollments:application_list')
    
    # Validate application is approved
    if not hasattr(application, 'status') or application.status != 'approved':
        messages.error(request, "Card can only be assigned to approved applications.")
        return redirect('enrollments:application_detail', pk=object_id, app_type=application.get_affiliation_type())
    
    # Check if card already exists
    existing_card = AffiliationCard.objects.filter(
        content_type=content_type,
        object_id=object_id,
        status__in=['assigned', 'active']
    ).first()
    
    if existing_card:
        messages.warning(request, f"Card already exists: {existing_card.card_number}")
        return redirect('affiliationcard:card_detail', pk=existing_card.pk)
    
    # Get council for template filtering
    council = application.get_council()
    
    if request.method == 'POST':
        form = CardAssignmentForm(request.POST, council=council, application=application)
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Create the card
                    card = AffiliationCard.objects.create(
                        content_type=content_type,
                        object_id=object_id,
                        card_template=form.cleaned_data.get('card_template'),
                        grace_period_days=form.cleaned_data['grace_period_days'],
                        assigned_by=request.user,
                        assigned_at=timezone.now(),
                        status='assigned'
                    )
                    
                    # Issue immediately if requested
                    if form.cleaned_data['issue_immediately']:
                        card.issue_card(issued_by=request.user)
                        messages.success(request, f"Card {card.card_number} assigned and issued successfully!")
                    else:
                        messages.success(request, f"Card {card.card_number} assigned successfully!")
                    
                    # Send email notification if requested
                    if form.cleaned_data['send_email_notification']:
                        try:
                            send_card_assignment_notification(card)
                            messages.info(request, "Email notification sent to affiliate.")
                        except Exception as e:
                            logger.error(f"Failed to send card notification: {e}")
                            messages.warning(request, "Card created but email notification failed.")
                    
                    return redirect('affiliationcard:card_detail', pk=card.pk)
                    
            except Exception as e:
                logger.error(f"Error assigning card: {e}")
                messages.error(request, "An error occurred while assigning the card.")
    else:
        form = CardAssignmentForm(council=council, application=application)
    
    context = {
        'form': form,
        'application': application,
        'council': council,
        'page_title': f'Assign Card - {application.get_display_name()}'
    }
    
    return render(request, 'affiliationcard/admin/assign_card.html', context)

def assign_card_programmatically(content_type, object_id, assigned_by, send_email=True, request=None):
    """
    Programmatically assign a digital card to an approved application.
    """
    try:
        # Get the application object
        application = content_type.get_object_for_this_type(pk=object_id)
        
        # Validate application is approved
        if not hasattr(application, 'status') or application.status != 'approved':
            raise ValueError("Card can only be assigned to approved applications")
        
        # Check if card already exists to prevent duplicates
        existing_card = AffiliationCard.objects.filter(
            content_type=content_type,
            object_id=object_id,
            status__in=['assigned', 'active']
        ).first()
        
        if existing_card:
            raise ValueError(f"Card already exists: {existing_card.card_number}")
        
        # Get council for template selection
        council = application.get_council()
        
        # Get default template for this council
        default_template = CardTemplate.objects.filter(
            council_code=council.code,
            is_default=True
        ).first()
        
        if not default_template:
            # Fallback to any template for this council
            default_template = CardTemplate.objects.filter(council_code=council.code).first()
        
        # Create the card with transaction safety
        with transaction.atomic():
            card = AffiliationCard.objects.create(
                content_type=content_type,
                object_id=object_id,
                card_template=default_template,
                grace_period_days=30,  # Default grace period
                assigned_by=assigned_by,
                assigned_at=timezone.now(),
                status='assigned'
            )
            
            # Issue the card immediately since this is an approved application
            card.issue_card(issued_by=assigned_by)
            
            # Send email notification if requested
            if send_email and card.affiliate_email:
                try:
                    # Pass the request object if available
                    send_digital_card_email(card, request)
                    
                    # Create delivery record
                    CardDelivery.objects.create(
                        card=card,
                        delivery_type='email_link',
                        recipient_email=card.affiliate_email,
                        recipient_name=card.get_display_name(),
                        initiated_by=assigned_by,
                        file_format='pdf',
                        status='completed',
                        completed_at=timezone.now()
                    )
                except Exception as e:
                    logger.error(f"Failed to send digital card email: {str(e)}")
                    # Don't re-raise the exception, just log it
                    # The card should still be created even if email fails
            
            logger.info(f"Digital card {card.card_number} programmatically assigned to application {application.application_number}")
            return card
            
    except Exception as e:
        logger.error(f"Failed to assign card programmatically: {str(e)}")
        raise



    
def send_digital_card_email(card, request=None):
    """
    Send digital card via email to the affiliate.
    
    This function sends the actual digital card file as a PDF attachment
    along with a welcome message and download link.
    
    Args:
        card: AffiliationCard instance
        request: HTTP request object for building absolute URLs
    """
    try:
        # Generate the card file
        file_content, filename, content_type_pdf = generate_card_file(card, 'pdf')
        
        # Prepare email context
        context = {
            'card': card,
            'applicant_name': card.get_display_name(),
            'card_number': card.card_number,
            'council_name': getattr(card, 'council_name', 'N/A'),
            'affiliation_type': getattr(card, 'affiliation_type', 'N/A').title() if hasattr(card, 'affiliation_type') else 'N/A',
            'verification_url': request.build_absolute_uri(
                reverse('affiliationcard:verify_token', args=[card.verification_token])
            ) if request else '#',
            'download_url': request.build_absolute_uri(
                reverse('affiliationcard:download_card', args=[card.verification_token])
            ) if request else '#',
            'current_year': timezone.now().year,
        }
        
        # Render email content - using Tailwind-compatible template
        subject = f"Your ACRP Digital Affiliation Card - {card.card_number}"
        html_message = render_to_string('affiliationcard/emails/digital_card_delivery.html', context)
        
        # Create email with both text and HTML versions
        text_message = f"""
        Your ACRP Digital Affiliation Card - {card.card_number}
        
        Dear {card.get_display_name()},
        
        Your digital affiliation card is ready. You can download it from:
        {request.build_absolute_uri(reverse('affiliationcard:download_card', args=[card.verification_token])) if request else '#'}
        
        Or verify your card at:
        {request.build_absolute_uri(reverse('affiliationcard:verify_token', args=[card.verification_token])) if request else '#'}
        
        Thank you,
        ACRP Team
        """
        
        # Create email
        email = EmailMessage(
            subject=subject,
            body='',
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[card.affiliate_email],
            reply_to=[settings.DEFAULT_REPLY_TO_EMAIL] if hasattr(settings, 'DEFAULT_REPLY_TO_EMAIL') else [settings.DEFAULT_FROM_EMAIL],
        )
        
        # Attach HTML alternative
        email.attach_alternative(html_message, "text/html")
        
        # Attach the PDF card
        email.attach(filename, file_content, content_type_pdf)
        
        # Send the email
        email.send()
        
        logger.info(f"Digital card email sent to {card.affiliate_email} for card {card.card_number}")
        
    except Exception as e:
        logger.error(f"Failed to send digital card email for card {card.card_number}: {str(e)}")




    
class CardListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """List view for all affiliation cards with filtering and search."""
    
    model = AffiliationCard
    template_name = 'affiliationcard/admin/card_list.html'
    context_object_name = 'cards'
    paginate_by = 25
    permission_required = 'affiliationcard.view_affiliationcard'
    
    def get_queryset(self):
        """Get filtered queryset based on search parameters."""
        queryset = AffiliationCard.objects.select_related().order_by('-created_at')
        
        # Apply filters
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(card_number__icontains=search) |
                Q(affiliate_full_name__icontains=search) |
                Q(affiliate_surname__icontains=search) |
                Q(affiliate_email__icontains=search)
            )
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        council = self.request.GET.get('council')
        if council:
            queryset = queryset.filter(council_code=council)
        
        affiliation_type = self.request.GET.get('affiliation_type')
        if affiliation_type:
            queryset = queryset.filter(affiliation_type=affiliation_type)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add filter options to context."""
        context = super().get_context_data(**kwargs)
        
        # Add filter choices
        context.update({
            'status_choices': AffiliationCard.STATUS_CHOICES,
            'council_choices': AffiliationCard.objects.values_list('council_code', 'council_name').distinct(),
            'affiliation_type_choices': AffiliationCard.objects.values_list('affiliation_type', 'affiliation_type').distinct(),
            'search_query': self.request.GET.get('search', ''),
            'selected_status': self.request.GET.get('status', ''),
            'selected_council': self.request.GET.get('council', ''),
            'selected_affiliation_type': self.request.GET.get('affiliation_type', ''),
            'page_title': 'Affiliation Cards'
        })
        
        return context


class CardDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """Detailed view of a single affiliation card."""
    
    model = AffiliationCard
    template_name = 'affiliationcard/admin/card_detail.html'
    context_object_name = 'card'
    permission_required = 'affiliationcard.view_affiliationcard'
    
    def get_object(self):
        """Get card with related data."""
        return get_object_or_404(
            AffiliationCard.objects.select_related('card_template').prefetch_related(
                'verifications', 'deliveries', 'status_changes'
            ),
            pk=self.kwargs['pk']
        )
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        card = self.object
        
        # Recent activity
        recent_verifications = card.verifications.order_by('-verified_at')[:10]
        recent_deliveries = card.deliveries.order_by('-initiated_at')[:5]
        status_history = card.status_changes.order_by('-changed_at')[:10]
        
        # Statistics
        verification_stats = {
            'total': card.total_verifications,
            'last_30_days': card.verifications.filter(
                verified_at__gte=timezone.now() - timedelta(days=30)
            ).count(),
            'last_7_days': card.verifications.filter(
                verified_at__gte=timezone.now() - timedelta(days=7)
            ).count(),
        }
        
        context.update({
            'recent_verifications': recent_verifications,
            'recent_deliveries': recent_deliveries,
            'status_history': status_history,
            'verification_stats': verification_stats,
            'page_title': f'Card {card.card_number}'
        })
        
        return context


@login_required
@user_passes_test(is_card_admin)
@require_POST
def update_card_status(request, pk):
    """Update card status (activate, suspend, revoke, etc.)."""
    card = get_object_or_404(AffiliationCard, pk=pk)
    form = CardStatusUpdateForm(request.POST, card=card)
    
    if form.is_valid():
        try:
            old_status = card.status
            new_status = form.cleaned_data['new_status']
            reason = form.cleaned_data['reason']
            
            with transaction.atomic():
                # Update card status
                card.status = new_status
                card.save()
                
                # Log status change
                CardStatusChange.objects.create(
                    card=card,
                    old_status=old_status,
                    new_status=new_status,
                    reason=reason,
                    changed_by=request.user,
                    ip_address=get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )
                
                # Send notification if requested
                if form.cleaned_data['notify_affiliate']:
                    try:
                        send_status_change_notification(card, old_status, new_status, reason)
                    except Exception as e:
                        logger.error(f"Failed to send status change notification: {e}")
                        messages.warning(request, "Status updated but notification failed.")
                
                messages.success(request, f"Card status updated to {card.get_status_display()}")
                
        except Exception as e:
            logger.error(f"Error updating card status: {e}")
            messages.error(request, "Failed to update card status.")
    else:
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"{field}: {error}")
    
    return redirect('affiliationcard:card_detail', pk=pk)


@login_required
@user_passes_test(is_card_admin)
def bulk_operations(request):
    """Handle bulk operations on multiple cards."""
    if request.method == 'POST':
        form = BulkCardOperationForm(request.POST)
        
        if form.is_valid():
            operation = form.cleaned_data['operation']
            card_ids = form.cleaned_data['card_ids']
            reason = form.cleaned_data.get('reason', '')
            send_notifications = form.cleaned_data['send_notifications']
            
            cards = AffiliationCard.objects.filter(id__in=card_ids)
            
            try:
                with transaction.atomic():
                    if operation == 'assign':
                        result = bulk_assign_cards(cards, request.user, reason, request=request)
                    elif operation == 'issue':
                        result = bulk_issue_cards(cards, request.user, reason, request=request)
                    elif operation == 'suspend':
                        result = bulk_suspend_cards(cards, request.user, reason, request=request)
                    elif operation == 'send_email':
                        result = bulk_send_emails(cards, request.user, request=request)
                    elif operation == 'regenerate':
                        result = bulk_regenerate_cards(cards, request.user, request=request)
                    else:
                        messages.error(request, "Invalid operation")
                        return redirect('affiliationcard:card_list')
                    
                    messages.success(request, f"Bulk operation completed: {result['message']}")
                    
            except Exception as e:
                logger.error(f"Bulk operation failed: {e}")
                messages.error(request, "Bulk operation failed.")
        else:
            messages.error(request, "Invalid form data.")
    
    return redirect('affiliationcard:card_list')





def process_email_link_delivery(delivery, kwargs):
    """
    Process email delivery with download link.
    
    This function sends an email containing a secure download link
    that allows the recipient to download their card.
    
    Args:
        delivery: CardDelivery instance
        kwargs: Additional parameters from delivery creation
    """
    try:
        card = delivery.card
        
        # Generate secure download token
        delivery.download_token = secrets.token_urlsafe(32)
        delivery.download_expires_at = timezone.now() + timedelta(days=30)  # 30-day expiry
        delivery.max_downloads = 5  # Allow up to 5 downloads
        delivery.save()
        
        # Build download URL
        download_url = reverse('affiliationcard:download_card', args=[delivery.download_token])
        if hasattr(kwargs, 'request') and kwargs['request']:
            download_url = kwargs['request'].build_absolute_uri(download_url)
        
        # Prepare email context
        context = {
            'card': card,
            'delivery': delivery,
            'recipient_name': delivery.recipient_name,
            'card_number': card.card_number,
            'download_url': download_url,
            'expiry_date': delivery.download_expires_at,
            'max_downloads': delivery.max_downloads,
            'council_name': getattr(card, 'council_name', 'N/A'),
            'affiliation_type': getattr(card, 'affiliation_type', 'N/A'),
            'current_year': timezone.now().year,
        }
        
        # Render email content
        subject = kwargs.get('email_subject', f"Download Your ACRP Digital Card - {card.card_number}")
        html_message = render_to_string('email_templates/affiliationcard/card_link_delivery.html', context)
        
        # Create and send email
        email = EmailMessage(
            subject=subject,
            body='',
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[delivery.recipient_email],
            reply_to=[settings.DEFAULT_FROM_EMAIL],
        )
        
        email.attach_alternative(html_message, "text/html")
        email.send()
        
        # Update delivery status
        delivery.status = 'completed'
        delivery.completed_at = timezone.now()
        delivery.delivery_notes = f"Download link sent successfully to {delivery.recipient_email}"
        delivery.save()
        
        logger.info(f"Link email delivery completed for card {card.card_number}")
        
    except Exception as e:
        delivery.status = 'failed'
        delivery.failure_reason = str(e)
        delivery.save()
        logger.error(f"Link email delivery failed for card {card.card_number}: {e}")
        raise


def process_direct_download_delivery(delivery, kwargs):
    """
    Process direct download generation.
    
    This function prepares a card for immediate download without email,
    typically used for admin-initiated downloads or API responses.
    
    Args:
        delivery: CardDelivery instance
        kwargs: Additional parameters from delivery creation
    """
    try:
        card = delivery.card
        
        # Generate download token for tracking
        delivery.download_token = secrets.token_urlsafe(32)
        delivery.download_expires_at = timezone.now() + timedelta(hours=24)  # 24-hour expiry
        delivery.max_downloads = 1  # Single download only
        
        # Generate the card file for immediate availability
        file_content, filename, content_type = generate_card_file(
            card, 
            delivery.file_format or 'pdf'
        )
        
        # Store file metadata for later retrieval
        delivery.file_size = len(file_content)
        delivery.generated_filename = filename
        
        # Update delivery status
        delivery.status = 'ready_for_download'
        delivery.completed_at = timezone.now()
        delivery.delivery_notes = f"Direct download prepared for {card.card_number}"
        delivery.save()
        
        logger.info(f"Direct download delivery prepared for card {card.card_number}")
        
        # Return file data for immediate download
        return file_content, filename, content_type
        
    except Exception as e:
        delivery.status = 'failed'
        delivery.failure_reason = str(e)
        delivery.save()
        logger.error(f"Direct download delivery failed for card {card.card_number}: {e}")
        raise


def create_custom_report(form_data):
    """
    Create custom report based on form data.
    
    This function processes form parameters to generate comprehensive
    reports about card usage, verification patterns, and system metrics.
    
    Args:
        form_data: Dictionary containing report parameters
    
    Returns:
        dict: Structured report data ready for rendering or export
    """
    try:
        # Extract form parameters
        report_type = form_data.get('report_type', 'overview')
        date_from = form_data.get('date_from', timezone.now().date() - timedelta(days=30))
        date_to = form_data.get('date_to', timezone.now().date())
        council_filter = form_data.get('council_code')
        affiliation_filter = form_data.get('affiliation_type')
        status_filter = form_data.get('status')
        
        # Base queryset for cards
        cards_queryset = AffiliationCard.objects.all()
        
        # Apply filters
        if council_filter:
            cards_queryset = cards_queryset.filter(council_code=council_filter)
        if affiliation_filter:
            cards_queryset = cards_queryset.filter(affiliation_type=affiliation_filter)
        if status_filter:
            cards_queryset = cards_queryset.filter(status=status_filter)
        
        # Generate report data based on type
        if report_type == 'overview':
            report_data = generate_overview_report(cards_queryset, date_from, date_to)
        else:
            report_data = generate_overview_report(cards_queryset, date_from, date_to)
        
        # Add metadata
        report_data.update({
            'report_type': report_type,
            'date_range': f"{date_from} to {date_to}",
            'generated_at': timezone.now(),
            'total_cards_analyzed': cards_queryset.count(),
            'filters_applied': {
                'council': council_filter,
                'affiliation_type': affiliation_filter,
                'status': status_filter
            }
        })
        
        return report_data
        
    except Exception as e:
        logger.error(f"Error creating custom report: {e}")
        return {
            'error': str(e),
            'report_type': 'error',
            'generated_at': timezone.now()
        }


def generate_overview_report(cards_queryset, date_from, date_to):
    """Generate overview report data."""
    # Card statistics
    card_stats = {
        'total': cards_queryset.count(),
        'by_status': dict(cards_queryset.values_list('status').annotate(Count('status'))),
        'by_council': dict(cards_queryset.values('council_code', 'council_name').annotate(
            count=Count('id')
        ).values_list('council_code', 'count')),
        'by_affiliation': dict(cards_queryset.values_list('affiliation_type').annotate(Count('affiliation_type'))),
    }
    
    # Verification statistics for date range
    verifications = CardVerification.objects.filter(
        verified_at__date__gte=date_from,
        verified_at__date__lte=date_to,
        card__in=cards_queryset
    )
    
    verification_stats = {
        'total': verifications.count(),
        'by_type': dict(verifications.values_list('verification_type').annotate(Count('verification_type'))),
        'daily_counts': list(verifications.extra(
            select={'day': 'date(verified_at)'}
        ).values('day').annotate(count=Count('id')).order_by('day')),
        'success_rate': verifications.filter(was_successful=True).count() / max(verifications.count(), 1) * 100
    }
    
    return {
        'card_statistics': card_stats,
        'verification_statistics': verification_stats,
        'summary': {
            'most_active_council': max(card_stats['by_council'].items(), key=lambda x: x[1])[0] if card_stats['by_council'] else 'N/A',
            'verification_rate': verification_stats['total'] / max(card_stats['total'], 1),
            'system_utilization': f"{(card_stats['by_status'].get('active', 0) / max(card_stats['total'], 1)) * 100:.1f}%"
        }
    }


def generate_pdf_report(report_data):
    """
    Generate PDF report from report data.
    
    Creates a professionally formatted PDF document with charts,
    tables, and summary information.
    
    Args:
        report_data: Dictionary containing structured report data
    
    Returns:
        HttpResponse: PDF file response
    """
    try:
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        
        # Create PDF buffer
        buffer = BytesIO()
        
        # Create document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Build document content
        story = []
        
        # Title
        story.append(Paragraph("ACRP Card Management Report", title_style))
        story.append(Spacer(1, 12))
        
        # Report metadata
        metadata = f"""
        <b>Report Type:</b> {report_data.get('report_type', 'N/A').title()}<br/>
        <b>Date Range:</b> {report_data.get('date_range', 'N/A')}<br/>
        <b>Generated:</b> {report_data.get('generated_at', timezone.now()).strftime('%Y-%m-%d %H:%M')}<br/>
        <b>Total Cards:</b> {report_data.get('total_cards_analyzed', 0)}
        """
        story.append(Paragraph(metadata, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Card Statistics Table
        if 'card_statistics' in report_data:
            story.append(Paragraph("Card Statistics", styles['Heading2']))
            
            card_stats = report_data['card_statistics']
            table_data = [['Metric', 'Value']]
            table_data.append(['Total Cards', str(card_stats.get('total', 0))])
            
            # Add status breakdown
            for status, count in card_stats.get('by_status', {}).items():
                table_data.append([f"{status.title()} Cards", str(count)])
            
            table = Table(table_data, colWidths=[3*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 20))
        
        # Verification Statistics
        if 'verification_statistics' in report_data:
            story.append(Paragraph("Verification Statistics", styles['Heading2']))
            
            verification_stats = report_data['verification_statistics']
            table_data = [['Metric', 'Value']]
            table_data.append(['Total Verifications', str(verification_stats.get('total', 0))])
            table_data.append(['Success Rate', f"{verification_stats.get('success_rate', 0):.1f}%"])
            
            # Add verification type breakdown
            for vtype, count in verification_stats.get('by_type', {}).items():
                table_data.append([f"{vtype.replace('_', ' ').title()}", str(count)])
            
            table = Table(table_data, colWidths=[3*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
        
        # Build PDF
        doc.build(story)
        
        # Get PDF data
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Create response
        response = HttpResponse(pdf_data, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="ACRP_Card_Report_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
        
        return response
        
    except Exception as e:
        logger.error(f"Error generating PDF report: {e}")
        raise


def generate_csv_report(report_data):
    """
    Generate CSV report from report data.
    
    Creates a CSV file with detailed card and verification data
    suitable for spreadsheet analysis.
    
    Args:
        report_data: Dictionary containing structured report data
    
    Returns:
        HttpResponse: CSV file response
    """
    try:
        import csv
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="ACRP_Card_Report_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
        
        writer = csv.writer(response)
        
        # Write header information
        writer.writerow(['ACRP Card Management Report'])
        writer.writerow(['Report Type', report_data.get('report_type', 'N/A')])
        writer.writerow(['Date Range', report_data.get('date_range', 'N/A')])
        writer.writerow(['Generated At', report_data.get('generated_at', timezone.now())])
        writer.writerow(['Total Cards Analyzed', report_data.get('total_cards_analyzed', 0)])
        writer.writerow([])  # Empty row
        
        # Card Statistics Section
        if 'card_statistics' in report_data:
            writer.writerow(['CARD STATISTICS'])
            writer.writerow(['Metric', 'Value'])
            
            card_stats = report_data['card_statistics']
            writer.writerow(['Total Cards', card_stats.get('total', 0)])
            
            # Status breakdown
            for status, count in card_stats.get('by_status', {}).items():
                writer.writerow([f"{status.title()} Cards", count])
            
            writer.writerow([])  # Empty row
            
            # Council breakdown
            writer.writerow(['COUNCIL BREAKDOWN'])
            writer.writerow(['Council Code', 'Card Count'])
            for council, count in card_stats.get('by_council', {}).items():
                writer.writerow([council, count])
            
            writer.writerow([])  # Empty row
        
        # Verification Statistics Section
        if 'verification_statistics' in report_data:
            writer.writerow(['VERIFICATION STATISTICS'])
            writer.writerow(['Metric', 'Value'])
            
            verification_stats = report_data['verification_statistics']
            writer.writerow(['Total Verifications', verification_stats.get('total', 0)])
            writer.writerow(['Success Rate %', f"{verification_stats.get('success_rate', 0):.1f}"])
            
            # Verification type breakdown
            for vtype, count in verification_stats.get('by_type', {}).items():
                writer.writerow([f"{vtype.replace('_', ' ').title()}", count])
            
            writer.writerow([])  # Empty row
            
            # Daily verification counts
            daily_counts = verification_stats.get('daily_counts', [])
            if daily_counts:
                writer.writerow(['DAILY VERIFICATION TRENDS'])
                writer.writerow(['Date', 'Verification Count'])
                for day_data in daily_counts:
                    writer.writerow([day_data.get('day', ''), day_data.get('count', 0)])
        
        return response
        
    except Exception as e:
        logger.error(f"Error generating CSV report: {e}")
        raise


def generate_excel_report(report_data):
    """
    Generate Excel report from report data.
    
    Creates a multi-sheet Excel workbook with formatted data,
    charts, and summary information.
    
    Args:
        report_data: Dictionary containing structured report data
    
    Returns:
        HttpResponse: Excel file response
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import PieChart, BarChart, Reference
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet and create custom sheets
        wb.remove(wb.active)
        
        # Summary Sheet
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet['A1'] = "ACRP Card Management Report"
        summary_sheet['A1'].font = Font(size=16, bold=True)
        
        # Add report metadata
        summary_sheet['A3'] = "Report Details"
        summary_sheet['A3'].font = Font(bold=True)
        summary_sheet['A4'] = "Report Type:"
        summary_sheet['B4'] = report_data.get('report_type', 'N/A').title()
        summary_sheet['A5'] = "Date Range:"
        summary_sheet['B5'] = report_data.get('date_range', 'N/A')
        summary_sheet['A6'] = "Generated At:"
        summary_sheet['B6'] = str(report_data.get('generated_at', timezone.now()))
        summary_sheet['A7'] = "Total Cards Analyzed:"
        summary_sheet['B7'] = report_data.get('total_cards_analyzed', 0)
        
        # Card Statistics Sheet
        if 'card_statistics' in report_data:
            card_sheet = wb.create_sheet("Card Statistics")
            card_sheet['A1'] = "Card Statistics"
            card_sheet['A1'].font = Font(size=14, bold=True)
            
            # Headers
            card_sheet['A3'] = "Metric"
            card_sheet['B3'] = "Value"
            card_sheet['A3'].font = Font(bold=True)
            card_sheet['B3'].font = Font(bold=True)
            
            # Data
            row = 4
            card_stats = report_data['card_statistics']
            card_sheet[f'A{row}'] = "Total Cards"
            card_sheet[f'B{row}'] = card_stats.get('total', 0)
            row += 1
            
            # Status breakdown
            for status, count in card_stats.get('by_status', {}).items():
                card_sheet[f'A{row}'] = f"{status.title()} Cards"
                card_sheet[f'B{row}'] = count
                row += 1
            
            # Council breakdown in separate section
            row += 2
            card_sheet[f'A{row}'] = "Council Breakdown"
            card_sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            card_sheet[f'A{row}'] = "Council Code"
            card_sheet[f'B{row}'] = "Card Count"
            card_sheet[f'A{row}'].font = Font(bold=True)
            card_sheet[f'B{row}'].font = Font(bold=True)
            row += 1
            
            for council, count in card_stats.get('by_council', {}).items():
                card_sheet[f'A{row}'] = council
                card_sheet[f'B{row}'] = count
                row += 1
        
        # Verification Statistics Sheet
        if 'verification_statistics' in report_data:
            verification_sheet = wb.create_sheet("Verification Statistics")
            verification_sheet['A1'] = "Verification Statistics"
            verification_sheet['A1'].font = Font(size=14, bold=True)
            
            # Headers
            verification_sheet['A3'] = "Metric"
            verification_sheet['B3'] = "Value"
            verification_sheet['A3'].font = Font(bold=True)
            verification_sheet['B3'].font = Font(bold=True)
            
            # Data
            row = 4
            verification_stats = report_data['verification_statistics']
            verification_sheet[f'A{row}'] = "Total Verifications"
            verification_sheet[f'B{row}'] = verification_stats.get('total', 0)
            row += 1
            
            verification_sheet[f'A{row}'] = "Success Rate %"
            verification_sheet[f'B{row}'] = f"{verification_stats.get('success_rate', 0):.1f}%"
            row += 1
            
            # Verification type breakdown
            for vtype, count in verification_stats.get('by_type', {}).items():
                verification_sheet[f'A{row}'] = f"{vtype.replace('_', ' ').title()}"
                verification_sheet[f'B{row}'] = count
                row += 1
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="ACRP_Card_Report_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        
        return response
        
    except ImportError:
        # Fallback to CSV if openpyxl is not installed
        logger.warning("openpyxl not installed, falling back to CSV export")
        return generate_csv_report(report_data)
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        raise





# ============================================================================
# PUBLIC VERIFICATION VIEWS
# ============================================================================

@never_cache
def verify_lookup(request):
    """
    Public card lookup/verification by card number.
    
    This is the main public verification page where anyone can
    enter a card number to verify affiliate status.
    """
    if not rate_limit_verification(request):
        return render(request, 'affiliationcard/public/rate_limited.html', status=429)
    
    verification_result = None
    
    if request.method == 'POST':
        form = CardLookupForm(request.POST)
        
        if form.is_valid():
            card_number = form.cleaned_data['card_number']
            purpose = form.cleaned_data.get('verification_purpose', '')
            
            try:
                # Find the card
                card = AffiliationCard.objects.get(card_number=card_number)
                
                # Record verification
                verification = CardVerification.objects.create(
                    card=card,
                    verification_type='manual_lookup',
                    verified_at=timezone.now(),
                    ip_address=get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                    was_successful=True,
                    card_status_at_time=card.status,
                    verification_purpose=purpose
                )
                
                # Update card verification count
                card.increment_verification_count()
                
                verification_result = {
                    'success': True,
                    'card': card,
                    'verification': verification
                }
                
            except AffiliationCard.DoesNotExist:
                verification_result = {
                    'success': False,
                    'error': 'Card not found',
                    'message': 'No card found with this number. Please check the card number and try again.'
                }
            except Exception as e:
                logger.error(f"Verification error: {e}")
                verification_result = {
                    'success': False,
                    'error': 'system_error',
                    'message': 'System error occurred during verification. Please try again later.'
                }
    else:
        form = CardLookupForm()
    
    context = {
        'form': form,
        'verification_result': verification_result,
        'page_title': 'Verify Affiliation Card'
    }
    
    return render(request, 'affiliationcard/public/verify_lookup.html', context)


@never_cache
def verify_token(request, token):
    """
    Verify card using QR code verification token.
    
    This is called when someone scans a QR code on a card.
    """
    if not rate_limit_verification(request):
        return render(request, 'affiliationcard/public/rate_limited.html', status=429)
    
    try:
        # Find card by verification token
        card = AffiliationCard.objects.get(verification_token=token)
        
        # Record verification
        verification = CardVerification.objects.create(
            card=card,
            verification_type='qr_scan',
            verified_at=timezone.now(),
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            referer=request.META.get('HTTP_REFERER', ''),
            was_successful=True,
            card_status_at_time=card.status
        )
        
        # Update card verification count
        card.increment_verification_count()
        
        verification_result = {
            'success': True,
            'card': card,
            'verification': verification
        }
        
    except AffiliationCard.DoesNotExist:
        verification_result = {
            'success': False,
            'error': 'invalid_token',
            'message': 'Invalid verification token. This may be a counterfeit card.'
        }
    except Exception as e:
        logger.error(f"Token verification error: {e}")
        verification_result = {
            'success': False,
            'error': 'system_error',
            'message': 'System error occurred during verification.'
        }
    
    context = {
        'verification_result': verification_result,
        'page_title': 'Card Verification Result'
    }
    
    return render(request, 'affiliationcard/public/verification_result.html', context)


@csrf_exempt
@require_POST
def api_verify(request):
    """
    API endpoint for card verification.
    
    Returns JSON response for programmatic verification.
    """
    if not rate_limit_verification(request):
        return JsonResponse({
            'success': False,
            'error': 'rate_limit_exceeded',
            'message': 'Too many verification attempts. Please try again later.'
        }, status=429)
    
    try:
        # Parse request data
        data = json.loads(request.body)
        card_number = data.get('card_number')
        token = data.get('verification_token')
        
        if not (card_number or token):
            return JsonResponse({
                'success': False,
                'error': 'missing_parameters',
                'message': 'Either card_number or verification_token is required'
            }, status=400)
        
        # Find card
        if token:
            card = AffiliationCard.objects.get(verification_token=token)
        else:
            card = AffiliationCard.objects.get(card_number=card_number)
        
        # Record verification
        verification = CardVerification.objects.create(
            card=card,
            verification_type='api_verification',
            verified_at=timezone.now(),
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            was_successful=True,
            card_status_at_time=card.status,
            verification_purpose=data.get('purpose', '')
        )
        
        # Update verification count
        card.increment_verification_count()
        
        # Return card information
        return JsonResponse({
            'success': True,
            'card': {
                'card_number': card.card_number,
                'affiliate_name': card.get_display_name(),
                'council': card.council_code,
                'affiliation_type': card.affiliation_type,
                'status': card.status,
                'is_active': card.is_active(),
                'is_expired': card.is_expired(),
                'date_issued': card.date_issued.isoformat() if card.date_issued else None,
                'date_expires': card.date_expires.isoformat() if card.date_expires else None,
                'days_until_expiry': card.days_until_expiry()
            },
            'verification_id': verification.id
        })
        
    except AffiliationCard.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'card_not_found',
            'message': 'Card not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'invalid_json',
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        logger.error(f"API verification error: {e}")
        return JsonResponse({
            'success': False,
            'error': 'system_error',
            'message': 'System error occurred'
        }, status=500)


# ============================================================================
# CARD DELIVERY VIEWS
# ============================================================================

# ============================================================================
# ENHANCED CARD DELIVERY PROCESSING FUNCTIONS
# ============================================================================

import base64
import secrets
from datetime import timedelta
from io import BytesIO
from django.conf import settings
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from mailjet_rest import Client
import logging

logger = logging.getLogger(__name__)

def get_mailjet_client():
    """
    Initialize and return Mailjet client.
    
    Requires MAILJET_API_KEY and MAILJET_SECRET_KEY in settings.
    """
    try:
        return Client(
            auth=(settings.MAILJET_API_KEY, settings.MAILJET_SECRET_KEY),
            version='v3.1'
        )
    except Exception as e:
        logger.error(f"Failed to initialize Mailjet client: {e}")
        raise


def create_card_delivery(card, delivery_method, recipient_email, recipient_name, **kwargs):
    """
    Create and process card delivery with enhanced error handling.
    
    Args:
        card: AffiliationCard instance
        delivery_method: String - 'email_pdf', 'email_link', or 'direct_download'
        recipient_email: String - recipient's email address
        recipient_name: String - recipient's full name
        **kwargs: Additional parameters (email_subject, email_message, file_format, etc.)
    
    Returns:
        CardDelivery instance
    """
    try:
        # Create delivery record with initial status
        delivery = CardDelivery.objects.create(
            card=card,
            delivery_type=delivery_method,
            recipient_email=recipient_email,
            recipient_name=recipient_name,
            initiated_by=kwargs.get('initiated_by'),
            file_format=kwargs.get('file_format', 'pdf'),
            status='processing',
            email_subject=kwargs.get('email_subject', ''),
            email_message=kwargs.get('email_message', ''),
            max_downloads=kwargs.get('max_downloads', 5),
            download_expires_at=timezone.now() + timedelta(days=30)
        )
        
        logger.info(f"Created card delivery {delivery.id} for card {card.card_number}")
        
        # Process delivery based on method
        if delivery_method == 'email_pdf':
            process_email_pdf_delivery(delivery, kwargs)
        elif delivery_method == 'email_link':
            process_email_link_delivery(delivery, kwargs)
        elif delivery_method == 'direct_download':
            process_direct_download_delivery(delivery, kwargs)
        else:
            raise ValueError(f"Unsupported delivery method: {delivery_method}")
        
        return delivery
        
    except Exception as e:
        logger.error(f"Failed to create card delivery: {e}")
        # Update delivery status if it was created
        if 'delivery' in locals():
            delivery.status = 'failed'
            delivery.failure_reason = str(e)
            delivery.save()
        raise


def process_email_link_delivery(delivery, kwargs):
    """
    Process email delivery with secure download link - FIXED DJANGO EMAIL VERSION
    """
    try:
        card = delivery.card
        request = kwargs.get('request')
        
        logger.info(f"Starting link email delivery for {card.card_number}")
        logger.info(f"Delivery ID: {delivery.id}, Status: {delivery.status}")
        
        # Update status to processing
        delivery.status = 'processing'
        delivery.save()
        logger.info("Updated status to processing")
        
        # Generate secure download token
        if not delivery.download_token:
            delivery.download_token = secrets.token_urlsafe(32)
            delivery.download_expires_at = timezone.now() + timedelta(days=30)
            delivery.max_downloads = kwargs.get('max_downloads', 5)
            delivery.save()
            
        logger.info(f"Generated download token: {delivery.download_token[:10]}...")
        
        # Build download URL
        download_path = reverse('affiliationcard:download_card', args=[delivery.download_token])
        if request:
            download_url = request.build_absolute_uri(download_path)
        else:
            download_url = f"http://localhost:8000{download_path}"
            
        logger.info(f"Download URL: {download_url}")
        
        # Prepare email context
        context = {
            'card': card,
            'delivery': delivery,
            'recipient_name': delivery.recipient_name,
            'card_number': card.card_number,
            'download_url': download_url,
            'expiry_date': delivery.download_expires_at,
            'max_downloads': delivery.max_downloads,
            'council_name': getattr(card, 'council_name', 'N/A'),
            'affiliation_type': getattr(card, 'affiliation_type', 'N/A').title() if hasattr(card, 'affiliation_type') else 'N/A',
            'current_year': timezone.now().year,
            'system_name': 'ACRP AMS',
        }
        
        # Render email content (try template first, fallback to simple)
        logger.info("Preparing email content...")
        try:
            html_content = render_to_string('email_templates/affiliationcard/card_link_delivery.html', context)
            logger.info("Used template for HTML content")
        except Exception as template_error:
            logger.warning(f"Template not found, using fallback HTML: {template_error}")
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>Download Your ACRP Digital Card</title>
            </head>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                    <h1 style="color: #1e40af;">Download Your ACRP Digital Card</h1>
                    
                    <p>Dear {delivery.recipient_name},</p>
                    
                    <p>Your ACRP digital affiliation card is ready for download.</p>
                    
                    <div style="background-color: #f8f9fa; padding: 20px; border-left: 4px solid #1e40af; margin: 20px 0;">
                        <h3 style="margin: 0 0 15px 0;">Card Details</h3>
                        <p style="margin: 0;"><strong>Card Number:</strong> {card.card_number}</p>
                        <p style="margin: 0;"><strong>Council:</strong> {context['council_name']}</p>
                        <p style="margin: 0;"><strong>Affiliation Type:</strong> {context['affiliation_type']}</p>
                    </div>
                    
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{download_url}" 
                           style="background-color: #1e40af; color: white; padding: 15px 30px; text-decoration: none; border-radius: 8px; font-weight: bold; display: inline-block;">
                            Download Your Card
                        </a>
                    </div>
                    
                    <div style="background-color: #fff3cd; padding: 15px; border-radius: 8px; margin: 20px 0;">
                        <h4 style="margin: 0 0 10px 0; color: #856404;">Important Information</h4>
                        <ul style="margin: 0; color: #856404;">
                            <li>Link expires on {delivery.download_expires_at.strftime('%B %d, %Y')}</li>
                            <li>Can be downloaded up to {delivery.max_downloads} times</li>
                            <li>Keep your card secure and do not share this link</li>
                        </ul>
                    </div>
                    
                    <p>If the button doesn't work, copy this link: <br>
                    <code style="background-color: #f8f9fa; padding: 8px; border-radius: 4px; display: block; margin: 10px 0; word-break: break-all;">{download_url}</code></p>
                    
                    <p>Best regards,<br>The ACRP Digital Cards Team</p>
                </div>
            </body>
            </html>
            """
        
        # Simple text content
        text_content = f"""Download Your ACRP Digital Card - {card.card_number}

Dear {delivery.recipient_name},

Your ACRP digital affiliation card is ready for download.

Download Link: {download_url}

Card Details:
- Card Number: {card.card_number}
- Council: {context['council_name']}
- Affiliation Type: {context['affiliation_type']}

Important Information:
- Link expires on {delivery.download_expires_at.strftime('%B %d, %Y')}
- Can be downloaded up to {delivery.max_downloads} times
- Keep your card secure and do not share this link

Best regards,
The ACRP Digital Cards Team"""
        
        # Create and send email using SAME METHOD as PDF function
        logger.info("Creating email message...")
        subject = kwargs.get('email_subject', f'Download Your ACRP Digital Card - {card.card_number}')
        
        from django.core.mail import EmailMultiAlternatives
        
        email = EmailMultiAlternatives(
            subject=subject,
            body=text_content,
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'dave@kreeck.com'),
            to=[delivery.recipient_email],
            reply_to=[getattr(settings, 'DEFAULT_FROM_EMAIL', 'dave@kreeck.com')]
        )
        
        # Add HTML version
        email.attach_alternative(html_content, "text/html")
        
        logger.info(f"Sending email to {delivery.recipient_email}...")
        
        # Send the email
        email.send()
        
        logger.info("Email sent successfully!")
        
        # Update delivery status
        delivery.status = 'completed'
        delivery.completed_at = timezone.now()
        delivery.delivery_notes = f"Download link sent successfully to {delivery.recipient_email}"
        delivery.save()
        
        logger.info(f"Link email delivery completed for card {card.card_number}")
        
    except Exception as e:
        logger.error(f"Link email delivery failed for card {card.card_number}: {e}")
        
        # Update delivery status on failure
        delivery.status = 'failed'
        delivery.failure_reason = str(e)
        delivery.save()
        raise


def process_direct_download_delivery(delivery, kwargs):
    """
    Process direct download generation - FIXED VERSION
    """
    try:
        card = delivery.card
        
        logger.info(f"Starting direct download for card {card.card_number}")
        logger.info(f"Delivery ID: {delivery.id}, Status: {delivery.status}")
        
        # Update status to processing
        delivery.status = 'processing'
        delivery.save()
        logger.info("Updated status to processing")
        
        # Generate download token for tracking
        if not delivery.download_token:
            delivery.download_token = secrets.token_urlsafe(32)
            delivery.download_expires_at = timezone.now() + timedelta(hours=24)
            delivery.max_downloads = 1  # Single download for direct downloads
            delivery.save()
        
        logger.info(f"Generated download token: {delivery.download_token[:10]}...")
        
        # Generate the card file for immediate availability
        logger.info("Generating card file...")
        file_content, filename, content_type = generate_card_file_simple(
            card, 
            delivery.file_format or 'pdf'
        )
        
        logger.info(f"Card file generated: {filename} ({len(file_content)} bytes)")
        
        # Store file metadata for tracking
        delivery.file_size = len(file_content)
        delivery.generated_filename = filename
        
        # Update delivery status to ready
        delivery.status = 'ready_for_download'
        delivery.completed_at = timezone.now()
        delivery.delivery_notes = f"Direct download prepared for {card.card_number}"
        delivery.save()
        
        logger.info(f"DIRECT DOWNLOAD COMPLETED: Status = {delivery.status}")
        logger.info(f"Direct download delivery prepared for card {card.card_number}")
        
        # Return file data for immediate download
        return file_content, filename, content_type
        
    except Exception as e:
        logger.error(f"Direct download delivery failed for card {card.card_number}: {e}")
        
        # Update delivery status on failure
        delivery.status = 'failed'
        delivery.failure_reason = str(e)
        delivery.save()
        raise



def generate_card_pdf(card):
    """
    Generate enhanced PDF version of the card with professional design.
    
    Creates a business card sized PDF with proper layout, QR code,
    and ACRP branding elements.
    
    Args:
        card: AffiliationCard instance
        
    Returns:
        tuple: (pdf_content, filename, content_type)
    """
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch, mm
        from reportlab.lib import colors
        from reportlab.graphics import renderPDF
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import HorizontalBarChart
        import qrcode
        from PIL import Image as PILImage
        
        # Create buffer for PDF
        buffer = BytesIO()
        
        # Standard business card size (3.5" x 2")
        card_width = 3.5 * inch
        card_height = 2 * inch
        
        # Create canvas with business card dimensions
        c = canvas.Canvas(buffer, pagesize=(card_width, card_height))
        
        # Set up colors
        primary_color = colors.HexColor('#1e40af')  # Blue
        secondary_color = colors.HexColor('#64748b')  # Gray
        text_color = colors.HexColor('#1f2937')  # Dark gray
        
        # Background
        c.setFillColor(colors.white)
        c.rect(0, 0, card_width, card_height, fill=1)
        
        # Header section with background
        c.setFillColor(primary_color)
        c.rect(0, card_height - 40, card_width, 40, fill=1)
        
        # ACRP Logo/Title
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(10, card_height - 25, "ACRP")
        
        c.setFont("Helvetica", 8)
        c.drawString(10, card_height - 35, "Digital Affiliation Card")
        
        # Card number (top right)
        c.setFont("Helvetica-Bold", 8)
        text_width = c.stringWidth(f"#{card.card_number}", "Helvetica-Bold", 8)
        c.drawString(card_width - text_width - 10, card_height - 25, f"#{card.card_number}")
        
        # Main content area
        c.setFillColor(text_color)
        y_position = card_height - 60
        
        # Affiliate name
        c.setFont("Helvetica-Bold", 12)
        affiliate_name = card.get_display_name()
        if len(affiliate_name) > 25:  # Truncate long names
            affiliate_name = affiliate_name[:22] + "..."
        c.drawString(10, y_position, affiliate_name)
        y_position -= 15
        
        # Council and affiliation type
        c.setFont("Helvetica", 8)
        c.setFillColor(secondary_color)
        council_text = f"{getattr(card, 'council_name', 'N/A')}"
        if len(council_text) > 30:
            council_text = council_text[:27] + "..."
        c.drawString(10, y_position, council_text)
        y_position -= 10
        
        affiliation_text = f"{getattr(card, 'affiliation_type', 'N/A').title()}"
        c.drawString(10, y_position, affiliation_text)
        y_position -= 15
        
        # Status and dates
        c.setFont("Helvetica", 7)
        c.setFillColor(text_color)
        
        status_text = f"Status: {card.get_status_display()}"
        c.drawString(10, y_position, status_text)
        y_position -= 8
        
        if card.date_issued:
            issued_text = f"Issued: {card.date_issued.strftime('%m/%d/%Y')}"
            c.drawString(10, y_position, issued_text)
            y_position -= 8
        
        if card.date_expires:
            expires_text = f"Expires: {card.date_expires.strftime('%m/%d/%Y')}"
            c.drawString(10, y_position, expires_text)
        
        # Generate and add QR code
        qr_data = card.qr_code_data
        qr = qrcode.QRCode(version=1, box_size=2, border=1)
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        # Create QR code image
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        # Save QR code to temporary buffer
        qr_buffer = BytesIO()
        qr_img.save(qr_buffer, format='PNG')
        qr_buffer.seek(0)
        
        # Add QR code to PDF (right side)
        qr_size = 50  # Size in points
        qr_x = card_width - qr_size - 10
        qr_y = 15
        
        c.drawInlineImage(qr_buffer, qr_x, qr_y, qr_size, qr_size)
        
        # QR code label
        c.setFont("Helvetica", 6)
        c.setFillColor(secondary_color)
        qr_label_width = c.stringWidth("Scan to verify", "Helvetica", 6)
        c.drawString(qr_x + (qr_size - qr_label_width) / 2, qr_y - 8, "Scan to verify")
        
        # Footer with verification URL (small text)
        c.setFont("Helvetica", 5)
        footer_text = f"Verify at: your-domain.com/verify/{card.verification_token[:8]}..."
        c.drawString(10, 5, footer_text)
        
        # Card border
        c.setStrokeColor(colors.lightgrey)
        c.setLineWidth(0.5)
        c.rect(0, 0, card_width, card_height, fill=0, stroke=1)
        
        # Save PDF
        c.save()
        
        # Get PDF content
        pdf_content = buffer.getvalue()
        buffer.close()
        
        filename = f"ACRP_Card_{card.card_number}.pdf"
        content_type = 'application/pdf'
        
        logger.info(f"Generated PDF card for {card.card_number}")
        return pdf_content, filename, content_type
        
    except Exception as e:
        logger.error(f"Failed to generate PDF card for {card.card_number}: {e}")
        raise


def generate_card_image(card, format_type):
    """
    Generate professional image version of the card.
    
    Creates a business card sized image with modern design,
    proper typography, and QR code integration.
    
    Args:
        card: AffiliationCard instance
        format_type: String - 'PNG' or 'JPEG'
        
    Returns:
        tuple: (image_content, filename, content_type)
    """
    try:
        from PIL import Image, ImageDraw, ImageFont, ImageColor
        import qrcode
        
        # Card dimensions (business card size at 300 DPI)
        width, height = 1050, 600  # 3.5" x 2" at 300 DPI
        
        # Create card background with gradient effect
        img = Image.new('RGB', (width, height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Colors
        primary_color = '#1e40af'
        secondary_color = '#64748b'
        text_color = '#1f2937'
        light_gray = '#f8fafc'
        
        # Header background
        header_height = 120
        draw.rectangle([(0, 0), (width, header_height)], fill=primary_color)
        
        # Try to load custom fonts, fallback to default if not available
        try:
            title_font = ImageFont.truetype("arial.ttf", 42)
            subtitle_font = ImageFont.truetype("arial.ttf", 24)
            name_font = ImageFont.truetype("arial.ttf", 36)
            info_font = ImageFont.truetype("arial.ttf", 20)
            small_font = ImageFont.truetype("arial.ttf", 16)
        except:
            # Fallback to default font
            title_font = ImageFont.load_default()
            subtitle_font = ImageFont.load_default()
            name_font = ImageFont.load_default()
            info_font = ImageFont.load_default()
            small_font = ImageFont.load_default()
        
        # ACRP Title
        draw.text((30, 25), "ACRP", fill='white', font=title_font)
        draw.text((30, 70), "Digital Affiliation Card", fill='white', font=subtitle_font)
        
        # Card number (top right)
        card_num_text = f"#{card.card_number}"
        card_num_bbox = draw.textbbox((0, 0), card_num_text, font=info_font)
        card_num_width = card_num_bbox[2] - card_num_bbox[0]
        draw.text((width - card_num_width - 30, 25), card_num_text, fill='white', font=info_font)
        
        # Main content area
        y_pos = header_height + 30
        
        # Affiliate name
        affiliate_name = card.get_display_name()
        if len(affiliate_name) > 20:  # Adjust for image layout
            affiliate_name = affiliate_name[:17] + "..."
        
        draw.text((30, y_pos), affiliate_name, fill=text_color, font=name_font)
        y_pos += 50
        
        # Council and affiliation
        council_text = getattr(card, 'council_name', 'N/A')
        if len(council_text) > 35:
            council_text = council_text[:32] + "..."
        
        draw.text((30, y_pos), council_text, fill=secondary_color, font=info_font)
        y_pos += 30
        
        affiliation_text = getattr(card, 'affiliation_type', 'N/A').title()
        draw.text((30, y_pos), affiliation_text, fill=secondary_color, font=info_font)
        y_pos += 40
        
        # Status and dates
        status_text = f"Status: {card.get_status_display()}"
        draw.text((30, y_pos), status_text, fill=text_color, font=small_font)
        y_pos += 25
        
        if card.date_issued:
            issued_text = f"Issued: {card.date_issued.strftime('%m/%d/%Y')}"
            draw.text((30, y_pos), issued_text, fill=secondary_color, font=small_font)
            y_pos += 20
        
        if card.date_expires:
            expires_text = f"Expires: {card.date_expires.strftime('%m/%d/%Y')}"
            draw.text((30, y_pos), expires_text, fill=secondary_color, font=small_font)
        
        # Generate QR code
        qr_data = card.qr_code_data
        qr = qrcode.QRCode(version=1, box_size=4, border=2)
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_size = 150
        qr_img = qr_img.resize((qr_size, qr_size))
        
        # Position QR code on the right side
        qr_x = width - qr_size - 30
        qr_y = height - qr_size - 30
        img.paste(qr_img, (qr_x, qr_y))
        
        # QR code label
        qr_label = "Scan to verify"
        qr_label_bbox = draw.textbbox((0, 0), qr_label, font=small_font)
        qr_label_width = qr_label_bbox[2] - qr_label_bbox[0]
        qr_label_x = qr_x + (qr_size - qr_label_width) // 2
        draw.text((qr_label_x, qr_y - 25), qr_label, fill=secondary_color, font=small_font)
        
        # Footer
        footer_text = f"Verify at: your-domain.com/verify/{card.verification_token[:8]}..."
        draw.text((30, height - 25), footer_text, fill=secondary_color, font=small_font)
        
        # Card border
        draw.rectangle([(0, 0), (width-1, height-1)], outline='lightgray', width=2)
        
        # Save image to buffer
        buffer = BytesIO()
        img.save(buffer, format=format_type, quality=95 if format_type == 'JPEG' else None)
        buffer.seek(0)
        
        # Prepare response data
        extension = format_type.lower()
        if extension == 'jpeg':
            extension = 'jpg'
        
        filename = f"ACRP_Card_{card.card_number}.{extension}"
        content_type = f'image/{extension}'
        
        logger.info(f"Generated {format_type} card image for {card.card_number}")
        return buffer.getvalue(), filename, content_type
        
    except Exception as e:
        logger.error(f"Failed to generate {format_type} card for {card.card_number}: {e}")
        raise




    
# ============================================================================
# SYSTEM ADMINISTRATION VIEWS
# ============================================================================

class CardTemplateListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """List all card templates."""
    
    model = CardTemplate
    template_name = 'affiliationcard/admin/template_list.html'
    context_object_name = 'templates'
    permission_required = 'affiliationcard.view_cardtemplate'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Card Templates'
        return context


class CardTemplateCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create new card template."""
    
    model = CardTemplate
    form_class = CardTemplateForm
    template_name = 'affiliationcard/admin/template_form.html'
    permission_required = 'affiliationcard.add_cardtemplate'
    success_url = reverse_lazy('affiliationcard:template_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, "Card template created successfully!")
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create Card Template'
        return context


@login_required
@user_passes_test(lambda u: u.is_superuser)
def system_settings(request):
    """Manage system settings."""
    settings_obj = CardSystemSettings.get_settings()
    
    if request.method == 'POST':
        form = SystemSettingsForm(request.POST, instance=settings_obj)
        
        if form.is_valid():
            form.instance.updated_by = request.user
            form.save()
            messages.success(request, "System settings updated successfully!")
            return redirect('affiliationcard:system_settings')
    else:
        form = SystemSettingsForm(instance=settings_obj)
    
    context = {
        'form': form,
        'settings': settings_obj,
        'page_title': 'Card System Settings'
    }
    
    return render(request, 'affiliationcard/admin/system_settings.html', context)


# ============================================================================
# REPORTING AND ANALYTICS VIEWS
# ============================================================================

@login_required
@user_passes_test(is_card_admin)
def analytics_dashboard(request):
    """Analytics dashboard with charts and statistics."""
    
    # Date range
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    # Card statistics
    card_stats = {
        'total': AffiliationCard.objects.count(),
        'active': AffiliationCard.objects.filter(status='active').count(),
        'pending': AffiliationCard.objects.filter(status='pending_assignment').count(),
        'expired': AffiliationCard.objects.filter(status='expired').count(),
        'suspended': AffiliationCard.objects.filter(status='suspended').count(),
    }
    
    # Verification statistics
    verification_stats = {
        'total': CardVerification.objects.count(),
        'last_30_days': CardVerification.objects.filter(
            verified_at__gte=timezone.now() - timedelta(days=30)
        ).count(),
        'qr_scans': CardVerification.objects.filter(verification_type='qr_scan').count(),
        'manual_lookups': CardVerification.objects.filter(verification_type='manual_lookup').count(),
    }
    
    # Council breakdown
    council_breakdown = AffiliationCard.objects.values('council_code', 'council_name').annotate(
        total=Count('id'),
        active=Count('id', filter=Q(status='active'))
    ).order_by('council_code')
    
    # Daily verification trend (last 30 days)
    daily_verifications = []
    for i in range(30):
        date = end_date - timedelta(days=i)
        count = CardVerification.objects.filter(
            verified_at__date=date
        ).count()
        daily_verifications.append({
            'date': date.isoformat(),
            'count': count
        })
    daily_verifications.reverse()
    
    context = {
        'card_stats': card_stats,
        'verification_stats': verification_stats,
        'council_breakdown': council_breakdown,
        'daily_verifications': daily_verifications,
        'date_range': f"{start_date} to {end_date}",
        'page_title': 'Card Analytics'
    }
    
    return render(request, 'affiliationcard/admin/analytics.html', context)


@login_required
@user_passes_test(is_card_admin)
def generate_report(request):
    """Generate custom reports."""
    if request.method == 'POST':
        form = CardReportForm(request.POST)
        
        if form.is_valid():
            report_data = create_custom_report(form.cleaned_data)
            
            output_format = form.cleaned_data['output_format']
            
            if output_format == 'pdf':
                return generate_pdf_report(report_data)
            elif output_format == 'csv':
                return generate_csv_report(report_data)
            elif output_format == 'excel':
                return generate_excel_report(report_data)
            else:
                context = {
                    'report_data': report_data,
                    'form': form,
                    'page_title': 'Card Report'
                }
                return render(request, 'affiliationcard/admin/report_results.html', context)
    else:
        form = CardReportForm()
    
    context = {
        'form': form,
        'page_title': 'Generate Report'
    }
    
    return render(request, 'affiliationcard/admin/generate_report.html', context)


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def send_card_assignment_notification(card, request):
    """Send email notification when card is assigned."""
    subject = f"Your ACRP Digital Affiliation Card is Ready - {card.card_number}"
    
    context = {
        'card': card,
        'download_url': request.build_absolute_uri(
            reverse('affiliationcard:download_card', args=[card.verification_token])
        )
    }
    
    message = render_to_string('affiliationcard/emails/card_assigned.html', context)
    
    send_mail(
        subject=subject,
        message=message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[card.affiliate_email],
        html_message=message
    )


def send_status_change_notification(card, old_status, new_status, reason):
    """Send notification when card status changes."""
    subject = f"ACRP Card Status Update - {card.card_number}"
    
    context = {
        'card': card,
        'old_status': old_status,
        'new_status': new_status,
        'reason': reason
    }
    
    message = render_to_string('affiliationcard/emails/status_change.html', context)
    
    send_mail(
        subject=subject,
        message=message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[card.affiliate_email],
        html_message=message
    )


def create_card_delivery(card, delivery_method, recipient_email, recipient_name, **kwargs):
    """Create and process card delivery."""
    delivery = CardDelivery.objects.create(
        card=card,
        delivery_type=delivery_method,
        recipient_email=recipient_email,
        recipient_name=recipient_name,
        initiated_by=kwargs.get('initiated_by'),
        file_format=kwargs.get('file_format', 'pdf')
    )
    
    # Process delivery based on method
    if delivery_method == 'email_pdf':
        process_email_pdf_delivery(delivery, kwargs)
    elif delivery_method == 'email_link':
        process_email_link_delivery(delivery, kwargs)
    elif delivery_method == 'direct_download':
        process_direct_download_delivery(delivery, kwargs)
    
    return delivery



from io import BytesIO
from PIL import Image, ImageDraw, ImageFilter, ImageFont
import qrcode
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch, mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import logging

logger = logging.getLogger(__name__)






def generate_card_pdf(card):
    """
    Generate a professional business card-sized PDF with proper spacing and coordinates.
    Creates a credit card-sized PDF (3.375" x 2.125") with carefully positioned elements.
    
    Key improvements:
    - Proper element spacing to prevent overlap
    - Optimized font sizes for card dimensions
    - Coordinated layout with clear visual hierarchy
    - Functional QR code with verification URL
    
    Args:
        card: AffiliationCard instance containing card data
        
    Returns:
        tuple: (pdf_content, filename, content_type)
        
    Raises:
        Exception: If PDF generation fails
    """
    try:
        # Create buffer for PDF output
        buffer = BytesIO()
        
        # Business card dimensions (3.375" x 2.125" - standard credit card size)
        card_width = 3.375 * inch
        card_height = 2.125 * inch
        
        # Create canvas with specified dimensions
        c = canvas.Canvas(buffer, pagesize=(card_width, card_height))
        
        # Professional color palette optimized for business cards
        primary_navy = colors.HexColor('#0f172a')       # Deep navy
        primary_blue = colors.HexColor('#1e40af')       # Professional blue
        accent_blue = colors.HexColor('#3b82f6')        # Bright blue
        accent_emerald = colors.HexColor('#059669')     # Emerald green
        accent_gold = colors.HexColor('#d97706')        # Warning gold
        text_primary = colors.HexColor('#111827')       # Almost black
        text_secondary = colors.HexColor('#4b5563')     # Medium gray
        text_muted = colors.HexColor('#9ca3af')         # Light gray
        
        # Create white background
        c.setFillColor(colors.white)
        c.rect(0, 0, card_width, card_height, fill=1)
        
        # ============================================================================
        # HEADER SECTION - Carefully measured and spaced
        # ============================================================================
        header_height = 0.75 * inch
        
        # Main header background (navy)
        c.setFillColor(primary_navy)
        c.rect(0, card_height - header_height, card_width, header_height, fill=1)
        
        # Accent stripe using path operations (ReportLab compatible)
        c.setFillColor(accent_blue)
        path = c.beginPath()
        path.moveTo(0, card_height - header_height + 0.25 * inch)
        path.lineTo(card_width * 0.4, card_height - header_height + 0.25 * inch)
        path.lineTo(card_width * 0.45, card_height - header_height)
        path.lineTo(0, card_height - header_height)
        path.close()
        c.drawPath(path, fill=1, stroke=0)
        
        # ACRP Logo - positioned to avoid crowding
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(0.2 * inch, card_height - 0.28 * inch, "ACRP")
        
        # Organization name - smaller and well-positioned
        c.setFont("Helvetica", 6)
        c.setFillColorRGB(1, 1, 1, alpha=0.9)
        c.drawString(0.2 * inch, card_height - 0.42 * inch, "ASSOCIATION OF CHRISTIAN")
        c.drawString(0.2 * inch, card_height - 0.5 * inch, "RELIGIOUS PRACTITIONERS")
        
        # Card type badge - compact and positioned
        c.setFillColor(accent_emerald)
        badge_x = 0.2 * inch
        badge_y = card_height - 0.68 * inch
        badge_width = 0.9 * inch
        badge_height = 0.1 * inch
        c.roundRect(badge_x, badge_y, badge_width, badge_height, 0.03 * inch, fill=1)
        
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 5)
        c.drawString(badge_x + 0.03 * inch, badge_y + 0.025 * inch, "DIGITAL AFFILIATION CARD")
        
        # Card number - top right, properly spaced
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 9)
        card_num_text = f"#{card.card_number}"
        text_width = c.stringWidth(card_num_text, "Helvetica-Bold", 9)
        c.drawString(card_width - text_width - 0.2 * inch, card_height - 0.25 * inch, card_num_text)
        
        # Status badge - positioned below card number
        status = card.get_status_display().upper()
        c.setFont("Helvetica-Bold", 5)
        
        # Status color logic
        if hasattr(card, 'status') and card.status == 'active':
            status_color = accent_emerald
        elif hasattr(card, 'status') and card.status == 'pending':
            status_color = accent_gold
        else:
            status_color = colors.HexColor('#ef4444')  # Red for inactive
        
        # Status badge dimensions and positioning
        status_width = c.stringWidth(status, "Helvetica-Bold", 5)
        status_x = card_width - status_width - 0.3 * inch
        status_y = card_height - 0.42 * inch
        
        c.setFillColor(status_color)
        c.roundRect(status_x - 0.03 * inch, status_y - 0.01 * inch, 
                   status_width + 0.06 * inch, 0.08 * inch, 0.02 * inch, fill=1)
        
        c.setFillColor(colors.white)
        c.drawString(status_x, status_y + 0.005 * inch, status)
        
        # ============================================================================
        # MAIN CONTENT AREA - Left side with proper vertical spacing
        # ============================================================================
        content_x = 0.2 * inch
        content_start_y = card_height - header_height - 0.15 * inch
        content_y = content_start_y
        
        # Available width for left content (leaving space for QR code)
        content_width = 2.2 * inch
        
        # Affiliate name - main focal point
        c.setFillColor(text_primary)
        c.setFont("Helvetica-Bold", 11)
        affiliate_name = card.get_display_name()
        
        # Smart name truncation to fit available width
        if c.stringWidth(affiliate_name, "Helvetica-Bold", 11) > content_width:
            while len(affiliate_name) > 8 and c.stringWidth(affiliate_name + "...", "Helvetica-Bold", 11) > content_width:
                affiliate_name = affiliate_name[:-1]
            if len(affiliate_name) <= 8:
                affiliate_name = affiliate_name[:8]
            affiliate_name += "..."
        
        c.drawString(content_x, content_y, affiliate_name)
        content_y -= 0.18 * inch  # Proper spacing
        
        # Affiliation type
        affiliation_text = getattr(card, 'affiliation_type', 'Member').title()
        c.setFont("Helvetica-Bold", 8)
        c.setFillColor(accent_blue)
        c.drawString(content_x, content_y, affiliation_text)
        content_y -= 0.15 * inch
        
        # Council name
        council_text = getattr(card, 'council_name', '')
        if council_text:
            c.setFont("Helvetica", 7)
            c.setFillColor(text_secondary)
            
            # Truncate council name if needed
            if c.stringWidth(council_text, "Helvetica", 7) > content_width:
                while len(council_text) > 8 and c.stringWidth(council_text + "...", "Helvetica", 7) > content_width:
                    council_text = council_text[:-1]
                council_text += "..."
            
            c.drawString(content_x, content_y, council_text)
            content_y -= 0.12 * inch
        
        # Issue date
        if hasattr(card, 'date_issued') and card.date_issued:
            c.setFont("Helvetica", 6)
            c.setFillColor(text_muted)
            issued_text = f"Issued: {card.date_issued.strftime('%B %d, %Y')}"
            c.drawString(content_x, content_y, issued_text)
            content_y -= 0.1 * inch
        
        # Expiration date with color coding
        if hasattr(card, 'date_expires') and card.date_expires:
            c.setFont("Helvetica", 6)
            
            # Color code based on expiration status
            from datetime import datetime, date
            today = date.today()
            if card.date_expires < today:
                c.setFillColor(colors.HexColor('#ef4444'))  # Red for expired
            elif (card.date_expires - today).days < 90:
                c.setFillColor(accent_gold)  # Gold for expiring soon
            else:
                c.setFillColor(text_muted)  # Normal color
                
            expires_text = f"Valid Until: {card.date_expires.strftime('%B %Y')}"
            c.drawString(content_x, content_y, expires_text)
        
        # ============================================================================
        # QR CODE SECTION - Right side, properly positioned
        # ============================================================================
        
        # Generate functional verification URL
        verification_token = getattr(card, 'verification_token', None)
        if not verification_token:
            # Fallback token generation
            import uuid
            verification_token = str(uuid.uuid4()).replace('-', '')[:32]
        
        qr_verification_url = f"https://kreeck.com/card/verify/{verification_token}/"
        
        # QR code configuration
        qr = qrcode.QRCode(
            version=2,
            box_size=4,
            border=2,
            error_correction=qrcode.constants.ERROR_CORRECT_M
        )
        qr.add_data(qr_verification_url)
        qr.make(fit=True)
        
        # Generate QR code image
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        # QR code positioning - right side, centered vertically
        qr_size = 0.65 * inch
        qr_x = card_width - qr_size - 0.2 * inch
        qr_y = card_height - header_height - qr_size - 0.15 * inch  # Aligned with content
        
        # QR code background frame
        frame_padding = 6
        c.setFillColorRGB(0, 0, 0, alpha=0.1)  # Subtle shadow
        c.roundRect(qr_x - frame_padding + 1, qr_y - frame_padding + 1, 
                   qr_size + frame_padding * 2, qr_size + frame_padding * 2, 4, fill=1)
        
        # White frame
        c.setFillColor(colors.white)
        c.setStrokeColor(text_muted)
        c.setLineWidth(0.5)
        c.roundRect(qr_x - frame_padding, qr_y - frame_padding, 
                   qr_size + frame_padding * 2, qr_size + frame_padding * 2, 4, fill=1, stroke=1)
        
        # Draw QR code image
        c.drawInlineImage(qr_img, qr_x, qr_y, qr_size, qr_size)
        
        # QR code labels - positioned below QR code
        label_y = qr_y - 0.12 * inch
        
        c.setFont("Helvetica-Bold", 5)
        c.setFillColor(text_primary)
        main_label = "SCAN TO VERIFY"
        label_width = c.stringWidth(main_label, "Helvetica-Bold", 5)
        c.drawString(qr_x + (qr_size - label_width) / 2, label_y, main_label)
        
        c.setFont("Helvetica", 4)
        c.setFillColor(text_muted)
        sub_label = "or visit kreeck.com"
        sub_width = c.stringWidth(sub_label, "Helvetica", 4)
        c.drawString(qr_x + (qr_size - sub_width) / 2, label_y - 0.08 * inch, sub_label)
        
        # ============================================================================
        # FOOTER SECTION - Clean and minimal
        # ============================================================================
        footer_height = 0.15 * inch
        
        # Footer background
        c.setFillColorRGB(0.97, 0.98, 0.99)
        c.rect(0, 0, card_width, footer_height, fill=1)
        
        # Footer separator line
        c.setStrokeColor(colors.HexColor('#e5e7eb'))
        c.setLineWidth(0.5)
        c.line(0.15 * inch, footer_height, card_width - 0.15 * inch, footer_height)
        
        # Organization information (left)
        c.setFont("Helvetica-Bold", 5)
        c.setFillColor(text_secondary)
        c.drawString(0.15 * inch, 0.08 * inch, "ACRP South Africa")
        
        c.setFont("Helvetica", 4)
        c.setFillColor(text_muted)
        c.drawString(0.15 * inch, 0.04 * inch, "Professional Religious Practitioners")
        
        # Verification info (right)
        c.setFont("Helvetica", 4)
        verify_text = "verify: ams.acrp.org/affiliationcard/verify"
        verify_width = c.stringWidth(verify_text, "Helvetica", 4)
        c.drawString(card_width - verify_width - 0.15 * inch, 0.04 * inch, verify_text)
        
        # ============================================================================
        # BORDER AND FINISHING TOUCHES
        # ============================================================================
        
        # Outer border
        c.setStrokeColor(text_muted)
        c.setLineWidth(1)
        c.roundRect(1, 1, card_width - 2, card_height - 2, 6, fill=0, stroke=1)
        
        # Inner accent border
        c.setStrokeColorRGB(0.2, 0.4, 0.8, alpha=0.3)
        c.setLineWidth(0.5)
        c.roundRect(2, 2, card_width - 4, card_height - 4, 4, fill=0, stroke=1)
        
        # Finalize PDF
        c.save()
        
        # Extract PDF content
        pdf_content = buffer.getvalue()
        buffer.close()
        
        # Generate filename and content type
        filename = f"ACRP_Card_{card.card_number}.pdf"
        content_type = 'application/pdf'
        
        logger.info(f"Generated properly spaced PDF card for {card.card_number} with verification URL: {qr_verification_url}")
        return pdf_content, filename, content_type
        
    except Exception as e:
        logger.error(f"Failed to generate PDF card for {card.card_number}: {e}")
        raise






def generate_card_image(card, fmt):
    """
    Generate premium affiliation card image matching the professional PDF design.
    Modern navy/blue color scheme with sophisticated layout and typography.
    """
    # Choose resampling constant compatible across Pillow versions
    try:
        RESAMPLE = Image.Resampling.LANCZOS
    except AttributeError:
        RESAMPLE = getattr(Image, "LANCZOS", getattr(Image, "ANTIALIAS", Image.NEAREST))

    # Enhanced layout constants - optimized for modern design
    W, H = 1400, 880                        # High-resolution canvas
    CARD_W, CARD_H = 1200, 760              # Card dimensions
    CARD_X = (W - CARD_W) // 2
    CARD_Y = (H - CARD_H) // 2
    RADIUS = 28                             # Rounded corners
    MARGIN = 50

    # Professional color palette - matching PDF design
    primary_navy = (15, 23, 42)      # Deep navy for sophistication
    primary_blue = (30, 64, 175)     # Professional blue
    accent_blue = (59, 130, 246)     # Bright blue for highlights
    accent_emerald = (5, 150, 105)   # Emerald for active status
    accent_gold = (217, 119, 6)      # Gold for premium feel
    text_primary = (17, 24, 39)      # Almost black for readability
    text_secondary = (75, 85, 99)    # Medium gray
    text_muted = (156, 163, 175)     # Light gray
    text_white = (255, 255, 255)     # Pure white
    bg_light = (248, 250, 252)       # Premium light background
    card_bg = (255, 255, 255)        # White card background

    # Enhanced gradient helper
    def create_gradient(size, colors_list, direction='vertical'):
        """Create smooth multi-color gradient"""
        w, h = size
        if len(colors_list) < 2:
            return Image.new('RGB', (w, h), colors_list[0])
            
        # Create smooth transition between colors
        gradient = Image.new('RGB', (w, h))
        pixels = []
        
        for y in range(h):
            for x in range(w):
                if direction == 'vertical':
                    ratio = y / (h - 1) if h > 1 else 0
                else:  # horizontal
                    ratio = x / (w - 1) if w > 1 else 0
                
                # Interpolate between first and last color
                c1, c2 = colors_list[0], colors_list[-1]
                r = int(c1[0] + (c2[0] - c1[0]) * ratio)
                g = int(c1[1] + (c2[1] - c1[1]) * ratio)
                b = int(c1[2] + (c2[2] - c1[2]) * ratio)
                pixels.append((r, g, b))
        
        gradient.putdata(pixels)
        return gradient

    # Canvas with premium background
    canvas = Image.new('RGB', (W, H), bg_light)
    
    # Enhanced shadow with multiple layers for depth
    shadow_layers = [
        (CARD_W + 16, CARD_H + 16, 8, (0, 0, 0, 60)),   # Close shadow
        (CARD_W + 32, CARD_H + 32, 16, (0, 0, 0, 30)),  # Medium shadow
        (CARD_W + 48, CARD_H + 48, 24, (0, 0, 0, 15)),  # Far shadow
    ]
    
    for shadow_w, shadow_h, blur, color in shadow_layers:
        shadow = Image.new('RGBA', (shadow_w, shadow_h), color)
        shadow_mask = Image.new('L', (shadow_w, shadow_h), 0)
        sd = ImageDraw.Draw(shadow_mask)
        sd.rounded_rectangle([0, 0, shadow_w, shadow_h], radius=RADIUS + 8, fill=255)
        shadow.putalpha(shadow_mask)
        shadow = shadow.filter(ImageFilter.GaussianBlur(blur))
        
        offset_x = CARD_X - (shadow_w - CARD_W) // 2
        offset_y = CARD_Y - (shadow_h - CARD_H) // 2 + 6
        canvas.paste(shadow, (offset_x, offset_y), shadow)

    # Create sophisticated card background
    card_img = Image.new('RGBA', (CARD_W, CARD_H), card_bg)
    
    # Header section with navy background
    header_h = 200
    header_bg = Image.new('RGB', (CARD_W, header_h), primary_navy)
    
    # Create diagonal accent stripe
    stripe_img = Image.new('RGBA', (CARD_W, header_h), (0, 0, 0, 0))
    stripe_draw = ImageDraw.Draw(stripe_img)
    
    # Diagonal stripe coordinates - matching PDF design
    stripe_points = [
        (0, header_h - 60),
        (int(CARD_W * 0.4), header_h - 60),
        (int(CARD_W * 0.45), header_h),
        (0, header_h)
    ]
    stripe_draw.polygon(stripe_points, fill=accent_blue)
    
    # Combine header elements
    header_bg.paste(stripe_img, (0, 0), stripe_img)
    card_img.paste(header_bg, (0, 0))

    # Card mask with rounded corners
    mask = Image.new('L', (CARD_W, CARD_H), 0)
    mdraw = ImageDraw.Draw(mask)
    mdraw.rounded_rectangle([0, 0, CARD_W, CARD_H], radius=RADIUS, fill=255)

    # Apply card to canvas
    canvas.paste(card_img.convert('RGB'), (CARD_X, CARD_Y), mask)
    draw = ImageDraw.Draw(canvas)

    # Enhanced font loading
    def load_font(name=None, size=36):
        font_paths = [
            '/System/Library/Fonts/Helvetica.ttc',
            '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
            '/Windows/Fonts/arial.ttf',
            'arial.ttf'
        ]
        
        for path in font_paths:
            try:
                return ImageFont.truetype(path, size)
            except:
                continue
        return ImageFont.load_default()

    # Font sizes optimized for the design
    font_logo = load_font(size=42)           # ACRP logo
    font_org = load_font(size=16)            # Organization name
    font_badge = load_font(size=12)          # Badge text
    font_name = load_font(size=38)           # Affiliate name
    font_title = load_font(size=24)          # Affiliation type
    font_meta = load_font(size=18)           # Metadata
    font_small = load_font(size=14)          # Small text
    font_tiny = load_font(size=12)           # Footer text

    # ============================================================================
    # HEADER SECTION - Matching PDF layout
    # ============================================================================
    header_y_start = CARD_Y + 30
    
    # ACRP Logo - positioned like PDF
    draw.text((CARD_X + MARGIN, header_y_start), "ACRP", font=font_logo, fill=text_white)
    
    # Organization name - smaller, positioned below
    org_y = header_y_start + 50
    draw.text((CARD_X + MARGIN, org_y), "ASSOCIATION OF CHRISTIAN", font=font_org, fill=text_white)
    draw.text((CARD_X + MARGIN, org_y + 20), "RELIGIOUS PRACTITIONERS", font=font_org, fill=text_white)
    
    # Digital card badge - matching PDF style
    badge_x = CARD_X + MARGIN
    badge_y = header_y_start + 100
    badge_width = 240
    badge_height = 28
    
    draw.rounded_rectangle([
        badge_x, badge_y,
        badge_x + badge_width, badge_y + badge_height
    ], radius=8, fill=accent_emerald)
    
    draw.text((badge_x + 8, badge_y + 6), "DIGITAL AFFILIATION CARD", 
              font=font_badge, fill=text_white)
    
    # Card number - top right
    card_num = f"#{card.card_number}"
    try:
        bbox = draw.textbbox((0, 0), card_num, font=font_meta)
        card_num_width = bbox[2] - bbox[0]
    except AttributeError:
        card_num_width = len(card_num) * 12
    
    draw.text((CARD_X + CARD_W - MARGIN - card_num_width, header_y_start + 5), 
              card_num, font=font_meta, fill=text_white)
    
    # Status badge - positioned below card number
    status = card.get_status_display().upper()
    
    # Status color logic - matching PDF
    if hasattr(card, 'status') and card.status == 'active':
        status_color = accent_emerald
    elif hasattr(card, 'status') and card.status == 'pending':
        status_color = accent_gold
    else:
        status_color = (239, 68, 68)  # Red for inactive
    
    try:
        status_bbox = draw.textbbox((0, 0), status, font=font_small)
        status_width = status_bbox[2] - status_bbox[0]
    except AttributeError:
        status_width = len(status) * 8
    
    # Status badge positioning
    status_badge_x = CARD_X + CARD_W - MARGIN - status_width - 20
    status_badge_y = header_y_start + 35
    
    draw.rounded_rectangle([
        status_badge_x, status_badge_y,
        status_badge_x + status_width + 20, status_badge_y + 22
    ], radius=6, fill=status_color)
    
    draw.text((status_badge_x + 10, status_badge_y + 4), status, 
              font=font_small, fill=text_white)

    # ============================================================================
    # MAIN CONTENT AREA - Left side text, right side QR
    # ============================================================================
    content_y = CARD_Y + header_h + 40
    content_x = CARD_X + MARGIN
    content_max_width = CARD_W - (MARGIN * 2) - 200  # Leave space for QR code
    
    # Affiliate name - prominent display
    name = card.get_display_name() if hasattr(card, 'get_display_name') else str(getattr(card, 'affiliate_full_name', ''))
    
    # Smart name truncation
    if len(name) > 25:
        name = name[:22] + "..."
    
    draw.text((content_x, content_y), name, font=font_name, fill=text_primary)
    content_y += 55
    
    # Affiliation type with accent color
    affiliation = getattr(card, 'affiliation_type', 'Member').title()
    draw.text((content_x, content_y), affiliation, font=font_title, fill=accent_blue)
    content_y += 35
    
    # Council information
    council = getattr(card, 'council_name', '')
    if council:
        # Truncate if too long
        if len(council) > 35:
            council = council[:32] + "..."
        draw.text((content_x, content_y), council, font=font_meta, fill=text_secondary)
        content_y += 28
    
    # Date information
    if hasattr(card, 'date_issued') and card.date_issued:
        issued_text = f"Issued: {card.date_issued.strftime('%B %d, %Y')}"
        draw.text((content_x, content_y), issued_text, font=font_small, fill=text_muted)
        content_y += 22
    
    # Expiration with color coding
    if hasattr(card, 'date_expires') and card.date_expires:
        from datetime import datetime, date
        today = date.today()
        
        if card.date_expires < today:
            expire_color = (239, 68, 68)  # Red for expired
        elif (card.date_expires - today).days < 90:
            expire_color = accent_gold  # Gold for expiring soon
        else:
            expire_color = text_muted  # Normal color
        
        expires_text = f"Valid Until: {card.date_expires.strftime('%B %Y')}"
        draw.text((content_x, content_y), expires_text, font=font_small, fill=expire_color)

    # ============================================================================
    # QR CODE SECTION - Right side, matching PDF design
    # ============================================================================
    
    # Generate functional verification URL - CORRECTED URL FORMAT
    verification_token = getattr(card, 'verification_token', None)
    if not verification_token:
        import uuid
        verification_token = str(uuid.uuid4()).replace('-', '')[:32]
    
    # Use the correct verification URL format
    qr_verification_url = f"https://kreeck.com/card/verify/{verification_token}/"
    
    logger.info(f"Generated QR code URL for image card {card.card_number}: {qr_verification_url}")
    
    # QR code configuration
    qr = qrcode.QRCode(
        version=2,
        box_size=6,
        border=3,
        error_correction=qrcode.constants.ERROR_CORRECT_M
    )
    qr.add_data(qr_verification_url)
    qr.make(fit=True)
    
    qr_img = qr.make_image(fill_color="black", back_color="white").convert('RGBA')
    qr_size = 180
    qr_img = qr_img.resize((qr_size, qr_size), RESAMPLE)

    # QR positioning - right side, vertically centered
    qr_x = CARD_X + CARD_W - qr_size - MARGIN - 20
    qr_y = CARD_Y + header_h + 60

    # QR background frame with shadow - matching PDF style
    qr_padding = 12
    
    # Subtle shadow
    shadow_offset = 2
    draw.rounded_rectangle([
        qr_x - qr_padding + shadow_offset, qr_y - qr_padding + shadow_offset,
        qr_x + qr_size + qr_padding + shadow_offset, qr_y + qr_size + qr_padding + shadow_offset
    ], radius=12, fill=(0, 0, 0, 25))
    
    # White frame
    draw.rounded_rectangle([
        qr_x - qr_padding, qr_y - qr_padding,
        qr_x + qr_size + qr_padding, qr_y + qr_size + qr_padding
    ], radius=12, fill=card_bg, outline=text_muted, width=1)
    
    # Paste QR code
    canvas.paste(qr_img, (qr_x, qr_y), qr_img)

    # QR labels - matching PDF style
    qr_label_y = qr_y + qr_size + 15
    
    main_label = "SCAN TO VERIFY"
    try:
        label_bbox = draw.textbbox((0, 0), main_label, font=font_small)
        label_width = label_bbox[2] - label_bbox[0]
    except AttributeError:
        label_width = len(main_label) * 8
    
    label_x = qr_x + (qr_size - label_width) // 2
    draw.text((label_x, qr_label_y), main_label, font=font_small, fill=text_primary)
    
    # Secondary label
    sub_label = "or visit kreeck.com"
    try:
        sub_bbox = draw.textbbox((0, 0), sub_label, font=font_tiny)
        sub_width = sub_bbox[2] - sub_bbox[0]
    except AttributeError:
        sub_width = len(sub_label) * 6
    
    sub_x = qr_x + (qr_size - sub_width) // 2
    draw.text((sub_x, qr_label_y + 20), sub_label, font=font_tiny, fill=text_muted)

    # ============================================================================
    # FOOTER SECTION - Clean and professional
    # ============================================================================
    footer_y = CARD_Y + CARD_H - 60
    
    # Footer background
    footer_bg = Image.new('RGBA', (CARD_W, 50), (247, 248, 249, 255))
    canvas.paste(footer_bg, (CARD_X, footer_y), footer_bg)
    
    # Footer separator line
    draw.line([(CARD_X + 30, footer_y), (CARD_X + CARD_W - 30, footer_y)], 
              fill=text_muted, width=1)
    
    # Organization info (left side)
    draw.text((CARD_X + MARGIN, footer_y + 12), "ACRP South Africa", 
              font=font_small, fill=text_secondary)
    draw.text((CARD_X + MARGIN, footer_y + 28), "Professional Religious Practitioners", 
              font=font_tiny, fill=text_muted)
    
    # Verification info (right side)
    verify_text = "verify: kreeck.com/card/verify"
    try:
        verify_bbox = draw.textbbox((0, 0), verify_text, font=font_tiny)
        verify_width = verify_bbox[2] - verify_bbox[0]
    except AttributeError:
        verify_width = len(verify_text) * 6
    
    draw.text((CARD_X + CARD_W - MARGIN - verify_width, footer_y + 28), 
              verify_text, font=font_tiny, fill=text_muted)

    # ============================================================================
    # FINAL TOUCHES - Borders and finishing
    # ============================================================================
    
    # Outer border
    draw.rounded_rectangle([CARD_X + 1, CARD_Y + 1, CARD_X + CARD_W - 1, CARD_Y + CARD_H - 1], 
                          radius=RADIUS, outline=text_muted, width=2)
    
    # Inner accent border
    draw.rounded_rectangle([CARD_X + 3, CARD_Y + 3, CARD_X + CARD_W - 3, CARD_Y + CARD_H - 3], 
                          radius=RADIUS - 2, outline=accent_blue + (100,), width=1)

    # Final image preparation
    out_img = canvas
    if fmt.upper() == 'JPEG' and out_img.mode != 'RGB':
        out_img = out_img.convert('RGB')

    # Save with high quality
    buffer = BytesIO()
    quality = 95 if fmt.upper() == 'JPEG' else None
    out_img.save(buffer, format=fmt, quality=quality, optimize=True)
    buffer.seek(0)
    data = buffer.getvalue()
    buffer.close()

    ext = 'png' if fmt.upper() == 'PNG' else 'jpg'
    filename = f"ACRP_Card_{card.card_number}.{ext}"
    content_type = 'image/png' if ext == 'png' else 'image/jpeg'
    
    logger.info(f"Generated premium {fmt} card image for {card.card_number} matching PDF design")
    return data, filename, content_type




def generate_card_file(card, file_format):
    """Generate card file in specified format with enhanced styling."""
    fmt = file_format.lower()
    if fmt == 'pdf':
        return generate_card_pdf(card)
    elif fmt == 'png':
        return generate_card_image(card, 'PNG')
    elif fmt in ('jpg', 'jpeg'):
        return generate_card_image(card, 'JPEG')
    else:
        raise ValueError(f"Unsupported file format: {file_format}")
    



















# Bulk operation functions
def bulk_assign_cards(cards, user, reason):
    """Bulk assign cards."""
    count = 0
    for card in cards:
        if card.status == 'pending_assignment':
            card.assign_card(assigned_by=user)
            count += 1
    
    return {'message': f'{count} cards assigned successfully'}


def bulk_issue_cards(cards, user, reason):
    """Bulk issue cards."""
    count = 0
    for card in cards:
        if card.status == 'assigned':
            card.issue_card(issued_by=user)
            count += 1
    
    return {'message': f'{count} cards issued successfully'}


def bulk_suspend_cards(cards, user, reason):
    """Bulk suspend cards."""
    count = 0
    for card in cards:
        if card.status == 'active':
            card.suspend_card(reason=reason)
            count += 1
    
    return {'message': f'{count} cards suspended successfully'}





def create_custom_report(form_data):
    """Create custom report based on form data."""
    # Implementation for custom report generation
    return {}



@never_cache
def download_card(request, token):
    """Secure card download using token. Robust binding + debug logging."""
    delivery = get_object_or_404(CardDelivery, download_token=token)

    # quick debug log: incoming request method and token
    logger.info("download_card called (method=%s) for token=%s (delivery_id=%s)",
                request.method, token, getattr(delivery, 'id', 'N/A'))

    # check expiry / validity first
    if not delivery.is_download_valid():
        logger.info("download_card: delivery not valid (expired or limit) for delivery_id=%s", delivery.id)
        return render(request, 'affiliationcard/public/download_expired.html', {
            'delivery': delivery,
            'page_title': 'Download Expired'
        })

    if request.method == 'POST':
        # Make a copy of POST and ensure the download_token is present in bound data
        post = request.POST.copy()
        # Put token into POST (defensive: template already supplies, but ensure it)
        post['download_token'] = token

        form = CardDownloadForm(post)

        logger.debug("CardDownloadForm bound: %s", form.is_bound)
        if not form.is_valid():
            logger.warning("CardDownloadForm invalid for delivery_id=%s: %s", delivery.id, form.errors)
            # show errors to user and re-render
            for field, errs in form.errors.items():
                for err in errs:
                    messages.error(request, f"{field}: {err}")
        else:
            file_format = form.cleaned_data.get('file_format', 'pdf')
            logger.info("CardDownloadForm valid: preparing file (format=%s) for delivery_id=%s", file_format, delivery.id)

            try:
                # generate_card_file must return bytes (or file-like content), filename, content_type
                file_content, filename, content_type = generate_card_file(delivery.card, file_format)

                # record the download in DB
                try:
                    delivery.record_download()
                except Exception:
                    logger.exception("Failed to record download for delivery_id=%s", delivery.id)

                # send file response
                response = HttpResponse(file_content, content_type=content_type)
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                response['Content-Length'] = str(len(file_content)) if hasattr(file_content, '__len__') else None
                logger.info("Download served for delivery_id=%s filename=%s", delivery.id, filename)
                return response

            except Exception as e:
                logger.exception("Card generation error for delivery_id=%s: %s", delivery.id, e)
                messages.error(request, "Failed to generate card file. Check server logs for details.")

    else:
        form = CardDownloadForm(initial={'download_token': token, 'file_format': 'pdf'})

    context = {
        'form': form,
        'delivery': delivery,
        'card': delivery.card,
        'page_title': 'Download Your Card'
    }

    return render(request, 'affiliationcard/public/download_card.html', context)
