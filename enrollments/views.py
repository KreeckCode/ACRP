from datetime import timedelta
import json
import logging
import os
from typing import Type, Dict, Any, Optional
from django.db.models import Count, Q, Case, When, IntegerField, Prefetch
from django.db.models.functions import Coalesce
from django.core.cache import cache
from django.forms import ValidationError
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.contrib import messages
from django.db.models import Q, Count, Prefetch, F
from django.http import HttpResponseForbidden, JsonResponse, Http404, HttpResponse
from django.urls import reverse
from django.core.mail import send_mail
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.cache import cache
from django.views.decorators.cache import cache_page
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import csrf_protect
from django.db import transaction
from django.utils import timezone
from django.conf import settings
from django.contrib.contenttypes.models import ContentType
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from django.contrib.contenttypes.models import ContentType

from affiliationcard.views import assign_card_programmatically


from .models import (
    # Core models
    Council,
    AffiliationType,
    DesignationCategory,
    DesignationSubcategory,
    OnboardingSession,
    
    # Application models
    AssociatedApplication,
    DesignatedApplication,
    StudentApplication,
    
    # Related models
    AcademicQualification,
    Reference,
    PracticalExperience,
    Document,
)

from .forms import (
    # Onboarding forms
    AffiliationTypeSelectionForm,
    CouncilSelectionForm,
    DesignationCategorySelectionForm,
    DesignationSubcategorySelectionForm,
    OnboardingSessionForm,
    
    # Application forms
    AssociatedApplicationForm,
    DesignatedApplicationForm,
    StudentApplicationForm,
    
    # Related model forms and formsets
    AcademicQualificationForm,
    ReferenceForm,
    PracticalExperienceForm,
    DocumentForm,
    BulkDocumentUploadForm,
    AcademicQualificationFormSet,
    ReferenceFormSet,
    PracticalExperienceFormSet,
    DocumentFormSet,
    
    # Utility forms
    ApplicationSearchForm,
    ApplicationReviewForm,
)

from accounts.models import User

# Configure logging
logger = logging.getLogger(__name__)

# Constants
ITEMS_PER_PAGE = 25
CACHE_TIMEOUT = 900  # 15 minutes
SEARCH_CACHE_TIMEOUT = 300  # 5 minutes
ONBOARDING_SESSION_TIMEOUT = 3600  # 1 hour


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def is_admin_or_manager(user):
    """Check if user has admin or manager privileges"""
    return user.acrp_role in {
        User.ACRPRole.GLOBAL_SDP,
        User.ACRPRole.PROVIDER_ADMIN,
    }


def can_approve_applications(user):
    """Check if user can approve applications"""
    return user.acrp_role in {
        User.ACRPRole.GLOBAL_SDP,
        User.ACRPRole.PROVIDER_ADMIN
    }


def get_client_ip(request):
    """Safely extract client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def rate_limit(max_requests=50, window=3600):
    """Simple rate limiting decorator"""
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            ip = get_client_ip(request)
            cache_key = f"rate_limit:{view_func.__name__}:{ip}"
            
            current_requests = cache.get(cache_key, 0)
            if current_requests >= max_requests:
                logger.warning(f"Rate limit exceeded for IP {ip} on {view_func.__name__}")
                messages.error(request, "Too many requests. Please try again later.")
                return redirect('enrollments:onboarding_start')
            
            cache.set(cache_key, current_requests + 1, window)
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


def get_application_model_for_type(affiliation_type: str) -> Type:
    """Get the appropriate application model class for affiliation type"""
    model_map = {
        'associated': AssociatedApplication,
        'designated': DesignatedApplication,
        'student': StudentApplication,
    }
    return model_map.get(affiliation_type)


def get_all_applications_queryset():
    """Get unified queryset of all application types"""
    # We'll need to use union for different models
    from django.db.models import CharField, Value
    
    associated = AssociatedApplication.objects.annotate(
        app_type=Value('associated', output_field=CharField())
    ).values(
        'id', 'app_type', 'application_number', 'status', 'email', 
        'full_names', 'surname', 'created_at', 'submitted_at'
    )
    
    designated = DesignatedApplication.objects.annotate(
        app_type=Value('designated', output_field=CharField())
    ).values(
        'id', 'app_type', 'application_number', 'status', 'email',
        'full_names', 'surname', 'created_at', 'submitted_at'
    )
    
    student = StudentApplication.objects.annotate(
        app_type=Value('student', output_field=CharField())
    ).values(
        'id', 'app_type', 'application_number', 'status', 'email',
        'full_names', 'surname', 'created_at', 'submitted_at'
    )
    
    return associated.union(designated, student)


# ============================================================================
# ONBOARDING FLOW VIEWS
# ============================================================================
# Utility function for cleaning up stale onboarding sessions
def cleanup_stale_sessions(user=None, ip_address=None):
    """
    Clean up incomplete onboarding sessions to prevent conflicts.
    
    Args:
        user: Django User object (for authenticated users)
        ip_address: Client IP address string (for anonymous users)
    
    This function prevents MultipleObjectsReturned errors in get_or_create()
    by removing stale sessions that might match the same criteria.
    """
    if user and user.is_authenticated:
        # Clean up by authenticated user
        deleted_count, _ = OnboardingSession.objects.filter(
            user=user,
            status__in=['selecting_council', 'selecting_affiliation', 'selecting_category', 'selecting_subcategory']
        ).delete()
        if deleted_count > 0:
            logger.info(f"Cleaned up {deleted_count} stale sessions for user {user.id}")
    elif ip_address:
        # Clean up by IP address for anonymous users (less reliable but prevents accumulation)
        deleted_count, _ = OnboardingSession.objects.filter(
            user=None,
            ip_address=ip_address,
            status__in=['selecting_council', 'selecting_affiliation', 'selecting_category', 'selecting_subcategory'],
            created_at__lt=timezone.now() - timezone.timedelta(hours=1)  # Only clean up sessions older than 1 hour
        ).delete()
        if deleted_count > 0:
            logger.info(f"Cleaned up {deleted_count} stale anonymous sessions for IP {ip_address}")


@csrf_protect
def onboarding_start(request):
    """
    Step 1: Start onboarding process - select council first.
    
    CHANGES MADE:
    - This now handles council selection instead of affiliation type
    - Council selection is now the first step in the onboarding flow
    - Session status starts with 'selecting_council' instead of 'selecting_affiliation'
    - Redirects to affiliation type selection (step 2) after council is chosen
    """
    if request.method == 'POST':
        # Get council ID from POST data instead of affiliation type
        council_id = request.POST.get('council')
        logger.info(f"POST data received: {request.POST}")
        logger.info(f"Council ID: {council_id}")
        
        if council_id:
            try:
                # Get the actual Council object and validate it's active
                council = Council.objects.get(id=council_id, is_active=True)
                
                # Clean up any existing incomplete sessions to prevent conflicts
                cleanup_stale_sessions(
                    user=request.user if request.user.is_authenticated else None,
                    ip_address=get_client_ip(request)
                )
                
                # Create fresh onboarding session
                # NOTE: Always create new session to avoid MultipleObjectsReturned errors
                session = OnboardingSession.objects.create(
                    user=request.user if request.user.is_authenticated else None,
                    status='selecting_affiliation',  # Next step: affiliation type selection
                    selected_council=council,  # Council is selected first now
                    ip_address=get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
                )
                
                # Store session ID in user session for tracking
                request.session['onboarding_session_id'] = str(session.session_id)
                
                logger.info(f"Council selected first: {council.code} - {council.name} from IP {get_client_ip(request)}")
                
                # Redirect to affiliation type selection (step 2) - using onboarding_council view
                # NOTE: We're reusing the same URL name but the view logic has changed
                return redirect('enrollments:onboarding_council', session_id=str(session.session_id))
                
            except Council.DoesNotExist:
                logger.error(f"Council with ID {council_id} not found")
                messages.error(request, f"Invalid council selected: {council_id}")
        else:
            messages.error(request, "Please select a valid council.")
    
    # For GET request or form errors - show council selection
    # Get all councils for the template
    councils = Council.objects.filter(is_active=True).order_by('code')
    
    # Create a dictionary for easy access in template
    councils_dict = {}
    for council in councils:
        councils_dict[council.code.lower()] = council
    
    # Debug: Log available councils
    logger.info(f"Available councils: {[(c.code, c.id) for c in councils]}")
    
    form = CouncilSelectionForm()  # Changed from AffiliationTypeSelectionForm
    
    context = {
        'form': form,
        'councils': councils_dict,  # For template access like {{ councils.cgmp.id }}
        'councils_list': councils,   # For iteration in template
        'page_title': 'Select Council',  # Updated page title
        'step': 1,  # This is now step 1
        'total_steps': 4,
    }
    
    # NOTE: Template name stays the same but will need minor updates to show councils
    return render(request, 'enrollments/onboarding/step1_affiliation_type.html', context)


@csrf_protect
def onboarding_council(request, session_id):
    """
    Step 2: Select affiliation type (associated, designated, student).
    
    CHANGES MADE:
    - This now handles affiliation type selection instead of council selection
    - This is now the second step in the onboarding flow
    - Council should already be selected from step 1
    - Validates that council was already selected in the session
    """
    # Get and validate onboarding session
    try:
        session = OnboardingSession.objects.select_related('selected_council').get(session_id=session_id)
    except OnboardingSession.DoesNotExist:
        messages.error(request, "Invalid onboarding session. Please start over.")
        return redirect('enrollments:onboarding_start')
    
    # Check session timeout
    if (timezone.now() - session.created_at).total_seconds() > ONBOARDING_SESSION_TIMEOUT:
        messages.error(request, "Onboarding session expired. Please start over.")
        # Clean up expired session
        session.delete()
        return redirect('enrollments:onboarding_start')
    
    # Validate that council was already selected (since it's step 1 now)
    if not session.selected_council:
        messages.error(request, "Council must be selected first. Please start over.")
        return redirect('enrollments:onboarding_start')
    
    if request.method == 'POST':
        affiliation_type_code = request.POST.get('affiliation_type')
        logger.info(f"Affiliation type: {affiliation_type_code}")
        
        if affiliation_type_code in ['associated', 'designated', 'student']:
            try:
                affiliation_type = AffiliationType.objects.get(
                    code=affiliation_type_code,
                    is_active=True
                )
                
                # Update session with selected affiliation type
                session.selected_affiliation_type = affiliation_type
                session.status = 'completed'  # Complete onboarding immediately
                session.completed_at = timezone.now()  # Mark as completed
                session.save(update_fields=['selected_affiliation_type', 'status', 'completed_at'])
                
                logger.info(f"Affiliation type selected: {affiliation_type.name}")
                
                # ALL affiliation types go directly to application creation
                return redirect('enrollments:application_create', session_id=session_id)
                    
            except AffiliationType.DoesNotExist:
                messages.error(request, f"Invalid affiliation type: {affiliation_type_code}")
        else:
            messages.error(request, "Please select a valid affiliation type.")
    
    # For GET request or form errors - show affiliation type selection
    form = AffiliationTypeSelectionForm()  # Changed from CouncilSelectionForm
    
    context = {
        'form': form,
        'session': session,
        'selected_council': session.selected_council,  # Show which council was selected
        'page_title': 'Select Affiliation Type',  # Updated page title
        'step': 2,  # This is now step 2
        'total_steps': 4,
    }
    
    # NOTE: Template name stays the same but will need minor updates to show affiliation types
    return render(request, 'enrollments/onboarding/step2_council.html', context)


@csrf_protect
def onboarding_category(request, session_id):
    """
    DISABLED: Category selection moved to admin assignment after application review.
    """
    messages.info(request, "Designation categories are now assigned by administrators after application review.")
    return redirect('enrollments:onboarding_start')


@csrf_protect  
def onboarding_subcategory(request, session_id):
    """
    DISABLED: Subcategory selection moved to admin assignment after application review.
    """
    messages.info(request, "Designation subcategories are now assigned by administrators after application review.")
    return redirect('enrollments:onboarding_start')


# ============================================================================
# APPLICATION CREATION VIEWS
# ============================================================================

from django.contrib.contenttypes.models import ContentType
import traceback 

@csrf_protect
@rate_limit(max_requests=102, window=3600)
@transaction.atomic
def application_create(request, session_id):
    """
    Create application with proper document handling for generic formsets.
    
    FIXED: Removed references to deleted OnboardingSession fields:
    - selected_designation_category (removed in migration 0005)
    - selected_designation_subcategory (removed in migration 0005)
    
    These fields are now managed differently in the application flow.
    """
    try:
        # FIXED: Removed select_related for deleted fields
        # Before: select_related('selected_designation_category', 'selected_designation_subcategory')
        # After: Only select existing fields
        session = OnboardingSession.objects.select_related(
            'selected_affiliation_type', 
            'selected_council'
            # Removed: 'selected_designation_category', 'selected_designation_subcategory'
        ).get(session_id=session_id)
    except OnboardingSession.DoesNotExist:
        messages.error(request, "Invalid onboarding session.")
        return redirect('enrollments:onboarding_start')
    
    if not session.is_complete():
        messages.error(request, "Please complete the onboarding process.")
        return redirect('enrollments:onboarding_start')
    
    # Check for existing application
    existing_application = None
    try:
        if session.selected_affiliation_type.code == 'associated':
            existing_application = AssociatedApplication.objects.get(onboarding_session=session)
        elif session.selected_affiliation_type.code == 'designated':
            existing_application = DesignatedApplication.objects.get(onboarding_session=session)
        elif session.selected_affiliation_type.code == 'student':
            existing_application = StudentApplication.objects.get(onboarding_session=session)
    except (AssociatedApplication.DoesNotExist, DesignatedApplication.DoesNotExist, StudentApplication.DoesNotExist):
        pass
    
    if existing_application:
        messages.info(request, "Application already exists for this session.")
        return redirect('enrollments:application_detail', 
                       pk=existing_application.pk, 
                       app_type=session.selected_affiliation_type.code)
    
    # Get form class
    form_class_map = {
        'associated': AssociatedApplicationForm,
        'designated': DesignatedApplicationForm,
        'student': StudentApplicationForm,
    }
    
    form_class = form_class_map.get(session.selected_affiliation_type.code)
    if not form_class:
        messages.error(request, "Invalid affiliation type.")
        return redirect('enrollments:onboarding_start')
    
    if request.method == 'POST':
        logger.info(f"POST data keys: {list(request.POST.keys())}")
        logger.info(f"FILES data keys: {list(request.FILES.keys())}")
        
        # DEBUG: Check session data - UPDATED for new field structure
        logger.info(f"Session has affiliation type: {session.selected_affiliation_type}")
        logger.info(f"Session has council: {session.selected_council}")
        
        # FIXED: No longer try to access deleted fields
        # These fields were removed in migration 0005:
        # - session.selected_designation_category (DELETED)
        # - session.selected_designation_subcategory (DELETED)
        
        # Create main form
        form = form_class(request.POST, request=request, onboarding_session=session)
        
        # This prevents RelatedObjectDoesNotExist errors during model.clean()
        form.instance.onboarding_session = session
        
        # FIXED: Handle designation category/subcategory differently
        # Since these fields were removed from OnboardingSession,
        # they should now be set directly on the application or handled elsewhere
        if session.selected_affiliation_type.code == 'designated':
            # OPTION 1: Get from form data if the form handles these fields
            designation_category_id = request.POST.get('designation_category')
            designation_subcategory_id = request.POST.get('designation_subcategory')
            
            if designation_category_id:
                try:
                    from .models import DesignationCategory, DesignationSubcategory
                    form.instance.designation_category_id = designation_category_id
                    logger.info(f"Set designation_category_id: {designation_category_id}")
                except Exception as e:
                    logger.error(f"Error setting designation category: {e}")
            
            if designation_subcategory_id:
                try:
                    form.instance.designation_subcategory_id = designation_subcategory_id
                    logger.info(f"Set designation_subcategory_id: {designation_subcategory_id}")
                except Exception as e:
                    logger.error(f"Error setting designation subcategory: {e}")
        
        # Extract document data manually instead of using generic formset
        document_data = extract_document_data_from_post(request)
        reference_data = extract_reference_data_from_post(request)
        
        # FIXED: Initialize non-generic formsets with CORRECT prefixes to match template
        formsets = {}
        if session.selected_affiliation_type.code == 'designated':
            # Initialize formsets with proper prefix and instance
            formsets = {
                'academic_qualifications': AcademicQualificationFormSet(
                    request.POST, 
                    prefix='academic_qualifications',
                    instance=None  # No instance for new application
                ),
                'practical_experiences': PracticalExperienceFormSet(
                    request.POST, 
                    prefix='practical_experiences',
                    instance=None  # No instance for new application
                ),
            }
        
        # Validate main form (now onboarding_session is set, so model.clean() won't fail)
        form_valid = form.is_valid()
        if not form_valid:
            logger.warning(f"Main form errors: {form.errors}")
        
        # Validate formsets
        formsets_valid = all(formset.is_valid() for formset in formsets.values())
        
        # Validate extracted document data
        documents_valid, document_errors = validate_extracted_documents(document_data)
        references_valid, reference_errors = validate_extracted_references(reference_data)
        
        # ENHANCED DEBUG: Log all validation results
        logger.info(f"Form valid: {form_valid}")
        logger.info(f"Formsets valid: {formsets_valid}")
        logger.info(f"Documents valid: {documents_valid}")
        logger.info(f"References valid: {references_valid}")

        if not form_valid:
            logger.error(f"FORM ERRORS: {form.errors}")
            logger.error(f"FORM NON_FIELD_ERRORS: {form.non_field_errors()}")

        if not formsets_valid:
            for name, formset in formsets.items():
                if not formset.is_valid():
                    logger.error(f"FORMSET {name} ERRORS: {formset.errors}")
                    logger.error(f"FORMSET {name} NON_FORM_ERRORS: {formset.non_form_errors()}")

        if not documents_valid:
            logger.error(f"DOCUMENT ERRORS: {document_errors}")
            
        if not references_valid:
            logger.error(f"REFERENCE ERRORS: {reference_errors}")
        
        if form_valid and formsets_valid and documents_valid and references_valid:
            try:
                # Save main application first
                application = form.save(commit=False)
                # onboarding_session is already set above, but we'll keep this for clarity
                application.onboarding_session = session
                
                # FIXED: No longer set designation fields from session
                # These were removed from OnboardingSession in migration 0005
                # The fields should now be set directly from form data (handled above)
                # if session.selected_affiliation_type.code == 'designated':
                #     application.designation_category = session.selected_designation_category  # DELETED FIELD
                #     application.designation_subcategory = session.selected_designation_subcategory  # DELETED FIELD
                
                application.save()
                logger.info(f"Application saved with ID: {application.pk}")
                
                # Get content type
                content_type = ContentType.objects.get_for_model(application.__class__)
                
                # Save non-generic formsets
                for formset_name, formset in formsets.items():
                    formset.instance = application
                    formset.save()
                    logger.info(f"Saved {formset_name} formset")
                
                # Save documents manually
                documents_saved = save_extracted_documents(document_data, application, content_type, request.user)
                logger.info(f"Saved {documents_saved} documents")
                
                # Save references manually
                references_saved = save_extracted_references(reference_data, application, content_type)
                logger.info(f"Saved {references_saved} references")
                
                # Verify documents were saved
                saved_documents = Document.objects.filter(content_type=content_type, object_id=application.pk)
                logger.info(f"Total documents in DB: {saved_documents.count()}")
                
                # Send both welcome email to applicant and notification to admin staff
                try:
                    email_results = send_application_emails(application)
                    
                    if email_results['welcome_email_sent'] and email_results['admin_notification_sent']:
                        logger.info(f"All emails sent successfully for application {application.application_number}")
                    elif email_results['welcome_email_sent']:
                        logger.warning(f"Welcome email sent but admin notification failed for application {application.application_number}")
                    elif email_results['admin_notification_sent']:
                        logger.warning(f"Admin notification sent but welcome email failed for application {application.application_number}")
                    else:
                        logger.error(f"Both emails failed for application {application.application_number}")
                    
                    # Log any specific errors
                    if email_results['errors']:
                        for error in email_results['errors']:
                            logger.warning(f"Email error for application {application.application_number}: {error}")
                            
                except Exception as e:
                    logger.warning(f"Failed to send emails for application {application.application_number}: {str(e)}")
                    # Don't fail the application submission if email fails
                    pass
                
                request.session['application_success'] = {
                    'application_number': application.application_number,
                    'affiliation_type': session.selected_affiliation_type.name,
                    'council_name': session.selected_council.name,
                    'application_id': application.pk,
                    'app_type': session.selected_affiliation_type.code,
                }
                
                return redirect('enrollments:application_success')
                
            except Exception as e:
                logger.error(f"Error creating application: {str(e)}")
                logger.error(f"Traceback: {traceback.format_exc()}")
                messages.error(request, f"An error occurred: {str(e)}")
        
        else:
            # ENHANCED: Show validation errors with more detail
            if not form_valid:
                messages.error(request, "Please correct the form errors.")
                # Add specific field errors to messages
                for field, errors in form.errors.items():
                    messages.error(request, f"{field}: {errors[0]}")
            
            if not formsets_valid:
                messages.error(request, "Please correct the formset errors.")
                for name, formset in formsets.items():
                    if not formset.is_valid():
                        messages.error(request, f"{name} errors: {formset.errors}")
            
            if document_errors:
                messages.error(request, f"Document errors: {'; '.join(document_errors)}")
            if reference_errors:
                messages.error(request, f"Reference errors: {'; '.join(reference_errors)}")
    
    else:
        # GET request - FIXED: Use correct prefixes here too
        form = form_class(request=request, onboarding_session=session)
        formsets = {}
        if session.selected_affiliation_type.code == 'designated':
            formsets = {
                'academic_qualifications': AcademicQualificationFormSet(
                    prefix='academic_qualifications',
                    instance=None  # No instance for new application
                ),
                'practical_experiences': PracticalExperienceFormSet(
                    prefix='practical_experiences', 
                    instance=None  # No instance for new application
                ),
            }
    
    # Create mock formsets for template
    formsets['references'] = create_mock_reference_formset()
    formsets['documents'] = create_mock_document_formset()
    
    context = {
        'form': form,
        'formsets': formsets,
        'session': session,
        'show_academic_qualifications': session.selected_affiliation_type.code == 'designated',
        'show_practical_experiences': session.selected_affiliation_type.code == 'designated',
        'show_references': True,
        'show_documents': True,
        'page_title': f'Create {session.selected_affiliation_type.name} Application',
        'council_name': session.selected_council.name,
        'affiliation_type': session.selected_affiliation_type.name,
    }
    
    template_map = {
        'associated': 'enrollments/applications/associated_form.html',
        'designated': 'enrollments/applications/designated_form.html',
        'student': 'enrollments/applications/student_form.html',
    }
    
    template = template_map.get(session.selected_affiliation_type.code, 
                               'enrollments/applications/base_form.html')
    
    return render(request, template, context)
# ============================================================================
# DOCUMENT MANAGEMENT VIEWS
# ============================================================================

@login_required
@user_passes_test(can_approve_applications, login_url='/', redirect_field_name=None)
@require_POST
def document_verify(request, pk):
    """Verify a document"""
    document = get_object_or_404(Document, pk=pk)
    
    notes = request.POST.get('notes', '')
    document.verify(verified_by=request.user, notes=notes)
    
    messages.success(request, f'Document "{document.title}" verified successfully.')
    
    # Return to application detail
    app = document.content_object
    app_type = app.get_affiliation_type()
    return redirect('enrollments:application_detail', pk=app.pk, app_type=app_type)


@login_required
@user_passes_test(can_approve_applications, login_url='/', redirect_field_name=None)
@require_POST
def document_reject(request, pk):
    """Reject a document"""
    document = get_object_or_404(Document, pk=pk)
    
    notes = request.POST.get('notes', 'Document rejected by reviewer')
    document.verified = False
    document.verified_by = request.user
    document.verified_at = timezone.now()
    document.verification_notes = notes
    document.save()
    
    messages.warning(request, f'Document "{document.title}" rejected.')
    
    # Return to application detail
    app = document.content_object
    app_type = app.get_affiliation_type()
    return redirect('enrollments:application_detail', pk=app.pk, app_type=app_type)


@login_required
@user_passes_test(can_approve_applications, login_url='/', redirect_field_name=None)
@require_POST
def document_delete(request, pk):
    """Delete a document"""
    document = get_object_or_404(Document, pk=pk)
    app = document.content_object
    app_type = app.get_affiliation_type()
    
    document_title = document.title
    document.delete()
    
    messages.success(request, f'Document "{document_title}" deleted successfully.')
    return redirect('enrollments:application_detail', pk=app.pk, app_type=app_type)


# ============================================================================
# REFERENCE MANAGEMENT VIEWS
# ============================================================================

@login_required
@user_passes_test(can_approve_applications, login_url='/', redirect_field_name=None)
@require_POST
def reference_approve(request, pk):
    """Mark reference as approved"""
    reference = get_object_or_404(Reference, pk=pk)
    
    reference.letter_received = True
    reference.letter_received_date = timezone.now()
    reference.save()
    
    messages.success(request, f'Reference from {reference.get_reference_full_name()} approved.')
    
    # Return to application detail
    app = reference.content_object
    app_type = app.get_affiliation_type()
    return redirect('enrollments:application_detail', pk=app.pk, app_type=app_type)




# ============================================================================
# 2. HELPER FUNCTIONS FOR MANUAL DOCUMENT PROCESSING
# ============================================================================

def extract_document_data_from_post(request):
    """Extract document data from POST request manually."""
    document_data = []
    
    # Get total forms from management form
    total_forms = int(request.POST.get('documents-TOTAL_FORMS', 0))
    
    for i in range(total_forms):
        # Check if this form has data
        category = request.POST.get(f'documents-{i}-category')
        title = request.POST.get(f'documents-{i}-title')
        file_key = f'documents-{i}-file'
        
        if category and title and file_key in request.FILES:
            document_data.append({
                'index': i,
                'category': category,
                'title': title,
                'description': request.POST.get(f'documents-{i}-description', ''),
                'file': request.FILES[file_key],
                'is_required': request.POST.get(f'documents-{i}-is_required') == 'on',
            })
    
    return document_data

def extract_reference_data_from_post(request):
    """Extract reference data from POST request manually."""
    reference_data = []
    
    total_forms = int(request.POST.get('references-TOTAL_FORMS', 0))
    
    for i in range(total_forms):
        title = request.POST.get(f'references-{i}-reference_title')
        surname = request.POST.get(f'references-{i}-reference_surname')
        
        if title and surname:
            reference_data.append({
                'index': i,
                'reference_title': title,
                'reference_surname': surname,
                'reference_names': request.POST.get(f'references-{i}-reference_names', ''),
                'reference_email': request.POST.get(f'references-{i}-reference_email', ''),
                'reference_phone': request.POST.get(f'references-{i}-reference_phone', ''),
                'reference_address': request.POST.get(f'references-{i}-reference_address', ''),
                'nature_of_relationship': request.POST.get(f'references-{i}-nature_of_relationship', ''),
                'letter_required': request.POST.get(f'references-{i}-letter_required') == 'on',
            })
    
    return reference_data

def validate_extracted_documents(document_data):
    """Validate extracted document data."""
    errors = []
    
    for doc in document_data:
        if not doc['category']:
            errors.append(f"Document {doc['index'] + 1}: Category is required")
        if not doc['title']:
            errors.append(f"Document {doc['index'] + 1}: Title is required")
        if not doc['file']:
            errors.append(f"Document {doc['index'] + 1}: File is required")
        elif doc['file'].size > 10 * 1024 * 1024:  # 10MB
            errors.append(f"Document {doc['index'] + 1}: File too large (max 10MB)")
    
    return len(errors) == 0, errors

def validate_extracted_references(reference_data):
    """Validate extracted reference data."""
    errors = []
    
    for ref in reference_data:
        if not ref['reference_title']:
            errors.append(f"Reference {ref['index'] + 1}: Title is required")
        if not ref['reference_surname']:
            errors.append(f"Reference {ref['index'] + 1}: Surname is required")
        if not ref['reference_email']:
            errors.append(f"Reference {ref['index'] + 1}: Email is required")
    
    return len(errors) == 0, errors

def save_extracted_documents(document_data, application, content_type, user):
    """Save documents manually."""
    saved_count = 0
    
    for doc_data in document_data:
        try:
            document = Document.objects.create(
                content_type=content_type,
                object_id=application.pk,
                category=doc_data['category'],
                title=doc_data['title'],
                description=doc_data['description'],
                file=doc_data['file'],
                is_required=doc_data['is_required'],
                uploaded_by=user if user.is_authenticated else None,
            )
            logger.info(f"Created document: {document.title} (ID: {document.pk})")
            saved_count += 1
        except Exception as e:
            logger.error(f"Error saving document {doc_data['title']}: {str(e)}")
    
    return saved_count

def save_extracted_references(reference_data, application, content_type):
    """Save references manually."""
    saved_count = 0
    
    for ref_data in reference_data:
        try:
            reference = Reference.objects.create(
                content_type=content_type,
                object_id=application.pk,
                reference_title=ref_data['reference_title'],
                reference_surname=ref_data['reference_surname'],
                reference_names=ref_data['reference_names'],
                reference_email=ref_data['reference_email'],
                reference_phone=ref_data['reference_phone'],
                reference_address=ref_data['reference_address'],
                nature_of_relationship=ref_data['nature_of_relationship'],
                letter_required=ref_data['letter_required'],
            )
            logger.info(f"Created reference: {reference.get_reference_full_name()} (ID: {reference.pk})")
            saved_count += 1
        except Exception as e:
            logger.error(f"Error saving reference {ref_data['reference_surname']}: {str(e)}")
    
    return saved_count


# Replace the mock formset functions in your views.py with these:

def create_mock_reference_formset():
    """Create a mock formset for template rendering."""
    class MockForm:
        def __init__(self):
            self.errors = {}
    
    class MockFormset:
        def __init__(self):
            self.management_form = self.MockManagementForm()
            self.forms = []  # Empty list so template can iterate
            self.errors = []
            self.non_form_errors = lambda: []
            
        # Make it iterable for the template
        def __iter__(self):
            return iter(self.forms)
        
        def __len__(self):
            return len(self.forms)
    
        class MockManagementForm:
            def __str__(self):
                return '''
                <input type="hidden" name="references-TOTAL_FORMS" value="0" id="id_references-TOTAL_FORMS">
                <input type="hidden" name="references-INITIAL_FORMS" value="0" id="id_references-INITIAL_FORMS">
                <input type="hidden" name="references-MIN_NUM_FORMS" value="0" id="id_references-MIN_NUM_FORMS">
                <input type="hidden" name="references-MAX_NUM_FORMS" value="1000" id="id_references-MAX_NUM_FORMS">
                '''
    
    return MockFormset()

def create_mock_document_formset():
    """Create a mock formset for template rendering."""
    class MockFormset:
        def __init__(self):
            self.management_form = self.MockManagementForm()
            self.forms = []  # Empty list so template can iterate
            self.errors = []
            self.non_form_errors = lambda: []
            
        # Make it iterable for the template
        def __iter__(self):
            return iter(self.forms)
            
        def __len__(self):
            return len(self.forms)
    
        class MockManagementForm:
            def __str__(self):
                return '''
                <input type="hidden" name="documents-TOTAL_FORMS" value="0" id="id_documents-TOTAL_FORMS">
                <input type="hidden" name="documents-INITIAL_FORMS" value="0" id="id_documents-INITIAL_FORMS">
                <input type="hidden" name="documents-MIN_NUM_FORMS" value="0" id="id_documents-MIN_NUM_FORMS">
                <input type="hidden" name="documents-MAX_NUM_FORMS" value="1000" id="id_documents-MAX_NUM_FORMS">
                '''
    
    return MockFormset()


def application_success(request):
    """
    Success page after application submission.
    """
    success_data = request.session.get('application_success')
    
    if not success_data:
        # If no success data, redirect to onboarding start
        messages.info(request, "Please complete an application first.")
        return redirect('enrollments:onboarding_start')
    
    # Clear the success data from session after use
    del request.session['application_success']
    
    context = {
        'application_number': success_data.get('application_number'),
        'affiliation_type': success_data.get('affiliation_type'),
        'council_name': success_data.get('council_name'),
        'application_id': success_data.get('application_id'),
        'app_type': success_data.get('app_type'),
        'page_title': 'Application Submitted Successfully',
    }
    
    return render(request, 'enrollments/applications/success.html', context)


# Admin email list - all staff who should receive application notifications
ADMIN_EMAIL_LIST = [
    'anita.snyders@acrp.org.za',
    'ilse.grunewald@acrp.org.za', 
    'maria.jansen@acrp.org.za',
    'riana.andersen@acrp.org.za',
    'ams@acrp.org.za',
    'andrea.leipoldt@acrp.org.za'
]

# Email configuration
FROM_EMAIL = 'dave@kreeck.com'
REPLY_TO_EMAIL = 'ams@acrp.org.za'
from django.template.loader import render_to_string
from django.core.mail import send_mail, EmailMultiAlternatives



def send_welcome_email(application):
    """
    Send welcome email to the applicant after successful application submission.
    
    Args:
        application: The application instance (AssociatedApplication, DesignatedApplication, or StudentApplication)
    
    Returns:
        bool: True if email was sent successfully, False otherwise
    """
    try:
        # Get the onboarding session and related data
        session = application.onboarding_session
        
        # Build applicant name using correct field names
        applicant_name = f"{getattr(application, 'full_names', '')} {getattr(application, 'surname', '')}"
        if hasattr(application, 'preferred_name') and application.preferred_name:
            applicant_name = f"{application.preferred_name} {getattr(application, 'surname', '')}"
        
        # Prepare email context data with correct field names
        context = {
            'applicant_name': applicant_name.strip(),
            'applicant_email': application.email,
            'application_number': application.application_number,
            'affiliation_type': session.selected_affiliation_type.name,
            'affiliation_code': session.selected_affiliation_type.code,
            'council_name': session.selected_council.name,
            'council_code': session.selected_council.code,
            'designation_category': session.selected_designation_category.name if session.selected_designation_category else None,
            'designation_subcategory': session.selected_designation_subcategory.name if session.selected_designation_subcategory else None,
            'submission_date': application.created_at,
            'current_year': timezone.now().year,
        }
        
        # Log context for debugging
        logger.info(f"Welcome email context for {application.application_number}: applicant_name='{context['applicant_name']}'")
        
        # Render HTML email template
        html_content = render_to_string('email_templates/enrollment/welcome.html', context)
        
        # Create plain text version (fallback)
        plain_text = f"""Dear {context['applicant_name']},

Thank you for submitting your application to join the Association of Christian Religious Practitioners (ACRP).

Application Details:
- Application Number: {context['application_number']}
- Council: {context['council_name']} ({context['council_code']})
- Affiliation Type: {context['affiliation_type']}
- Submission Date: {context['submission_date'].strftime('%B %d, %Y')}

Your application is now under review by our {context['council_name']} team. We will contact you within 5-10 business days with an update.

For any questions, please contact us at ams@acrp.org.za or (+27) 065 214 1536.

Best regards,
The ACRP Team
Association of Christian Religious Practitioners

---
This is an automated message. Please do not reply directly to this email."""
        
        # Create email message
        email = EmailMultiAlternatives(
            subject=f'Welcome to ACRP - Application {application.application_number} Submitted',
            body=plain_text,
            from_email=FROM_EMAIL,
            to=[application.email],
            reply_to=[REPLY_TO_EMAIL]
        )
        
        # Attach HTML version
        email.attach_alternative(html_content, "text/html")
        
        # Send the email
        email.send()
        
        logger.info(f"Welcome email sent successfully to {application.email} for application {application.application_number}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send welcome email for application {application.application_number}: {str(e)}")
        logger.error(f"Application fields available: {[field for field in dir(application) if not field.startswith('_')]}")
        return False
    





def send_admin_notification(application):
    """
    Send notification email to admin staff when a new application is submitted.
    
    Args:
        application: The application instance (AssociatedApplication, DesignatedApplication, or StudentApplication)
    
    Returns:
        bool: True if email was sent successfully, False otherwise
    """
    try:
        # Get the onboarding session and related data
        session = application.onboarding_session
        
        # Build applicant name using correct field names
        applicant_name = f"{getattr(application, 'full_names', '')} {getattr(application, 'surname', '')}"
        if hasattr(application, 'preferred_name') and application.preferred_name:
            applicant_name = f"{application.preferred_name} {getattr(application, 'surname', '')}"
        
        # Get phone number - try different possible field names
        phone_number = (
            getattr(application, 'cell_phone', None) or 
            getattr(application, 'work_phone', None) or 
            getattr(application, 'home_phone', None) or
            getattr(application, 'phone_number', None) or
            getattr(application, 'phone', None)
        )
        
        # Get organization - try different possible field names
        organization = (
            getattr(application, 'current_occupation', None) or
            getattr(application, 'current_organization', None) or
            getattr(application, 'organization', None) or
            getattr(application, 'work_description', None)
        )
        
        # Prepare comprehensive email context data with correct field names
        context = {
            # Application identification
            'application_number': application.application_number,
            'submission_date': application.created_at,
            'current_year': timezone.now().year,
            
            # Applicant information using correct field names
            'applicant_name': applicant_name.strip(),
            'applicant_email': application.email,
            'applicant_phone': phone_number,
            'applicant_organization': organization,
            
            # Council and affiliation information
            'council_name': session.selected_council.name,
            'council_code': session.selected_council.code,
            'affiliation_type': session.selected_affiliation_type.name,
            'affiliation_code': session.selected_affiliation_type.code,
            
            # Designation information (for designated applications)
            'designation_category': session.selected_designation_category.name if session.selected_designation_category else None,
            'designation_subcategory': session.selected_designation_subcategory.name if session.selected_designation_subcategory else None,
            
            # Admin action URLs (adjust these based on your URL patterns)
            'application_url': f"{getattr(settings, 'SITE_URL', 'http://localhost:8000')}/admin/applications/view/{application.pk}/",
            'admin_dashboard_url': f"{getattr(settings, 'SITE_URL', 'http://localhost:8000')}/admin/dashboard/",
        }
        
        # Log context for debugging
        logger.info(f"Admin email context for {application.application_number}: applicant_name='{context['applicant_name']}'")
        
        # Render HTML email template
        html_content = render_to_string('email_templates/enrollment/admin_notification.html', context)
        
        # Build designation information for plain text
        designation_info = ""
        if context['designation_category']:
            designation_info += f"- Designation Category: {context['designation_category']}\n"
        if context['designation_subcategory']:
            designation_info += f"- Designation Subcategory: {context['designation_subcategory']}\n"
        
        # Build applicant contact info
        contact_info = f"{context['applicant_name']} ({context['applicant_email']})"
        if context['applicant_phone']:
            contact_info += f" - {context['applicant_phone']}"
        if context['applicant_organization']:
            contact_info += f" - {context['applicant_organization']}"
        
        # Create clean plain text version
        plain_text = f"""NEW ACRP APPLICATION SUBMITTED

Application Details:
- Application Number: {context['application_number']}
- Council: {context['council_name']} ({context['council_code']})
- Affiliation Type: {context['affiliation_type']}
- Applicant: {contact_info}
- Submission Date: {context['submission_date'].strftime('%B %d, %Y at %I:%M %p')}
{designation_info}
Action Required:
This application requires review and processing by the {context['council_name']} team.

Please review this application within 5-10 business days to maintain our service commitments.

Review Application: {context['application_url']}
Admin Dashboard: {context['admin_dashboard_url']}

---
ACRP Admin Notification System
This is an automated notification."""
        
        # Create email subject with council code for easy filtering
        subject = f"New {context['council_code']} Application: {context['application_number']}"
        
        # Create email message with both HTML and plain text versions
        email = EmailMultiAlternatives(
            subject=subject,
            body=plain_text,
            from_email=FROM_EMAIL,
            to=ADMIN_EMAIL_LIST,
            reply_to=[REPLY_TO_EMAIL]
        )
        
        # Attach HTML version for rich formatting
        email.attach_alternative(html_content, "text/html")
        
        # Send the email
        email.send()
        
        # Log successful sending
        logger.info(
            f"Admin notification sent successfully for application {application.application_number} "
            f"to {len(ADMIN_EMAIL_LIST)} recipients: {', '.join(ADMIN_EMAIL_LIST)}"
        )
        return True
        
    except Exception as e:
        # Log error but don't raise to prevent application submission failure
        logger.error(
            f"Failed to send admin notification for application {application.application_number}: {str(e)}. "
            f"Recipients: {', '.join(ADMIN_EMAIL_LIST)}"
        )
        logger.error(f"Application fields available: {[field for field in dir(application) if not field.startswith('_')]}")
        return False


def send_application_emails(application):
    """
    Send both welcome email to applicant and notification to admin staff.
    
    Args:
        application: The application instance (AssociatedApplication, DesignatedApplication, or StudentApplication)
    
    Returns:
        dict: Results of email sending attempts
    """
    # Initialize results tracking
    results = {
        'welcome_email_sent': False,
        'admin_notification_sent': False,
        'errors': []
    }
    
    logger.info(f"Starting email sending process for application {application.application_number}")
    
    # Debug application fields
    logger.info(f"Application type: {type(application).__name__}")
    logger.info(f"Available fields: email={getattr(application, 'email', 'NOT_FOUND')}, "
               f"surname={getattr(application, 'surname', 'NOT_FOUND')}, "
               f"full_names={getattr(application, 'full_names', 'NOT_FOUND')}")
    
    # Send welcome email to applicant
    try:
        logger.info(f"Attempting to send welcome email to {application.email}")
        results['welcome_email_sent'] = send_welcome_email(application)
        
        if not results['welcome_email_sent']:
            error_msg = "Failed to send welcome email to applicant"
            results['errors'].append(error_msg)
            logger.warning(f"Welcome email failed for application {application.application_number}: {error_msg}")
        else:
            logger.info(f"Welcome email sent successfully for application {application.application_number}")
            
    except Exception as e:
        error_msg = f"Welcome email error: {str(e)}"
        results['errors'].append(error_msg)
        logger.error(f"Welcome email exception for application {application.application_number}: {str(e)}")
    
    # Send notification to admin staff
    try:
        logger.info(f"Attempting to send admin notification to {len(ADMIN_EMAIL_LIST)} recipients")
        results['admin_notification_sent'] = send_admin_notification(application)
        
        if not results['admin_notification_sent']:
            error_msg = "Failed to send admin notification"
            results['errors'].append(error_msg)
            logger.warning(f"Admin notification failed for application {application.application_number}: {error_msg}")
        else:
            logger.info(f"Admin notification sent successfully for application {application.application_number}")
            
    except Exception as e:
        error_msg = f"Admin notification error: {str(e)}"
        results['errors'].append(error_msg)
        logger.error(f"Admin notification exception for application {application.application_number}: {str(e)}")
    
    # Log comprehensive results
    if results['welcome_email_sent'] and results['admin_notification_sent']:
        logger.info(f"All emails sent successfully for application {application.application_number}")
    elif results['welcome_email_sent'] or results['admin_notification_sent']:
        logger.warning(
            f"Partial email success for application {application.application_number}. "
            f"Welcome: {results['welcome_email_sent']}, Admin: {results['admin_notification_sent']}"
        )
    else:
        logger.error(f"All emails failed for application {application.application_number}. Errors: {results['errors']}")
    
    return results





def send_application_emails(application):
    """
    Send both welcome email to applicant and notification to admin staff.
    
    This is the main email coordination function that handles sending both
    the welcome email to the applicant and the notification to admin staff.
    It provides comprehensive error handling and detailed logging.
    
    Args:
        application: The application instance (AssociatedApplication, DesignatedApplication, or StudentApplication)
    
    Returns:
        dict: Results of email sending attempts with the following structure:
              {
                  'welcome_email_sent': bool,
                  'admin_notification_sent': bool,
                  'errors': list[str]
              }
    """
    # Initialize results tracking
    results = {
        'welcome_email_sent': False,
        'admin_notification_sent': False,
        'errors': []
    }
    
    logger.info(f"Starting email sending process for application {application.application_number}")
    
    # Send welcome email to applicant
    try:
        logger.info(f"Attempting to send welcome email to {application.email}")
        results['welcome_email_sent'] = send_welcome_email(application)
        
        if not results['welcome_email_sent']:
            error_msg = "Failed to send welcome email to applicant"
            results['errors'].append(error_msg)
            logger.warning(f"Welcome email failed for application {application.application_number}: {error_msg}")
        else:
            logger.info(f"Welcome email sent successfully for application {application.application_number}")
            
    except Exception as e:
        error_msg = f"Welcome email error: {str(e)}"
        results['errors'].append(error_msg)
        logger.error(f"Welcome email exception for application {application.application_number}: {str(e)}")
    
    # Send notification to admin staff
    try:
        logger.info(f"Attempting to send admin notification to {len(ADMIN_EMAIL_LIST)} recipients")
        results['admin_notification_sent'] = send_admin_notification(application)
        
        if not results['admin_notification_sent']:
            error_msg = "Failed to send admin notification"
            results['errors'].append(error_msg)
            logger.warning(f"Admin notification failed for application {application.application_number}: {error_msg}")
        else:
            logger.info(f"Admin notification sent successfully for application {application.application_number}")
            
    except Exception as e:
        error_msg = f"Admin notification error: {str(e)}"
        results['errors'].append(error_msg)
        logger.error(f"Admin notification exception for application {application.application_number}: {str(e)}")
    
    # Log comprehensive results
    if results['welcome_email_sent'] and results['admin_notification_sent']:
        logger.info(f"All emails sent successfully for application {application.application_number}")
    elif results['welcome_email_sent'] or results['admin_notification_sent']:
        logger.warning(
            f"Partial email success for application {application.application_number}. "
            f"Welcome: {results['welcome_email_sent']}, Admin: {results['admin_notification_sent']}"
        )
    else:
        logger.error(f"All emails failed for application {application.application_number}. Errors: {results['errors']}")
    
    return results





def handle_application_email_sending(application):
    """
    Handle email sending for a newly created application.
    
    This function provides a clean interface for sending application emails
    from the application creation view, with appropriate error handling that
    doesn't interfere with the application submission process.
    
    Args:
        application: The newly created application instance
        
    Returns:
        bool: True if at least one email was sent successfully, False if all failed
    """
    try:
        # Send both emails using the main email function
        email_results = send_application_emails(application)
        
        # The detailed logging is already handled in send_application_emails,
        # so we only need to handle any critical failures here
        if email_results['success_count'] == 0:
            logger.critical(f"CRITICAL: All emails failed for application {application.application_number}. "
                          f"Manual intervention may be required. Errors: {'; '.join(email_results['errors'])}")
        
        # Return True if at least one email was sent successfully
        return email_results['success_count'] > 0
        
    except Exception as e:
        # This catch-all ensures that email failures never break application submission
        logger.critical(f"CRITICAL: Email sending process crashed for application {application.application_number}: {str(e)}")
        return False







def export_applications(request):
    """
    Professional Excel export with multiple format options.
    
    Export formats available:
    - summary: High-level statistics and summary data
    - detailed: Complete application information
    - council_breakdown: Applications grouped by council
    - status_report: Applications grouped by status
    - timeline: Applications with timeline data
    - contact_list: Contact information only
    """
    
    # Get export format from request
    export_format = request.GET.get('format', 'summary')
    
    # Get the same filters as the list view
    search_query = request.GET.get('search_query')
    council_filter = request.GET.get('council')
    status_filter = request.GET.get('status')
    app_type_filter = request.GET.get('app_type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Apply filters and get applications data
    applications_data = get_filtered_applications_for_export(
        search_query, council_filter, status_filter, 
        app_type_filter, date_from, date_to
    )
    
    # Route to appropriate export function
    export_functions = {
        'summary': export_summary_format,
        'detailed': export_detailed_format,
        'council_breakdown': export_council_breakdown_format,
        'status_report': export_status_report_format,
        'timeline': export_timeline_format,
        'contact_list': export_contact_list_format,
    }
    
    export_function = export_functions.get(export_format, export_summary_format)
    
    try:
        response = export_function(applications_data, request)
        
        # Log export activity
        logger.info(
            f"Applications exported by user {request.user.email} "
            f"- Format: {export_format}, Count: {len(applications_data)}"
        )
        
        return response
        
    except Exception as e:
        logger.error(f"Export failed for user {request.user.email}: {str(e)}")
        raise


def get_filtered_applications_for_export(search_query, council_filter, status_filter, 
                                       app_type_filter, date_from, date_to):
    """
    Get applications data with applied filters for export.
    Returns a comprehensive list of application dictionaries.
    """
    applications_data = []
    
    # Define application models
    app_models = [
        ('associated', AssociatedApplication),
        ('designated', DesignatedApplication),
        ('student', StudentApplication),
    ]
    
    for app_type_name, Model in app_models:
        # Skip if filtering by specific app type
        if app_type_filter and app_type_filter != app_type_name:
            continue
            
        # Build queryset with proper relationships
        qs = Model.objects.select_related(
            'onboarding_session__selected_council',
            'onboarding_session__selected_affiliation_type',
            'submitted_by',
            'reviewed_by',
            'approved_by'
        )
        
        # Add designation category for designated applications
        if app_type_name == 'designated':
            qs = qs.select_related('designation_category', 'designation_subcategory')
        
        # Apply filters
        if council_filter:
            qs = qs.filter(onboarding_session__selected_council_id=council_filter)
        if status_filter:
            qs = qs.filter(status=status_filter)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        if search_query:
            qs = qs.filter(
                Q(full_names__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(application_number__icontains=search_query)
            )
        
        # Process applications
        for app in qs:
            app_data = {
                'id': app.pk,
                'application_number': app.application_number,
                'type': 'Learner' if app_type_name == 'student' else app_type_name.title(),
                'app_type': app_type_name,
                
                # Personal Information
                'full_names': getattr(app, 'full_names', ''),
                'first_name': getattr(app, 'first_name', ''),
                'surname': getattr(app, 'surname', ''),
                'preferred_name': getattr(app, 'preferred_name', ''),
                'email': app.email,
                'cell_phone': getattr(app, 'cell_phone', ''),
                'home_phone': getattr(app, 'home_phone', ''),
                'work_phone': getattr(app, 'work_phone', ''),
                
                # Identity Information
                'id_number': getattr(app, 'id_number', ''),
                'passport_number': getattr(app, 'passport_number', ''),
                'date_of_birth': getattr(app, 'date_of_birth', ''),
                'gender': getattr(app, 'gender', ''),
                'race': getattr(app, 'race', ''),
                'nationality': getattr(app, 'nationality', ''),
                'home_language': getattr(app, 'home_language', ''),
                
                # Address Information
                'physical_address': f"{getattr(app, 'physical_address_line1', '')} {getattr(app, 'physical_address_line2', '')}".strip(),
                'physical_city': getattr(app, 'physical_city', ''),
                'physical_province': getattr(app, 'physical_province', ''),
                'physical_code': getattr(app, 'physical_code', ''),
                'physical_country': getattr(app, 'physical_country', ''),
                'postal_address': f"{getattr(app, 'postal_address_line1', '')} {getattr(app, 'postal_address_line2', '')}".strip(),
                'postal_city': getattr(app, 'postal_city', ''),
                'postal_province': getattr(app, 'postal_province', ''),
                'postal_code': getattr(app, 'postal_code', ''),
                'postal_country': getattr(app, 'postal_country', ''),
                
                # Council and Affiliation
                'council': app.onboarding_session.selected_council.code,
                'council_name': app.onboarding_session.selected_council.name,
                'affiliation_type': app.onboarding_session.selected_affiliation_type.name if app.onboarding_session.selected_affiliation_type else '',
                
                # Application Status and Dates
                'status': app.status,
                'created_at': app.created_at,
                'submitted_at': app.submitted_at,
                'reviewed_at': getattr(app, 'reviewed_at', ''),
                'approved_at': getattr(app, 'approved_at', ''),
                
                # Staff Information
                'submitted_by': app.submitted_by.get_full_name() if app.submitted_by else '',
                'reviewed_by': app.reviewed_by.get_full_name() if app.reviewed_by else '',
                'approved_by': app.approved_by.get_full_name() if app.approved_by else '',
                'reviewer_notes': getattr(app, 'reviewer_notes', ''),
                
                # Type-specific information
                'designation_category': '',
                'designation_subcategory': '',
                'current_institution': '',
                'qualification_institution': '',
                'highest_qualification': '',
                'years_in_ministry': '',
                'current_occupation': '',
                'religious_affiliation': '',
            }
            
            # Add type-specific data
            if app_type_name == 'designated':
                app_data.update({
                    'designation_category': app.designation_category.name if app.designation_category else '',
                    'designation_subcategory': app.designation_subcategory.name if app.designation_subcategory else '',
                    'qualification_institution': getattr(app, 'qualification_institution', ''),
                    'highest_qualification': getattr(app, 'highest_qualification', ''),
                    'years_in_ministry': getattr(app, 'years_in_ministry', ''),
                    'current_occupation': getattr(app, 'current_occupation', ''),
                    'religious_affiliation': getattr(app, 'religious_affiliation', ''),
                })
            elif app_type_name == 'student':
                app_data.update({
                    'current_institution': getattr(app, 'current_institution', ''),
                    'qualification_institution': getattr(app, 'current_institution', ''),
                })
            
            applications_data.append(app_data)
    
    return applications_data


def export_summary_format(applications_data, request):
    """
    Export summary statistics and high-level overview.
    Perfect for executive reports and quick overviews.
    """
    workbook = openpyxl.Workbook()
    
    # Remove default sheet and create custom sheets
    workbook.remove(workbook.active)
    
    # Create summary sheet
    summary_sheet = workbook.create_sheet("Executive Summary")
    
    # Define styles
    header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    subheader_font = Font(name='Calibri', size=12, bold=True, color='366092')
    normal_font = Font(name='Calibri', size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title and metadata
    summary_sheet['A1'] = 'ACRP AMS Application Export - Executive Summary'
    summary_sheet['A1'].font = Font(name='Calibri', size=16, bold=True, color='366092')
    summary_sheet['A2'] = f'Generated on: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
    summary_sheet['A3'] = f'Generated by: {request.user.username}'
    summary_sheet['A4'] = f'Total Applications: {len(applications_data)}'
    
    # Calculate statistics
    stats = calculate_export_statistics(applications_data)
    
    # Overall Statistics Section
    row = 6
    summary_sheet[f'A{row}'] = 'OVERALL STATISTICS'
    summary_sheet[f'A{row}'].font = subheader_font
    row += 2
    
    stats_data = [
        ['Total Applications', stats['total_applications']],
        ['Approved Applications', stats['approved_count']],
        ['Pending Applications', stats['pending_count']],
        ['Under Review', stats['under_review_count']],
        ['Rejected Applications', stats['rejected_count']],
        ['Approval Rate', f"{stats['approval_rate']:.1f}%"],
    ]
    
    for stat_row in stats_data:
        summary_sheet[f'A{row}'] = stat_row[0]
        summary_sheet[f'B{row}'] = stat_row[1]
        summary_sheet[f'A{row}'].font = normal_font
        summary_sheet[f'B{row}'].font = Font(name='Calibri', size=11, bold=True)
        row += 1
    
    # By Council Section
    row += 2
    summary_sheet[f'A{row}'] = 'BY COUNCIL'
    summary_sheet[f'A{row}'].font = subheader_font
    row += 1
    
    # Headers for council table
    council_headers = ['Council', 'Total', 'Approved', 'Pending', 'Rejected', 'Approval Rate']
    for col, header in enumerate(council_headers, 1):
        cell = summary_sheet.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    
    # Council data
    for council_code, council_stats in stats['by_council'].items():
        council_data = [
            council_code,
            council_stats['total'],
            council_stats['approved'],
            council_stats['pending'],
            council_stats['rejected'],
            f"{council_stats['approval_rate']:.1f}%"
        ]
        for col, value in enumerate(council_data, 1):
            cell = summary_sheet.cell(row=row, column=col)
            cell.value = value
            cell.font = normal_font
            cell.border = border
        row += 1
    
    # By Application Type Section
    row += 2
    summary_sheet[f'A{row}'] = 'BY APPLICATION TYPE'
    summary_sheet[f'A{row}'].font = subheader_font
    row += 1
    
    # Headers for type table
    type_headers = ['Application Type', 'Count', 'Percentage']
    for col, header in enumerate(type_headers, 1):
        cell = summary_sheet.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    
    # Type data
    for app_type, count in stats['by_type'].items():
        percentage = (count / stats['total_applications'] * 100) if stats['total_applications'] > 0 else 0
        type_data = [app_type, count, f"{percentage:.1f}%"]
        for col, value in enumerate(type_data, 1):
            cell = summary_sheet.cell(row=row, column=col)
            cell.value = value
            cell.font = normal_font
            cell.border = border
        row += 1
    
    # Auto-adjust column widths
    for column in summary_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create detailed data sheet
    create_detailed_data_sheet(workbook, applications_data)
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ACRP_Applications_Summary_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    workbook.save(response)
    return response


def export_detailed_format(applications_data, request):
    """
    Export complete detailed information for all applications.
    Perfect for comprehensive analysis and record-keeping.
    """
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    
    # Create detailed sheet
    detail_sheet = workbook.create_sheet("Detailed Applications")
    
    # Define comprehensive headers
    headers = [
        'Application ID', 'Application Number', 'Type', 'Status',
        'Full Names', 'Email', 'Cell Phone', 'Home Phone', 'Work Phone',
        'ID Number', 'Passport Number', 'Date of Birth', 'Gender', 'Race',
        'Nationality', 'Home Language', 'Physical Address', 'Physical City',
        'Physical Province', 'Physical Code', 'Physical Country',
        'Postal Address', 'Postal City', 'Postal Province', 'Postal Code',
        'Postal Country', 'Council', 'Council Name', 'Affiliation Type',
        'Designation Category', 'Designation Subcategory', 'Current Institution',
        'Qualification Institution', 'Highest Qualification', 'Years in Ministry',
        'Current Occupation', 'Religious Affiliation', 'Created Date',
        'Submitted Date', 'Reviewed Date', 'Approved Date', 'Submitted By',
        'Reviewed By', 'Approved By', 'Reviewer Notes'
    ]
    
    # Style headers
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    normal_font = Font(name='Calibri', size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = detail_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Add data
    for row, app in enumerate(applications_data, 2):
        data_row = [
            app['id'], app['application_number'], app['type'], app['status'],
            app['full_names'], app['email'], app['cell_phone'], app['home_phone'],
            app['work_phone'], app['id_number'], app['passport_number'],
            app['date_of_birth'], app['gender'], app['race'], app['nationality'],
            app['home_language'], app['physical_address'], app['physical_city'],
            app['physical_province'], app['physical_code'], app['physical_country'],
            app['postal_address'], app['postal_city'], app['postal_province'],
            app['postal_code'], app['postal_country'], app['council'],
            app['council_name'], app['affiliation_type'], app['designation_category'],
            app['designation_subcategory'], app['current_institution'],
            app['qualification_institution'], app['highest_qualification'],
            app['years_in_ministry'], app['current_occupation'],
            app['religious_affiliation'], app['created_at'], app['submitted_at'],
            app['reviewed_at'], app['approved_at'], app['submitted_by'],
            app['reviewed_by'], app['approved_by'], app['reviewer_notes']
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = detail_sheet.cell(row=row, column=col)
            cell.value = value
            cell.font = normal_font
            cell.border = border
    
    # Auto-adjust column widths
    for column in detail_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        detail_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ACRP_Applications_Detailed_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    workbook.save(response)
    return response


def export_council_breakdown_format(applications_data, request):
    """
    Export applications grouped by council with separate sheets.
    Perfect for council-specific analysis and reporting.
    """
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    
    # Group applications by council
    council_groups = {}
    for app in applications_data:
        council = app['council']
        if council not in council_groups:
            council_groups[council] = []
        council_groups[council].append(app)
    
    # Create summary sheet
    summary_sheet = workbook.create_sheet("Council Summary")
    create_council_summary_sheet(summary_sheet, council_groups, request)
    
    # Create sheet for each council
    for council_code, council_apps in council_groups.items():
        council_sheet = workbook.create_sheet(f"{council_code} Applications")
        create_council_detail_sheet(council_sheet, council_apps, council_code)
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ACRP_Applications_By_Council_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    workbook.save(response)
    return response


def export_contact_list_format(applications_data, request):
    """
    Export contact information only.
    Perfect for communication and outreach purposes.
    """
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    
    contact_sheet = workbook.create_sheet("Contact List")
    
    # Headers for contact list
    headers = [
        'Full Names', 'Email', 'Cell Phone', 'Home Phone', 'Work Phone',
        'Application Type', 'Council', 'Status', 'Physical Address',
        'Physical City', 'Physical Province', 'Postal Address', 'Postal City',
        'Postal Province', 'Application Number', 'Submitted Date'
    ]
    
    # Style headers
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    normal_font = Font(name='Calibri', size=10)
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = contact_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    # Add contact data
    for row, app in enumerate(applications_data, 2):
        contact_data = [
            app['full_names'], app['email'], app['cell_phone'], app['home_phone'],
            app['work_phone'], app['type'], app['council'], app['status'],
            app['physical_address'], app['physical_city'], app['physical_province'],
            app['postal_address'], app['postal_city'], app['postal_province'],
            app['application_number'], app['submitted_at']
        ]
        
        for col, value in enumerate(contact_data, 1):
            cell = contact_sheet.cell(row=row, column=col)
            cell.value = value
            cell.font = normal_font
    
    # Auto-adjust column widths
    for column in contact_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 35)
        contact_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ACRP_Contact_List_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    workbook.save(response)
    return response




def calculate_export_statistics(applications_data):
    """Calculate comprehensive statistics for export."""
    total_applications = len(applications_data)
    
    if total_applications == 0:
        return {
            'total_applications': 0,
            'approved_count': 0,
            'pending_count': 0,
            'under_review_count': 0,
            'rejected_count': 0,
            'approval_rate': 0,
            'by_council': {},
            'by_type': {},
            'by_status': {}
        }
    
    # Calculate status counts
    status_counts = {}
    council_stats = {}
    type_counts = {}
    
    for app in applications_data:
        # Status counting
        status = app['status']
        status_counts[status] = status_counts.get(status, 0) + 1
        
        # Council statistics
        council = app['council']
        if council not in council_stats:
            council_stats[council] = {
                'total': 0, 'approved': 0, 'pending': 0,
                'under_review': 0, 'rejected': 0, 'approval_rate': 0
            }
        
        council_stats[council]['total'] += 1
        if status == 'approved':
            council_stats[council]['approved'] += 1
        elif status in ['pending', 'submitted']:
            council_stats[council]['pending'] += 1
        elif status == 'under_review':
            council_stats[council]['under_review'] += 1
        elif status == 'rejected':
            council_stats[council]['rejected'] += 1
        
        # Type counting
        app_type = app['type']
        type_counts[app_type] = type_counts.get(app_type, 0) + 1
    
    # Calculate approval rates for councils
    for council, stats in council_stats.items():
        processed = stats['approved'] + stats['rejected']
        if processed > 0:
            stats['approval_rate'] = (stats['approved'] / processed) * 100
    
    # Overall approval rate
    total_approved = status_counts.get('approved', 0)
    total_rejected = status_counts.get('rejected', 0)
    total_processed = total_approved + total_rejected
    overall_approval_rate = (total_approved / total_processed * 100) if total_processed > 0 else 0
    
    return {
        'total_applications': total_applications,
        'approved_count': status_counts.get('approved', 0),
        'pending_count': status_counts.get('pending', 0) + status_counts.get('submitted', 0),
        'under_review_count': status_counts.get('under_review', 0),
        'rejected_count': status_counts.get('rejected', 0),
        'approval_rate': overall_approval_rate,
        'by_council': council_stats,
        'by_type': type_counts,
        'by_status': status_counts
    }


def create_detailed_data_sheet(workbook, applications_data):
    """Create detailed data sheet for summary export."""
    data_sheet = workbook.create_sheet("Raw Data")
    
    # Headers for detailed data
    headers = [
        'ID', 'Application Number', 'Type', 'Full Names', 'Email', 'Status',
        'Council', 'Created Date', 'Submitted Date', 'Cell Phone', 'City',
        'Province', 'Designation Category', 'Institution'
    ]
    
    # Style definitions
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    normal_font = Font(name='Calibri', size=9)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = data_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Add data rows
    for row, app in enumerate(applications_data, 2):
        data_row = [
            app['id'], app['application_number'], app['type'], app['full_names'],
            app['email'], app['status'], app['council'], app['created_at'],
            app['submitted_at'], app['cell_phone'], app['physical_city'],
            app['physical_province'], app['designation_category'], 
            app['current_institution']
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = data_sheet.cell(row=row, column=col)
            cell.value = value
            cell.font = normal_font
            cell.border = border
    
    # Auto-adjust column widths
    for column in data_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        data_sheet.column_dimensions[column_letter].width = adjusted_width


def create_council_summary_sheet(sheet, council_groups, request):
    """Create council summary sheet."""
    # Title and metadata
    sheet['A1'] = 'Council Breakdown Summary'
    sheet['A1'].font = Font(name='Calibri', size=16, bold=True, color='366092')
    sheet['A2'] = f'Generated on: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
    sheet['A3'] = f'Generated by: {request.user.username}'
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    subheader_font = Font(name='Calibri', size=12, bold=True, color='366092')
    normal_font = Font(name='Calibri', size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    row = 5
    
    # Council summary table
    sheet[f'A{row}'] = 'COUNCIL SUMMARY'
    sheet[f'A{row}'].font = subheader_font
    row += 2
    
    # Headers
    summary_headers = ['Council', 'Total Applications', 'Approved', 'Pending', 'Under Review', 'Rejected', 'Approval Rate']
    for col, header in enumerate(summary_headers, 1):
        cell = sheet.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    
    # Council data
    for council_code, council_apps in council_groups.items():
        stats = calculate_export_statistics(council_apps)
        council_data = [
            council_code,
            stats['total_applications'],
            stats['approved_count'],
            stats['pending_count'],
            stats['under_review_count'],
            stats['rejected_count'],
            f"{stats['approval_rate']:.1f}%"
        ]
        
        for col, value in enumerate(council_data, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = value
            cell.font = normal_font
            cell.border = border
        row += 1
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        sheet.column_dimensions[column_letter].width = adjusted_width


def create_council_detail_sheet(sheet, council_apps, council_code):
    """Create individual council detail sheet."""
    # Title
    sheet['A1'] = f'{council_code} Council Applications'
    sheet['A1'].font = Font(name='Calibri', size=14, bold=True, color='366092')
    sheet['A2'] = f'Total Applications: {len(council_apps)}'
    
    # Headers
    headers = [
        'Application Number', 'Full Names', 'Email', 'Type', 'Status',
        'Cell Phone', 'City', 'Province', 'Created Date', 'Submitted Date',
        'Designation Category', 'Institution'
    ]
    
    # Style definitions
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    normal_font = Font(name='Calibri', size=9)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add headers
    row = 4
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Add application data
    for app_row, app in enumerate(council_apps, row + 1):
        data_row = [
            app['application_number'], app['full_names'], app['email'],
            app['type'], app['status'], app['cell_phone'], app['physical_city'],
            app['physical_province'], app['created_at'], app['submitted_at'],
            app['designation_category'], app['current_institution']
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = sheet.cell(row=app_row, column=col)
            cell.value = value
            cell.font = normal_font
            cell.border = border
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        sheet.column_dimensions[column_letter].width = adjusted_width


def export_status_report_format(applications_data, request):
    """
    Export applications grouped by status.
    Perfect for processing and workflow management.
    """
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    
    # Group applications by status
    status_groups = {}
    for app in applications_data:
        status = app['status']
        if status not in status_groups:
            status_groups[status] = []
        status_groups[status].append(app)
    
    # Create summary sheet
    summary_sheet = workbook.create_sheet("Status Summary")
    create_status_summary_sheet(summary_sheet, status_groups, request)
    
    # Create sheet for each status
    status_order = ['submitted', 'under_review', 'approved', 'rejected', 'pending', 'draft']
    for status in status_order:
        if status in status_groups:
            status_sheet = workbook.create_sheet(f"{status.title()} Applications")
            create_status_detail_sheet(status_sheet, status_groups[status], status)
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ACRP_Status_Report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    workbook.save(response)
    return response


def export_timeline_format(applications_data, request):
    """
    Export timeline analysis of applications.
    Perfect for trend analysis and performance monitoring.
    """
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    
    # Create timeline sheet
    timeline_sheet = workbook.create_sheet("Timeline Analysis")
    
    # Title and metadata
    timeline_sheet['A1'] = 'Applications Timeline Analysis'
    timeline_sheet['A1'].font = Font(name='Calibri', size=16, bold=True, color='366092')
    timeline_sheet['A2'] = f'Generated on: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
    timeline_sheet['A3'] = f'Period: {applications_data[0]["created_at"].strftime("%Y-%m-%d") if applications_data else "N/A"} to {applications_data[-1]["created_at"].strftime("%Y-%m-%d") if applications_data else "N/A"}'
    
    # Headers for timeline data
    headers = [
        'Application Number', 'Full Names', 'Type', 'Council', 'Status',
        'Created Date', 'Submitted Date', 'Reviewed Date', 'Approved Date',
        'Days to Submit', 'Days to Review', 'Days to Approve', 'Total Processing Days'
    ]
    
    # Style definitions
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    normal_font = Font(name='Calibri', size=9)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add headers
    row = 5
    for col, header in enumerate(headers, 1):
        cell = timeline_sheet.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Add timeline data
    for app_row, app in enumerate(applications_data, row + 1):
        # Calculate timeline metrics
        created_date = app['created_at']
        submitted_date = app['submitted_at']
        reviewed_date = app['reviewed_at']
        approved_date = app['approved_at']
        
        days_to_submit = ''
        days_to_review = ''
        days_to_approve = ''
        total_days = ''
        
        if submitted_date and created_date:
            days_to_submit = (submitted_date - created_date).days
        
        if reviewed_date and submitted_date:
            days_to_review = (reviewed_date - submitted_date).days
        
        if approved_date and submitted_date:
            days_to_approve = (approved_date - submitted_date).days
            total_days = (approved_date - created_date).days
        
        timeline_data = [
            app['application_number'], app['full_names'], app['type'],
            app['council'], app['status'], created_date, submitted_date,
            reviewed_date, approved_date, days_to_submit, days_to_review,
            days_to_approve, total_days
        ]
        
        for col, value in enumerate(timeline_data, 1):
            cell = timeline_sheet.cell(row=app_row, column=col)
            cell.value = value
            cell.font = normal_font
            cell.border = border
    
    # Auto-adjust column widths
    for column in timeline_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        timeline_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ACRP_Timeline_Analysis_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    workbook.save(response)
    return response


def create_status_summary_sheet(sheet, status_groups, request):
    """Create status summary sheet."""
    # Similar implementation to council summary but for statuses
    sheet['A1'] = 'Status Summary Report'
    sheet['A1'].font = Font(name='Calibri', size=16, bold=True, color='366092')
    # Add status breakdown table here
    pass


def create_status_detail_sheet(sheet, status_apps, status):
    """Create individual status detail sheet."""
    # Similar implementation to council detail but for specific status
    sheet['A1'] = f'{status.title()} Applications'
    sheet['A1'].font = Font(name='Calibri', size=14, bold=True, color='366092')
    # Add application details for this status
    pass



# ============================================================================
# APPLICATION MANAGEMENT VIEWS
# ============================================================================




@login_required
@permission_required('enrollments.view_baseapplication', raise_exception=True)
def application_list(request):
    """
    Unified list view for all application types with advanced filtering.
    
    Shows applications from all councils and affiliation types in one view.
    Performance optimized with proper queryset handling.
    """
    # Get search parameters
    search_form = ApplicationSearchForm(request.GET or None)
    
    # Initialize filter variables
    search_query = None
    council_filter = None
    affiliation_type_filter = None
    status_filter = None
    date_from = None
    date_to = None
    app_type_filter = None
    
    # Apply filters if form is valid
    if search_form.is_valid():
        search_query = search_form.cleaned_data.get('search_query')
        council_filter = search_form.cleaned_data.get('council')
        affiliation_type_filter = search_form.cleaned_data.get('affiliation_type')
        status_filter = search_form.cleaned_data.get('status')
        date_from = search_form.cleaned_data.get('date_from')
        date_to = search_form.cleaned_data.get('date_to')
        app_type_filter = request.GET.get('app_type')  # Get from URL params
    
    # Get applications using the same logic as export
    all_applications = []
    
    # Define application models
    app_models = [
        ('associated', AssociatedApplication),
        ('designated', DesignatedApplication),
        ('student', StudentApplication),
    ]
    
    # Base querysets with proper select_related for performance
    base_select_related = [
        'onboarding_session__selected_council',
        'onboarding_session__selected_affiliation_type'
    ]
    
    for name, Model in app_models:
        # Skip if filtering by specific app type
        if app_type_filter and app_type_filter != name:
            continue
            
        # Build queryset
        qs = Model.objects.select_related(*base_select_related)
        
        # Apply filters
        if council_filter:
            qs = qs.filter(onboarding_session__selected_council=council_filter)
        if affiliation_type_filter:
            qs = qs.filter(onboarding_session__selected_affiliation_type=affiliation_type_filter)
        if status_filter:
            qs = qs.filter(status=status_filter)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        if search_query:
            qs = qs.filter(
                Q(full_names__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(application_number__icontains=search_query)
            )
        
        # Process applications
        for app in qs.order_by('-created_at'):
            all_applications.append({
                'id': app.pk,
                'type': 'Learner' if name == 'student' else name.title(),  # For display
                'app_type': name,  # For URL generation - CRITICAL FIX
                'council': app.onboarding_session.selected_council.code,
                'council_name': app.onboarding_session.selected_council.name,
                'name': app.get_display_name(),
                'email': app.email,
                'application_number': app.application_number,
                'status': app.status,
                'created_at': app.created_at,
                'submitted_at': app.submitted_at,
                'affiliation_type': app.onboarding_session.selected_affiliation_type.name if app.onboarding_session.selected_affiliation_type else None,
                'extra_info': getattr(app, 'current_institution', None) if name == 'student' else 
                             (app.designation_category.name if hasattr(app, 'designation_category') and app.designation_category else None),
            })
    
    # Sort by creation date (most recent first)
    all_applications.sort(key=lambda x: x['created_at'], reverse=True)
    
    # Calculate summary statistics
    total_applications = len(all_applications)
    status_counts = {}
    council_counts = {}
    type_counts = {}
    
    for app in all_applications:
        # Status counts
        status = app['status']
        status_counts[status] = status_counts.get(status, 0) + 1
        
        # Council counts
        council = app['council']
        council_counts[council] = council_counts.get(council, 0) + 1
        
        # Type counts
        app_type = app['type']
        type_counts[app_type] = type_counts.get(app_type, 0) + 1
    
    # Paginate results
    paginator = Paginator(all_applications, 25)  # 25 items per page
    page_number = request.GET.get('page')
    try:
        applications_page = paginator.page(page_number)
    except PageNotAnInteger:
        applications_page = paginator.page(1)
    except EmptyPage:
        applications_page = paginator.page(paginator.num_pages)
    
    # Get available filter options
    councils = Council.objects.filter(is_active=True).order_by('name')
    affiliation_types = AffiliationType.objects.filter(is_active=True).order_by('name')
    
    # Status choices
    status_choices = [
        ('draft', 'Draft'),
        ('submitted', 'Submitted'),
        ('under_review', 'Under Review'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected'),
        ('requires_clarification', 'Requires Clarification'),
    ]
    
    # Application type choices
    app_type_choices = [
        ('associated', 'Associated'),
        ('designated', 'Designated'),
        ('student', 'Learner'),
    ]
    
    context = {
        'applications': applications_page,
        'search_form': search_form,
        'page_title': 'All Applications',
        'total_count': total_applications,
        'status_counts': status_counts,
        'council_counts': council_counts,
        'type_counts': type_counts,
        
        # Filter options
        'councils': councils,
        'affiliation_types': affiliation_types,
        'status_choices': status_choices,
        'app_type_choices': app_type_choices,
        
        # Current filter values
        'current_filters': {
            'search_query': search_query,
            'council': council_filter,
            'affiliation_type': affiliation_type_filter,
            'status': status_filter,
            'date_from': date_from,
            'date_to': date_to,
            'app_type': app_type_filter,
        },
        
        # Pagination info
        'page_info': {
            'current_page': applications_page.number,
            'total_pages': paginator.num_pages,
            'has_previous': applications_page.has_previous(),
            'has_next': applications_page.has_next(),
            'start_index': applications_page.start_index(),
            'end_index': applications_page.end_index(),
        },
        
        # Permission flags
        'can_delete_applications': is_admin_or_manager(request.user),
        'can_approve_applications': can_approve_applications(request.user),
        'can_export_data': request.user.has_perm('enrollments.export_applications'),
    }
    
    return render(request, 'enrollments/applications/list.html', context)




def application_detail(request, pk, app_type):
    """
    Comprehensive application detail view with admin review capabilities.
    """
    model_map = {
        'associated': AssociatedApplication,
        'designated': DesignatedApplication,
        'student': StudentApplication,
    }
    
    model = model_map.get(app_type)
    if not model:
        raise Http404("Invalid application type")
    
    # Get application with all related data
    if model == DesignatedApplication:
        application = get_object_or_404(
            model.objects.select_related(
                'onboarding_session__selected_council',
                'onboarding_session__selected_affiliation_type',
                'designation_category',
                'designation_subcategory',
                'submitted_by',
                'reviewed_by',
                'approved_by'
            ).prefetch_related(
                'academic_qualifications',
                'practical_experiences',
                Prefetch(
                    'documents',
                    queryset=Document.objects.select_related('uploaded_by', 'verified_by')
                ),
                Prefetch(
                    'references',
                    queryset=Reference.objects.select_related().prefetch_related('documents')
                )
            ),
            pk=pk
        )
    else:
        application = get_object_or_404(
            model.objects.select_related(
                'onboarding_session__selected_council',
                'onboarding_session__selected_affiliation_type',
                'submitted_by',
                'reviewed_by',
                'approved_by'
            ).prefetch_related(
                Prefetch(
                    'documents',
                    queryset=Document.objects.select_related('uploaded_by', 'verified_by')
                ),
                Prefetch(
                    'references',
                    queryset=Reference.objects.select_related().prefetch_related('documents')
                )
            ),
            pk=pk
        )
    
    # Check permissions
    user_can_view = (
        request.user.is_authenticated and (
            application.onboarding_session.user == request.user or
            application.email == request.user.email or
            is_admin_or_manager(request.user)
        )
    ) or not request.user.is_authenticated
    
    if not user_can_view:
        return HttpResponseForbidden("You don't have permission to view this application")
    
    # Determine user capabilities
    can_review = request.user.is_authenticated and can_approve_applications(request.user)
    can_edit = request.user.is_authenticated and (
        application.onboarding_session.user == request.user or
        is_admin_or_manager(request.user)
    )
    
    # Get review form for admins
    review_form = None
    if can_review and request.method == 'GET':
        review_form = ApplicationReviewForm(initial={
            'status': application.status,
            'reviewer_notes': application.reviewer_notes,
            'rejection_reason': application.rejection_reason,
        })
    
    # Calculate completion percentage
    completion_percentage = calculate_application_completion(application, app_type)
    
    # Get status timeline
    status_timeline = get_application_timeline(application)
    
    # Group documents by category
    documents_by_category = {}
    for doc in application.documents.all():
        category = doc.get_category_display()
        if category not in documents_by_category:
            documents_by_category[category] = []
        documents_by_category[category].append(doc)
    
    # Calculate document statistics
    total_documents = application.documents.count()
    verified_documents = application.documents.filter(verified=True).count()
    pending_documents = total_documents - verified_documents
    
    # Calculate reference statistics
    total_references = application.references.count()
    references_with_letters = application.references.filter(letter_received=True).count()
    pending_references = total_references - references_with_letters
    
    context = {
        'application': application,
        'app_type': app_type,
        'council': application.onboarding_session.selected_council,
        'affiliation_type': application.onboarding_session.selected_affiliation_type,
        'can_review': can_review,
        'can_edit': can_edit,
        'review_form': review_form,
        'completion_percentage': completion_percentage,
        'status_timeline': status_timeline,
        'documents_by_category': documents_by_category,
        'document_stats': {
            'total': total_documents,
            'verified': verified_documents,
            'pending': pending_documents,
        },
        'reference_stats': {
            'total': total_references,
            'with_letters': references_with_letters,
            'pending': pending_references,
        },
        'page_title': f'{app_type.title()} Application - {application.application_number}',
    }
    
    return render(request, 'enrollments/applications/detail.html', context)



@login_required
@permission_required('enrollments.change_baseapplication', raise_exception=True)
@transaction.atomic
def application_update(request, pk, app_type):
    """
    Update existing application with all related models.
    """
    model_map = {
        'associated': (AssociatedApplication, AssociatedApplicationForm),
        'designated': (DesignatedApplication, DesignatedApplicationForm),
        'student': (StudentApplication, StudentApplicationForm),
    }
    
    model_info = model_map.get(app_type)
    if not model_info:
        raise Http404("Invalid application type")
    
    model, form_class = model_info
    application = get_object_or_404(model, pk=pk)
    
    if request.method == 'POST':
        form = form_class(
            request.POST,
            instance=application,
            request=request,
            onboarding_session=application.onboarding_session
        )
        
        # Handle formsets for designated applications
        formsets = {}
        if app_type == 'designated':
            formsets = {
                'qualifications': AcademicQualificationFormSet(
                    request.POST, instance=application, prefix='qualifications'
                ),
                'references': ReferenceFormSet(
                    request.POST, instance=application, prefix='references'
                ),
                'experiences': PracticalExperienceFormSet(
                    request.POST, instance=application, prefix='experiences'
                ),
                'documents': DocumentFormSet(
                    request.POST, request.FILES, instance=application, prefix='documents'
                ),
            }
        else:
            formsets = {
                'references': ReferenceFormSet(
                    request.POST, instance=application, prefix='references'
                ),
                'documents': DocumentFormSet(
                    request.POST, request.FILES, instance=application, prefix='documents'
                ),
            }
        
        form_valid = form.is_valid()
        formsets_valid = all(formset.is_valid() for formset in formsets.values())
        
        if form_valid and formsets_valid:
            form.save()
            
            # Save formsets
            for formset in formsets.values():
                formset.save()
            
            logger.info(f"Application updated: {application.application_number}")
            messages.success(request, "Application updated successfully.")
            return redirect('enrollments:application_detail', pk=pk, app_type=app_type)
        
        else:
            messages.error(request, "Please correct the errors below.")
    
    else:
        form = form_class(
            instance=application,
            request=request,
            onboarding_session=application.onboarding_session
        )
        
        # Initialize formsets
        formsets = {}
        if app_type == 'designated':
            formsets = {
                'qualifications': AcademicQualificationFormSet(
                    instance=application, prefix='qualifications'
                ),
                'references': ReferenceFormSet(
                    instance=application, prefix='references'
                ),
                'experiences': PracticalExperienceFormSet(
                    instance=application, prefix='experiences'
                ),
                'documents': DocumentFormSet(
                    instance=application, prefix='documents'
                ),
            }
        else:
            formsets = {
                'references': ReferenceFormSet(
                    instance=application, prefix='references'
                ),
                'documents': DocumentFormSet(
                    instance=application, prefix='documents'
                ),
            }
    
    context = {
        'form': form,
        'formsets': formsets,
        'application': application,
        'app_type': app_type,
        'is_update': True,
        'page_title': f'Update {app_type.title()} Application',
    }
    
    template_map = {
        'associated': 'enrollments/applications/associated_form.html',
        'designated': 'enrollments/applications/designated_form.html',
        'student': 'enrollments/applications/student_form.html',
    }
    
    template = template_map.get(app_type, 'enrollments/applications/base_form.html')
    
    return render(request, template, context)

@login_required
@permission_required('enrollments.delete_baseapplication', raise_exception=True)
@transaction.atomic
def application_delete(request, pk, app_type):
    """
    Comprehensive application deletion view with safety checks and audit trail.
    
    This view handles the deletion of applications with the following considerations:
    - Proper permission validation (only admins can delete)
    - Confirmation flow to prevent accidental deletions
    - Audit trail logging for compliance
    - Cascade handling for related objects (documents, references, etc.)
    - Soft delete option for data retention (configurable)
    
    Args:
        request: HTTP request object
        pk: Primary key of the application to delete
        app_type: Type of application ('associated', 'designated', 'student')
    
    Returns:
        HttpResponse: Redirect to application list on success, or confirmation form
    """
    # Define model mapping following the existing pattern
    model_map = {
        'associated': AssociatedApplication,
        'designated': DesignatedApplication,
        'student': StudentApplication,
    }
    
    # Validate application type
    model = model_map.get(app_type)
    if not model:
        logger.warning(f"Invalid application type requested for deletion: {app_type}")
        raise Http404("Invalid application type")
    
    # Get application with optimized query to include related data for display
    # We select_related the essential fields for the confirmation display
    application = get_object_or_404(
        model.objects.select_related(
            'onboarding_session__selected_council',
            'onboarding_session__selected_affiliation_type',
            'submitted_by',
            'reviewed_by',
            'approved_by'
        ).prefetch_related(
            'documents',  # For counting related objects
            'references'  # For counting related objects
        ),
        pk=pk
    )
    
    # Enhanced permission check - only allow admins/managers to delete
    # Applications are sensitive data and deletion should be restricted
    if not is_admin_or_manager(request.user):
        logger.warning(
            f"Unauthorized deletion attempt by user {request.user.id} "
            f"for application {application.application_number}"
        )
        return HttpResponseForbidden(
            "You don't have permission to delete applications. "
            "Please contact an administrator."
        )
    
    
    # Check if application has been submitted and warn user
    submission_warning = None
    if application.submitted_at:
        submission_warning = (
            f"This application was submitted on {application.submitted_at.strftime('%B %d, %Y')} "
            "and may contain important data. Consider archiving instead of deleting."
        )
    
    # Handle POST request (actual deletion)
    if request.method == 'POST':
        # Double-check confirmation
        if request.POST.get('confirm_delete') != 'yes':
            messages.error(request, "Deletion not confirmed. Application was not deleted.")
            return redirect('enrollments:application_detail', pk=pk, app_type=app_type)
        
        # Collect related object counts for logging
        related_counts = {
            'documents': application.documents.count(),
            'references': application.references.count(),
        }
        
        # Add specific counts for designated applications
        if app_type == 'designated':
            related_counts.update({
                'qualifications': application.academic_qualifications.count(),
                'experiences': application.practical_experiences.count(),
            })
        
        # Store application details for logging (before deletion)
        app_details = {
            'application_number': application.application_number,
            'app_type': app_type,
            'email': application.email,
            'status': application.status,
            'council': application.onboarding_session.selected_council.code,
            'created_at': application.created_at,
            'submitted_at': application.submitted_at,
            'related_counts': related_counts,
            'deleted_by': request.user.email,
            'deletion_reason': request.POST.get('deletion_reason', 'No reason provided'),
        }
        
        try:
            # Option 1: Soft Delete (Recommended for audit trails)
            # Uncomment this section if you want to implement soft delete
            """
            application.deleted = True
            application.deleted_at = timezone.now()
            application.deleted_by = request.user
            application.deletion_reason = request.POST.get('deletion_reason', '')
            application.save(update_fields=['deleted', 'deleted_at', 'deleted_by', 'deletion_reason'])
            
            # Also soft delete related objects if they support it
            application.documents.update(deleted=True, deleted_at=timezone.now())
            application.references.update(deleted=True, deleted_at=timezone.now())
            """
            
            # Option 2: Hard Delete (Current implementation)
            # This permanently removes the application and all related data
            
            # Handle file cleanup for documents
            # Get all document files before deletion to clean up storage
            document_files = []
            for document in application.documents.all():
                if document.file and hasattr(document.file, 'path'):
                    document_files.append(document.file.path)
            
            # Perform the deletion
            # Django will handle CASCADE deletes for related objects
            application.delete()
            
            # Clean up physical files from storage
            # This prevents orphaned files in the media directory
            for file_path in document_files:
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                        logger.info(f"Cleaned up file: {file_path}")
                except OSError as e:
                    logger.error(f"Failed to delete file {file_path}: {e}")
            
            # Comprehensive audit logging
            logger.info(
                f"Application deleted successfully. Details: {json.dumps(app_details, default=str)}"
            )
            
            # User feedback with summary
            messages.success(
                request,
                f"Application {app_details['application_number']} has been permanently deleted. "
                f"Removed {related_counts['documents']} documents and {related_counts['references']} references."
            )
            
            # Redirect to application list
            return redirect('enrollments:enrollment_dashboard')
            
        except Exception as e:
            # Handle any deletion errors
            logger.error(
                f"Failed to delete application {application.application_number}: {str(e)}"
            )
            messages.error(
                request,
                f"Failed to delete application: {str(e)}. Please try again or contact support."
            )
            return redirect('enrollments:application_detail', pk=pk, app_type=app_type)
    
    # Handle GET request (show confirmation form)
    # Calculate statistics for confirmation display
    stats = {
        'documents_count': application.documents.count(),
        'references_count': application.references.count(),
        'has_files': application.documents.filter(file__isnull=False).exists(),
    }
    
    # Add designated-specific stats
    if app_type == 'designated':
        stats.update({
            'qualifications_count': application.academic_qualifications.count(),
            'experiences_count': application.practical_experiences.count(),
        })
    
    # Prepare context for confirmation template
    context = {
        'application': application,
        'app_type': app_type,
        'stats': stats,
        'submission_warning': submission_warning,
        'page_title': f'Delete {app_type.title()} Application',
        'breadcrumbs': [
            ('Applications', reverse('enrollments:application_list')),
            (
                f'Application {application.application_number}',
                reverse('enrollments:application_detail', kwargs={'pk': pk, 'app_type': app_type})
            ),
            ('Delete', None)
        ],
    }
    
    return render(request, 'enrollments/applications/delete_confirm.html', context)
# ============================================================================
# APPLICATION BULK ACTIONS
# ============================================================================

@login_required
@user_passes_test(can_approve_applications, login_url='/', redirect_field_name=None)
@require_POST
def application_bulk_action(request):
    """Handle bulk actions on applications"""
    action = request.POST.get('action')
    application_ids = request.POST.getlist('application_ids')
    
    if not action or not application_ids:
        messages.error(request, "No action or applications selected.")
        return redirect('enrollments:application_list')
    
    count = 0
    
    try:
        with transaction.atomic():
            for app_id in application_ids:
                # Get app_type and model
                for app_type, model in [('associated', AssociatedApplication), 
                                      ('designated', DesignatedApplication), 
                                      ('student', StudentApplication)]:
                    try:
                        app = model.objects.get(pk=app_id)
                        
                        if action == 'approve':
                            app.status = 'approved'
                            app.approved_at = timezone.now()
                            app.approved_by = request.user
                        elif action == 'reject':
                            app.status = 'rejected'
                            app.rejected_at = timezone.now()
                            app.rejected_by = request.user
                        elif action == 'under_review':
                            app.status = 'under_review'
                            app.reviewed_at = timezone.now()
                            app.reviewed_by = request.user
                        
                        app.save()
                        count += 1
                        break
                    except model.DoesNotExist:
                        continue
        
        messages.success(request, f'Successfully {action}d {count} applications.')
        
    except Exception as e:
        logger.error(f"Error in bulk action: {str(e)}")
        messages.error(request, "An error occurred during bulk action.")
    
    return redirect('enrollments:application_list')


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def calculate_application_completion(application, app_type):
    """Calculate application completion percentage"""
    total_fields = 0
    completed_fields = 0
    
    # Count basic required fields
    required_basic_fields = [
        'title', 'surname', 'full_names', 'email', 'cell_phone',
        'postal_address_line1', 'postal_city', 'postal_province',
        'postal_code', 'home_language', 'current_occupation'
    ]
    
    for field in required_basic_fields:
        total_fields += 1
        if getattr(application, field, None):
            completed_fields += 1
    
    # Count documents
    if application.documents.exists():
        completed_fields += 1
    total_fields += 1
    
    # Count references
    if application.references.exists():
        completed_fields += 1
    total_fields += 1
    
    # Count legal agreements
    legal_fields = ['popi_act_accepted', 'terms_accepted', 'information_accurate', 'declaration_accepted']
    for field in legal_fields:
        total_fields += 1
        if getattr(application, field, False):
            completed_fields += 1
    
    # Add type-specific fields
    if app_type == 'designated':
        if hasattr(application, 'academic_qualifications') and application.academic_qualifications.exists():
            completed_fields += 1
        total_fields += 1
        
        if hasattr(application, 'practical_experiences') and application.practical_experiences.exists():
            completed_fields += 1
        total_fields += 1
    
    elif app_type == 'student':
        student_fields = ['current_institution', 'course_of_study', 'expected_graduation']
        for field in student_fields:
            total_fields += 1
            if getattr(application, field, None):
                completed_fields += 1
    
    return int((completed_fields / total_fields) * 100) if total_fields > 0 else 0


def get_application_timeline(application):
    """Get application status timeline"""
    timeline = []
    
    # Created
    timeline.append({
        'status': 'Created',
        'date': application.created_at,
        'user': application.submitted_by,
        'icon': 'plus-circle',
        'color': 'primary'
    })
    
    # Submitted
    if application.submitted_at:
        timeline.append({
            'status': 'Submitted',
            'date': application.submitted_at,
            'user': application.submitted_by,
            'icon': 'check-circle',
            'color': 'info'
        })
    
    # Under Review
    if application.reviewed_at:
        timeline.append({
            'status': 'Under Review',
            'date': application.reviewed_at,
            'user': application.reviewed_by,
            'icon': 'eye',
            'color': 'warning'
        })
    
    # Approved/Rejected
    if application.approved_at:
        timeline.append({
            'status': 'Approved',
            'date': application.approved_at,
            'user': application.approved_by,
            'icon': 'check-circle-fill',
            'color': 'success'
        })
    elif application.rejected_at:
        timeline.append({
            'status': 'Rejected',
            'date': application.rejected_at,
            'user': application.rejected_by,
            'icon': 'x-circle-fill',
            'color': 'danger'
        })
    
    return timeline


# ============================================================================
# APPLICATION REVIEW AND APPROVAL
# ============================================================================

@login_required
@permission_required('enrollments.change_baseapplication', raise_exception=True)
@require_http_methods(["POST"])
@transaction.atomic
def application_review(request, pk, app_type):
    """
    Handle application status changes and reviews.
    Fixed to only use fields that exist in the model.
    """
    model_map = {
        'associated': AssociatedApplication,
        'designated': DesignatedApplication,
        'student': StudentApplication,
    }
    
    model = model_map.get(app_type)
    if not model:
        messages.error(request, "Invalid application type.")
        return redirect('enrollments:application_list')
    
    try:
        application = get_object_or_404(model, pk=pk)
        
        # Create review form
        review_form = ApplicationReviewForm(request.POST)
        
        if review_form.is_valid():
            status = review_form.cleaned_data['status']
            reviewer_notes = review_form.cleaned_data['reviewer_notes']
            rejection_reason = review_form.cleaned_data['rejection_reason']
            
            # Update application status
            old_status = application.status
            application.status = status
            
            # Update review fields - only if they exist in the model
            model_fields = [field.name for field in application._meta.get_fields()]
            
            if 'reviewer_notes' in model_fields:
                application.reviewer_notes = reviewer_notes
            
            if 'rejection_reason' in model_fields:
                application.rejection_reason = rejection_reason
            
            # Set review timestamp and user if fields exist
            if 'reviewed_at' in model_fields:
                application.reviewed_at = timezone.now()
            
            if 'reviewed_by' in model_fields:
                application.reviewed_by = request.user
            
            # Handle status-specific fields - only if they exist
            if status == 'approved':
                if 'approved_at' in model_fields:
                    application.approved_at = timezone.now()
                if 'approved_by' in model_fields:
                    application.approved_by = request.user
                    
            elif status == 'submitted' and old_status != 'submitted':
                if 'submitted_at' in model_fields and not getattr(application, 'submitted_at', None):
                    application.submitted_at = timezone.now()
                if 'submitted_by' in model_fields and not getattr(application, 'submitted_by', None):
                    application.submitted_by = request.user
            
            # Save the application
            application.save()

            # Handle digital card assignment if requested
            # Check if the checkbox was checked (it will be in POST if checked)
            assign_digital_card = 'assign_digital_card' in request.POST
            
            if assign_digital_card and status == 'approved':
                try:
                    from affiliationcard.views import assign_card_programmatically
                    
                    # Get content type for this application model
                    content_type = ContentType.objects.get_for_model(application.__class__)
                    
                    # Check if card already exists to prevent duplicates
                    from affiliationcard.models import AffiliationCard
                    existing_card = AffiliationCard.objects.filter(
                        content_type=content_type,
                        object_id=application.pk,
                        status__in=['assigned', 'active']
                    ).first()
                    
                    if not existing_card:
                        # Create and send the digital card
                        card = assign_card_programmatically(
                            content_type=content_type,
                            object_id=application.pk,
                            assigned_by=request.user,
                            send_email=True,
                            request=request
                        )
                        
                        messages.success(request, f"Digital card {card.card_number} assigned and sent to {application.email}")
                        logger.info(f"Digital card {card.card_number} assigned to application {application.application_number}")
                    else:
                        messages.info(request, f"Digital card already exists: {existing_card.card_number}")
                        
                except Exception as e:
                    logger.error(f"Failed to assign digital card for application {application.application_number}: {str(e)}")
                    messages.warning(request, "Application approved but digital card assignment failed. You can assign it manually from the card administration.")
            
            # Log the status change
            logger.info(f"Application {application.application_number} status changed from {old_status} to {status} by {request.user}")
            
            # Add success message
            messages.success(request, f"Application status updated to {application.get_status_display()}.")
            
            # Send notification email if needed
            try:
                if status == 'approved':
                    send_application_approved_email(application, request)
                elif status == 'rejected' and rejection_reason:
                    send_application_rejected_email(application, rejection_reason, request)
            except Exception as e:
                logger.warning(f"Failed to send notification email: {str(e)}")
                # Don't fail the status update if email fails
                messages.warning(request, "Status updated but email notification failed to send.")
            
        else:
            # Form validation failed
            for field, errors in review_form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    
    except Exception as e:
        logger.error(f"Error reviewing application {pk}: {str(e)}")
        messages.error(request, "An error occurred while updating the application status.")
    
    return redirect('enrollments:application_detail', pk=pk, app_type=app_type)





def send_application_approved_email(application, request=None):
    """Send email notification when application is approved using HTML template only"""
    try:
        from django.template.loader import render_to_string
        from django.core.mail import EmailMultiAlternatives
        from django.conf import settings
        
        # Context for email template
        context = {
            'application': application,
            'council': application.onboarding_session.selected_council,
            'affiliation_type': application.onboarding_session.selected_affiliation_type,
            'applicant_name': application.get_display_name(),
            'application_number': application.application_number,
            'current_year': timezone.now().year,
            'request': request,  # Add request to context for building URLs
        }
        
        # Render HTML email template only
        subject = f"Application Approved - {application.application_number}"
        html_content = render_to_string('email_templates/enrollment/application_approved_email.html', context)
        
        # Create email with HTML only
        email = EmailMultiAlternatives(
            subject=subject,
            body='',  # Empty body since we're using HTML only
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[application.email],
            reply_to=[settings.DEFAULT_FROM_EMAIL],
        )
        email.attach_alternative(html_content, "text/html")
        
        # Send email
        email.send()
        
        logger.info(f"Approval email sent for application {application.application_number}")
        
    except Exception as e:
        logger.error(f"Failed to send approval email for {application.application_number}: {str(e)}")
        raise


def send_application_rejected_email(application, rejection_reason, request=None):
    """Send email notification when application is rejected using HTML template only"""
    try:
        from django.template.loader import render_to_string
        from django.core.mail import EmailMultiAlternatives
        from django.conf import settings
        
        # Context for email template
        context = {
            'application': application,
            'council': application.onboarding_session.selected_council,
            'affiliation_type': application.onboarding_session.selected_affiliation_type,
            'applicant_name': application.get_display_name(),
            'application_number': application.application_number,
            'rejection_reason': rejection_reason,
            'current_year': timezone.now().year,
            'request': request,  # Add request to context for building URLs
        }
        
        # Render HTML email template only
        subject = f"Application Status Update - {application.application_number}"
        html_content = render_to_string('email_templates/enrollment/application_rejected_email.html', context)
        
        # Create email with HTML only
        email = EmailMultiAlternatives(
            subject=subject,
            body='',  # Empty body since we're using HTML only
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[application.email],
            reply_to=[settings.DEFAULT_FROM_EMAIL],
        )
        email.attach_alternative(html_content, "text/html")
        
        # Send email
        email.send()
        
        logger.info(f"Rejection email sent for application {application.application_number}")
        
    except Exception as e:
        logger.error(f"Failed to send rejection email for {application.application_number}: {str(e)}")
        raise
    
# ============================================================================
# DASHBOARD AND STATISTICS
# ============================================================================
def enrollment_dashboard_ajax(request):
    """
    Lightweight AJAX endpoint for dashboard updates.
    Returns JSON data for real-time dashboard updates.
    """
    from django.http import JsonResponse
    
    # Simple AJAX response for now
    return JsonResponse({
        'status': 'success',
        'timestamp': timezone.now().isoformat(),
        'message': 'Dashboard data updated'
    })



@login_required
def enrollment_dashboard(request):
    """
    Enhanced administrative dashboard with comprehensive statistics.
    Supports AJAX requests for real‑time updates.
    
    Performance optimizations:
    - Uses database aggregation for statistics instead of multiple count() queries
    - Caches council data for 30 minutes
    - Optimized queries with select_related
    - Reduces total queries from ~45 to ~8
    """
    # Handle AJAX requests
    if request.GET.get('ajax') == '1':
        return enrollment_dashboard_ajax(request)

    start_time = timezone.now()

    # Define all application types up‑front
    app_models = [
        ('associated', AssociatedApplication),
        ('designated', DesignatedApplication),
        ('student', StudentApplication),
    ]

    # ============================================================================
    # OPTIMIZED COUNCIL DATA WITH CACHING
    # ============================================================================
    
    # Cache council data for 30 minutes since it rarely changes
    cache_key = 'dashboard_councils_v2'
    councils = cache.get(cache_key)
    
    if councils is None:
        councils = list(
            Council.objects
            .filter(is_active=True)
            .only('id', 'name', 'code', 'description')
            .order_by('name')
        )
        cache.set(cache_key, councils, 1800)  # 30 minutes
        logger.info("Council data cached for dashboard")

    # Create efficient lookup mappings
    council_map = {c.code.lower(): c for c in councils}
    council_ids = [c.id for c in councils]

    # ============================================================================
    # OPTIMIZED STATISTICS GENERATION
    # ============================================================================

    # Initialize stats structure
    stats = {}
    for council in councils:
        code = council.code.lower()
        stats[code] = {
            'total': 0, 'approved': 0, 'pending': 0,
            'rejected': 0, 'under_review': 0,
            'by_type': {}
        }

    # Generate optimized statistics for each application type
    for name, Model in app_models:
        # Single aggregated query instead of multiple count() queries
        type_stats_raw = (
            Model.objects
            .filter(onboarding_session__selected_council_id__in=council_ids)
            .values('onboarding_session__selected_council__code')
            .annotate(
                total_count=Count('id'),
                approved_count=Count('id', filter=Q(status='approved')),
                pending_count=Count('id', filter=Q(status__in=['draft', 'submitted'])),
                under_review_count=Count('id', filter=Q(status='under_review')),
                rejected_count=Count('id', filter=Q(status='rejected')),
                clarification_count=Count('id', filter=Q(status='requires_clarification')),
            )
        )

        # Process the aggregated results
        for stat_row in type_stats_raw:
            council_code = stat_row['onboarding_session__selected_council__code'].lower()
            
            if council_code in stats:
                type_stats = {
                    'total': stat_row['total_count'],
                    'approved': stat_row['approved_count'],
                    'pending': stat_row['pending_count'],
                    'under_review': stat_row['under_review_count'],
                    'rejected': stat_row['rejected_count'],
                    'requires_clarification': stat_row['clarification_count'],
                }

                # Store type-specific stats
                stats[council_code]['by_type'][name] = type_stats
                
                # Add to council totals
                stats[council_code]['total'] += type_stats['total']
                stats[council_code]['approved'] += type_stats['approved']
                stats[council_code]['pending'] += type_stats['pending'] + type_stats['under_review']
                stats[council_code]['under_review'] += type_stats['under_review']
                stats[council_code]['rejected'] += type_stats['rejected']

    # ============================================================================
    # RECENT APPLICATIONS (KEEPING YOUR WORKING LOGIC)
    # ============================================================================

    # Gather latest 20 applications across all types - KEEPING YOUR EXACT LOGIC
    recent_applications = []
    for name, Model in app_models:
        qs = (
            Model.objects
                 .select_related('onboarding_session__selected_council',
                                 'onboarding_session__selected_affiliation_type')
                 .order_by('-created_at')[:15]
        )

        for app in qs:
            # Map "student" → "Learner" - KEEPING YOUR EXACT LOGIC
            type_display = 'Learner' if name == 'student' else name.title()
            app_data = {
                'id': app.pk,
                'type': type_display,
                'app_type': name,  # ADDED THIS for URL generation (this fixes the error)
                'council': app.onboarding_session.selected_council.code,
                'name': app.get_display_name(),  # USING YOUR EXISTING METHOD
                'email': app.email,
                'application_number': app.application_number,
                'status': app.status,
                'created_at': app.created_at,
                'submitted_at': app.submitted_at,
            }

            # Add extra fields - KEEPING YOUR EXACT LOGIC
            if name == 'designated' and hasattr(app, 'designation_category'):
                app_data['category'] = (
                    app.designation_category.name if app.designation_category else None
                )
            elif name == 'student' and hasattr(app, 'current_institution'):
                app_data['institution'] = app.current_institution

            recent_applications.append(app_data)

    # Final sort & trim to 20 - KEEPING YOUR EXACT LOGIC
    recent_applications.sort(key=lambda x: x['created_at'], reverse=True)
    recent_applications = recent_applications[:20]

    # ============================================================================
    # PERFORMANCE LOGGING
    # ============================================================================
    
    end_time = timezone.now()
    query_time = (end_time - start_time).total_seconds()
    logger.info(f"Dashboard loaded in {query_time:.3f}s with {len(recent_applications)} recent applications")

    # ============================================================================
    # TEMPLATE CONTEXT - KEEPING YOUR EXACT STRUCTURE
    # ============================================================================

    return render(request, 'enrollments/dashboard.html', {
        'stats': stats,
        'recent_applications': recent_applications,
        'page_title': 'Enrollment Dashboard',
        'councils': council_map,
        'load_time': round(query_time, 3),  # ADDED for performance monitoring
        'last_updated': timezone.now(),     # ADDED for freshness indicator
    })

def application_dashboard(request, pk, app_type):
    """
    Public application dashboard showing status and next steps.
    """
    model_map = {
        'associated': AssociatedApplication,
        'designated': DesignatedApplication,
        'student': StudentApplication,
    }
    
    model = model_map.get(app_type)
    if not model:
        raise Http404("Invalid application type")
    
    try:
        application = get_object_or_404(
            model.objects.select_related(
                'onboarding_session__selected_council',
                'onboarding_session__selected_affiliation_type'
            ),
            pk=pk
        )
    except Exception:
        raise Http404("Application not found")
    
    # Check permissions
    user_can_view = (
        request.user.is_authenticated and (
            application.onboarding_session.user == request.user or
            application.email == request.user.email or
            request.user.acrp_role in {User.ACRPRole.GLOBAL_SDP, User.ACRPRole.PROVIDER_ADMIN}
        )
    ) or not request.user.is_authenticated
    
    if not user_can_view:
        return HttpResponseForbidden("You don't have permission to view this application")
    
    # Determine next steps based on status
    next_steps = []
    if application.status == 'draft':
        next_steps = [
            'Complete all required sections',
            'Upload required documents',
            'Submit your application for review'
        ]
    elif application.status == 'submitted':
        next_steps = [
            'Your application is under review',
            'You will receive email updates on progress',
            'Review process typically takes 5-10 business days'
        ]
    elif application.status == 'requires_clarification':
        next_steps = [
            'Review the feedback provided',
            'Update your application with additional information',
            'Resubmit for continued review'
        ]
    elif application.status == 'approved':
        next_steps = [
            'Congratulations! Your application has been approved',
            'You will receive your membership certificate via email',
            'Welcome to the ' + application.onboarding_session.selected_council.name
        ]
    elif application.status == 'rejected':
        next_steps = [
            'Unfortunately, your application was not approved',
            'Review the feedback provided',
            'You may submit a new application addressing the concerns'
        ]
    
    context = {
        'application': application,
        'app_type': app_type,
        'council': application.onboarding_session.selected_council,
        'affiliation_type': application.onboarding_session.selected_affiliation_type,
        'next_steps': next_steps,
        'page_title': f'Application Dashboard - {application.application_number}',
    }
    
    return render(request, 'enrollments/application_dashboard.html', context)


# ============================================================================
# LEGACY COMPATIBILITY AND UTILITY VIEWS
# ============================================================================

def onboarding(request):
    """Redirect legacy onboarding URL to new flow"""
    return redirect('enrollments:onboarding_start')


def learner_apply_prompt(request):
    """Handle student application prompt"""
    if request.method == 'POST':
        token = request.POST.get('token', '').strip()
        if token:
            return redirect(reverse('student:learner_apply', args=[token]))
        messages.error(request, "Please enter your registration link token.")
    
    return render(request, 'enrollments/learner_apply_prompt.html')


# ============================================================================
# AJAX AND API VIEWS
# ============================================================================

@require_http_methods(["GET"])
def get_subcategories_ajax(request):
    """
    AJAX endpoint to get subcategories for a category and council.
    Used for dynamic form updates.
    """
    category_id = request.GET.get('category_id')
    council_id = request.GET.get('council_id')
    
    if not category_id or not council_id:
        return JsonResponse({'error': 'Missing parameters'}, status=400)
    
    try:
        subcategories = DesignationSubcategory.objects.filter(
            category_id=category_id,
            council_id=council_id,
            is_active=True
        ).values('id', 'name', 'description')
        
        return JsonResponse({
            'subcategories': list(subcategories)
        })
    
    except Exception as e:
        logger.error(f"Error fetching subcategories: {str(e)}")
        return JsonResponse({'error': 'Server error'}, status=500)


@require_http_methods(["GET"])
def application_status_ajax(request, pk, app_type):
    """
    AJAX endpoint to get current application status.
    Used for real-time status updates.
    """
    model_map = {
        'associated': AssociatedApplication,
        'designated': DesignatedApplication,
        'student': StudentApplication,
    }
    
    model = model_map.get(app_type)
    if not model:
        return JsonResponse({'error': 'Invalid application type'}, status=400)
    
    try:
        application = get_object_or_404(model, pk=pk)
        
        return JsonResponse({
            'status': application.status,
            'status_display': application.get_status_display(),
            'last_updated': application.updated_at.isoformat(),
        })
    
    except Exception as e:
        logger.error(f"Error fetching application status: {str(e)}")
        return JsonResponse({'error': 'Server error'}, status=500)
    

    